import json
from utils.permissions import require_module_access
from functools import wraps
from django.shortcuts import redirect as _redirect
from django.contrib import messages as _messages

def _is_manager_bar(user):
    from utils.permissions import _is_manager
    return _is_manager(user)


def require_bar_gestion(view_func):
    """Réservé Manager Général et Manager Cuisine — pas les caissiers."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        from utils.permissions import _is_caissier
        if _is_caissier(request.user):
            _messages.error(request, "Accès refusé — cette section est réservée à la gestion.")
            return _redirect('bar:tpe')
        return view_func(request, *args, **kwargs)
    return wrapper
import json
import logging
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.db.models import Sum, Q
from django.db import transaction
from django.utils import timezone
from django.http import JsonResponse
from django.urls import reverse

logger = logging.getLogger(__name__)
from .models import (
    BonCommandeBar, LigneBonCommandeBar, BoissonBar,
    MouvementStockBar, CategorieBar, UniteVente, Client,
    BonReceptionBar, LigneBonReceptionBar,
    InventaireBar, LigneInventaireBar,
    CasseBar, LigneCasseBar,
    ParametrageShot, VenteShot,
    VenteCave, LigneVenteCave,
)
from .models import FournisseurBar


@require_module_access('bar')
def bar_dashboard(request):
    # Les caissiers vont directement au TPE — pas besoin du dashboard
    from utils.permissions import _is_caissier
    if _is_caissier(request.user):
        return redirect('bar:tpe')
    context = {'page_title': 'Tableau de Bord - Cave'}
    return render(request, 'bar/dashboard.html', context)


@require_module_access('bar')
@require_bar_gestion
def stock_management(request):
    boissons = BoissonBar.objects.filter(statut='actif', est_shot=False)

    total_articles = boissons.count()
    valeur_stock = sum(b.prix_achat * b.quantite_stock for b in boissons)
    stock_bas = boissons.filter(quantite_stock__lte=10, quantite_stock__gt=0).count()
    ruptures = boissons.filter(quantite_stock=0).count()
    commandes_en_cours = BonCommandeBar.objects.filter(statut__in=['brouillon', 'confirme', 'envoye', 'partiel']).count()
    total_unites = boissons.aggregate(total=Sum('quantite_stock'))['total'] or 0

    aucune_rupture = ruptures == 0
    seuils_definis = not boissons.filter(seuil_alerte=0).exists()
    prix_renseignes = not boissons.filter(prix=0).exists()
    categories_definies = not boissons.filter(categorie__isnull=True).exists()

    bonnes_pratiques = [
        {'ok': aucune_rupture, 'titre': 'Aucune rupture de stock', 'detail': 'Tous les articles sont disponibles.' if aucune_rupture else f'{ruptures} article(s) en rupture.'},
        {'ok': seuils_definis, 'titre': 'Seuils minimum définis pour tous les articles', 'detail': "Tous les articles ont un seuil d'alerte configuré." if seuils_definis else "Certains articles n'ont pas de seuil défini."},
        {'ok': prix_renseignes, 'titre': 'Prix de vente renseigné pour tous les articles', 'detail': 'Tous les prix sont renseignés.' if prix_renseignes else "Certains articles n'ont pas de prix."},
        {'ok': categories_definies, 'titre': 'Catégories définies pour tous les articles', 'detail': 'Tous les articles ont une catégorie.' if categories_definies else "Certains articles n'ont pas de catégorie."},
    ]
    score = sum(1 for p in bonnes_pratiques if p['ok'])
    articles_critiques = boissons.filter(Q(quantite_stock=0) | Q(quantite_stock__lte=10)).order_by('quantite_stock')

    # Bons de commande pour l'onglet
    bons = BonCommandeBar.objects.select_related('fournisseur', 'client', 'cree_par').all()
    # Filtres bons de commande
    bc_type = request.GET.get('bc_type', '')
    bc_statut = request.GET.get('bc_statut', '')
    bc_q = request.GET.get('bc_q', '')
    if bc_type:
        bons = bons.filter(type_commande=bc_type)
    if bc_statut:
        bons = bons.filter(statut=bc_statut)
    if bc_q:
        bons = bons.filter(Q(numero__icontains=bc_q) | Q(fournisseur__nom__icontains=bc_q) | Q(client__nom__icontains=bc_q))

    # Stats bons de commande
    bc_total = BonCommandeBar.objects.count()
    bc_brouillons = BonCommandeBar.objects.filter(statut='brouillon').count()
    bc_en_cours = BonCommandeBar.objects.filter(statut__in=['confirme', 'envoye', 'partiel']).count()
    bc_en_retard = sum(1 for b in BonCommandeBar.objects.filter(statut__in=['confirme', 'envoye', 'partiel']) if b.est_en_retard)

    # Mouvements du mois
    debut_mois = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    _raw_mvts = (MouvementStockBar.objects
        .exclude(commentaire__startswith='TPE Cave')
        .select_related('boisson', 'utilisateur')
        .order_by('-date'))[:500]
    _seen_keys = {}
    mouvements = []
    for _mv in _raw_mvts:
        _key = (_mv.commentaire, _mv.boisson_id, _mv.type_mouvement)
        if _key in _seen_keys:
            _seen_keys[_key].quantite += _mv.quantite
        else:
            _seen_keys[_key] = _mv
            mouvements.append(_mv)
    mouvements = mouvements[:100]
    mv_entrees_mois = MouvementStockBar.objects.filter(type_mouvement='entree', date__gte=debut_mois).count()
    mv_sorties_mois = MouvementStockBar.objects.filter(type_mouvement__in=['sortie', 'prelevement'], date__gte=debut_mois).count()
    mv_casses_mois = MouvementStockBar.objects.filter(type_mouvement='casse', date__gte=debut_mois).count()
    mv_total_mois = MouvementStockBar.objects.filter(date__gte=debut_mois).count()

    # Bons de réception
    br_q = request.GET.get('br_q', '')
    br_statut = request.GET.get('br_statut', '')
    receptions = BonReceptionBar.objects.select_related('fournisseur', 'operateur', 'bon_commande').all()
    if br_q:
        receptions = receptions.filter(Q(numero__icontains=br_q) | Q(fournisseur__nom__icontains=br_q) | Q(numero_document_fournisseur__icontains=br_q))
    if br_statut:
        receptions = receptions.filter(statut=br_statut)

    br_total = BonReceptionBar.objects.count()
    br_valides = BonReceptionBar.objects.filter(statut='valide').count()
    br_en_cours = BonReceptionBar.objects.filter(statut__in=['brouillon', 'en_cours']).count()
    br_avec_ecarts = sum(1 for br in BonReceptionBar.objects.filter(statut='valide') if br.a_des_ecarts)

    context = {
        'page_title': 'Stock',
        'total_articles': total_articles,
        'valeur_stock': valeur_stock,
        'stock_bas': stock_bas,
        'ruptures': ruptures,
        'commandes_en_cours': commandes_en_cours,
        'total_unites': total_unites,
        'bonnes_pratiques': bonnes_pratiques,
        'score': score,
        'score_total': len(bonnes_pratiques),
        'articles_critiques': articles_critiques,
        'boissons': BoissonBar.objects.exclude(statut='supprime'),
        'categories': CategorieBar.objects.all(),
        # Bons de commande
        'bons': bons,
        'bc_total': bc_total,
        'bc_brouillons': bc_brouillons,
        'bc_en_cours': bc_en_cours,
        'bc_en_retard': bc_en_retard,
        'fournisseurs': FournisseurBar.objects.filter(actif=True).order_by('nom'),
        'clients': Client.objects.filter(actif=True),
        'bc_type': bc_type,
        'bc_statut': bc_statut,
        'bc_q': bc_q,
        # Bons de réception
        'receptions': receptions,
        'br_total': br_total,
        'br_valides': br_valides,
        'br_en_cours': br_en_cours,
        'br_avec_ecarts': br_avec_ecarts,
        'br_q': br_q,
        'br_statut': br_statut,
        # Casses
        'casses': CasseBar.objects.select_related('declare_par').all()[:50],
        'css_total': CasseBar.objects.count(),
        'css_mois': CasseBar.objects.filter(date_casse__gte=timezone.now().replace(day=1).date()).count(),
        'css_valeur_mois': sum(c.total_valeur for c in CasseBar.objects.filter(date_casse__gte=timezone.now().replace(day=1).date(), statut='valide')),
        'css_en_attente': CasseBar.objects.filter(statut='declare').count(),
        # Inventaires
        'inventaires': InventaireBar.objects.select_related('cree_par').all()[:20],
        'inv_total': InventaireBar.objects.count(),
        'inv_en_cours': InventaireBar.objects.filter(statut='brouillon').count(),
        'inv_valides': InventaireBar.objects.filter(statut='valide').count(),
        # Mouvements
        'mouvements': mouvements,
        'mv_entrees_mois': mv_entrees_mois,
        'mv_sorties_mois': mv_sorties_mois,
        'mv_casses_mois': mv_casses_mois,
        'mv_total_mois': mv_total_mois,
        'is_manager': _is_manager_bar(request.user),
    }
    return render(request, 'bar/stock_management.html', context)


# ===== ARTICLES =====

@require_module_access('bar')
@require_bar_gestion
def articles_list(request):
    articles = BoissonBar.objects.exclude(statut='supprime').filter(est_shot=False)
    categories = CategorieBar.objects.all()
    q = request.GET.get('q', '')
    categorie_id = request.GET.get('categorie', '')
    statut = request.GET.get('statut', '')
    if q:
        articles = articles.filter(Q(nom__icontains=q) | Q(reference__icontains=q))
    if categorie_id:
        articles = articles.filter(categorie_id=categorie_id)
    if statut:
        articles = articles.filter(statut=statut)
    context = {
        'page_title': 'Articles',
        'articles': articles,
        'categories': categories,
        'q': q,
        'categorie_id': categorie_id,
        'statut_filtre': statut,
        'total': articles.count(),
    }
    return render(request, 'bar/articles_list.html', context)


@require_module_access('bar')
@require_bar_gestion
def article_create(request):
    categories = CategorieBar.objects.all()
    unites = UniteVente.objects.all()
    if request.method == 'POST':
        article = BoissonBar(
            nom=request.POST.get('nom'),
            reference=request.POST.get('reference') or None,
            categorie_id=request.POST.get('categorie'),
            description=request.POST.get('description', ''),
            mode_prix=request.POST.get('mode_prix', 'manuel'),
            prix_achat=request.POST.get('prix_achat', 0) or 0,
            marge=request.POST.get('marge', 0) or 0,
            prix=request.POST.get('prix_calcule', 0) or 0
                if request.POST.get('mode_prix') == 'marge'
                else request.POST.get('prix_manuel', 0) or 0,
            unite_standard=request.POST.get('unite_standard', 'bouteille'),
            unite_personnalisee_id=request.POST.get('unite_personnalisee') or None,
            seuil_alerte=request.POST.get('seuil_alerte', 10) or 10,
            disponible=request.POST.get('disponible') == 'on',
            statut=request.POST.get('statut', 'actif'),
            est_compose=request.POST.get('est_compose') == 'on',
        )
        if request.FILES.get('image'):
            article.image = request.FILES.get('image')
        article.save()
        messages.success(request, f"Article '{article.nom}' créé avec succès.")
        return redirect(reverse('bar:stock_management') + '?tab=articles')
    context = {'page_title': 'Nouvel Article', 'categories': categories, 'unites': unites, 'mode': 'create'}
    return render(request, 'bar/article_form.html', context)


@require_module_access('bar')
@require_bar_gestion
def article_edit(request, pk):
    article = get_object_or_404(BoissonBar, pk=pk)
    categories = CategorieBar.objects.all()
    unites = UniteVente.objects.all()
    if request.method == 'POST':
        article.nom = request.POST.get('nom') or article.nom
        article.reference = request.POST.get('reference') or None
        article.categorie_id = request.POST.get('categorie') or article.categorie_id
        article.description = request.POST.get('description', '')
        article.mode_prix = request.POST.get('mode_prix') or article.mode_prix
        article.prix_achat = request.POST.get('prix_achat') or article.prix_achat
        article.marge = request.POST.get('marge') or article.marge
        mode = request.POST.get('mode_prix') or article.mode_prix
        if mode == 'marge':
            article.prix = request.POST.get('prix_calcule') or article.prix
        else:
            article.prix = request.POST.get('prix_manuel') or article.prix
        article.unite_standard = request.POST.get('unite_standard') or article.unite_standard
        article.unite_personnalisee_id = request.POST.get('unite_personnalisee') or None
        article.seuil_alerte = request.POST.get('seuil_alerte') or article.seuil_alerte
        article.disponible = request.POST.get('disponible') == 'on'
        article.statut = request.POST.get('statut') or article.statut
        article.est_compose = request.POST.get('est_compose') == 'on'
        if request.FILES.get('image'):
            article.image = request.FILES.get('image')
        article.save()
        messages.success(request, f"Article '{article.nom}' modifié avec succès.")
        return redirect(reverse('bar:stock_management') + '?tab=articles')
    context = {'page_title': 'Modifier Article', 'article': article, 'categories': categories, 'unites': unites, 'mode': 'edit'}
    return render(request, 'bar/article_form.html', context)


@require_module_access('bar')
def article_delete(request, pk):
    from utils.permissions import _is_manager
    if not _is_manager(request.user):
        messages.error(request, "Accès refusé — suppression réservée aux managers.")
        return redirect(reverse('bar:stock_management') + '?tab=articles')
    article = get_object_or_404(BoissonBar, pk=pk)
    if request.method == 'POST':
        nom = article.nom
        article.statut = 'supprime'
        article.save()
        messages.success(request, f"Article '{nom}' supprimé.")
        return redirect(reverse('bar:stock_management') + '?tab=articles')
    return render(request, 'bar/article_confirm_delete.html', {'article': article})


@require_module_access('bar')
@require_bar_gestion
def article_dupliquer(request, pk):
    article = get_object_or_404(BoissonBar, pk=pk)
    article.pk = None
    article.reference = None
    article.nom = f"{article.nom} (copie)"
    article.save()
    messages.success(request, "Article dupliqué avec succès.")
    return redirect('bar:article_edit', pk=article.pk)


@require_module_access('bar')
@require_bar_gestion
def article_sommeil(request, pk):
    article = get_object_or_404(BoissonBar, pk=pk)
    if article.statut == 'actif':
        article.statut = 'sommeil'
        msg = f"Article '{article.nom}' mis en sommeil."
    else:
        article.statut = 'actif'
        msg = f"Article '{article.nom}' réactivé."
    article.save()
    messages.success(request, msg)
    return redirect(reverse('bar:stock_management') + '?tab=articles')


# ===== BONS DE COMMANDE =====

@require_module_access('bar')
@require_bar_gestion
def bon_commande_list(request):
    return redirect(reverse('bar:stock_management') + '?tab=commandes')


@require_module_access('bar')
@require_bar_gestion
def bon_commande_create(request):
    type_commande = request.GET.get('type', 'achat')
    fournisseurs = FournisseurBar.objects.filter(actif=True).order_by('nom')
    clients = Client.objects.filter(actif=True)
    articles = BoissonBar.objects.exclude(statut='supprime')

    if request.method == 'POST':
        type_cmd = request.POST.get('type_commande', 'achat')
        bon = BonCommandeBar(
            type_commande=type_cmd,
            statut=request.POST.get('statut', 'brouillon'),
            fournisseur_id=request.POST.get('fournisseur') or None,
            client_id=request.POST.get('client') or None,
            date_commande=request.POST.get('date_commande') or timezone.now().date(),
            date_livraison_prevue=request.POST.get('date_livraison_prevue') or None,
            notes=request.POST.get('notes', ''),
            cree_par=request.user,
        )
        bon.save()

        # Lignes
        article_ids = request.POST.getlist('article_id[]')
        quantites = request.POST.getlist('quantite[]')
        prix_list = request.POST.getlist('prix_unitaire[]')
        notes_list = request.POST.getlist('notes_ligne[]')

        for i, art_id in enumerate(article_ids):
            if art_id and quantites[i]:
                LigneBonCommandeBar.objects.create(
                    bon=bon,
                    article_id=art_id,
                    quantite_commandee=quantites[i],
                    prix_unitaire=prix_list[i] if prix_list[i] else 0,
                    notes_ligne=notes_list[i] if i < len(notes_list) else '',
                )

        messages.success(request, f"Bon de commande {bon.numero} créé avec succès.")
        return redirect(f'/bar/stock/?tab=commandes')

    context = {
        'page_title': 'Nouveau Bon de Commande',
        'type_commande': type_commande,
        'fournisseurs': fournisseurs,
        'clients': clients,
        'articles': articles,
        'mode': 'create',
    }
    return render(request, 'bar/bon_commande_form.html', context)


@require_module_access('bar')
@require_bar_gestion
def bon_commande_detail(request, pk):
    bon = get_object_or_404(BonCommandeBar, pk=pk)
    lignes = bon.lignes.select_related('article').all()
    context = {
        'page_title': f'Bon {bon.numero}',
        'bon': bon,
        'lignes': lignes,
    }
    return render(request, 'bar/bon_commande_detail.html', context)


@require_module_access('bar')
@require_bar_gestion
def bon_commande_edit(request, pk):
    bon = get_object_or_404(BonCommandeBar, pk=pk)
    fournisseurs = FournisseurBar.objects.filter(actif=True).order_by('nom')
    clients = Client.objects.filter(actif=True)
    articles = BoissonBar.objects.exclude(statut='supprime')

    if request.method == 'POST':
        bon.statut = request.POST.get('statut', bon.statut)
        bon.fournisseur_id = request.POST.get('fournisseur') or None
        bon.client_id = request.POST.get('client') or None
        bon.date_commande = request.POST.get('date_commande') or bon.date_commande
        bon.date_livraison_prevue = request.POST.get('date_livraison_prevue') or None
        bon.notes = request.POST.get('notes', '')
        bon.save()

        # Mise à jour lignes
        bon.lignes.all().delete()
        article_ids = request.POST.getlist('article_id[]')
        quantites = request.POST.getlist('quantite[]')
        prix_list = request.POST.getlist('prix_unitaire[]')
        notes_list = request.POST.getlist('notes_ligne[]')

        for i, art_id in enumerate(article_ids):
            if art_id and quantites[i]:
                LigneBonCommandeBar.objects.create(
                    bon=bon,
                    article_id=art_id,
                    quantite_commandee=quantites[i],
                    prix_unitaire=prix_list[i] if prix_list[i] else 0,
                    notes_ligne=notes_list[i] if i < len(notes_list) else '',
                )

        messages.success(request, f"Bon {bon.numero} modifié avec succès.")
        return redirect(f'/bar/stock/?tab=commandes')

    context = {
        'page_title': f'Modifier {bon.numero}',
        'bon': bon,
        'lignes': bon.lignes.all(),
        'fournisseurs': fournisseurs,
        'clients': clients,
        'articles': articles,
        'mode': 'edit',
    }
    return render(request, 'bar/bon_commande_form.html', context)


@require_module_access('bar')
def fournisseur_list(request):
    from .models import FournisseurBar
    fournisseurs = FournisseurBar.objects.filter(actif=True).order_by('nom')
    q = request.GET.get('q', '')
    if q:
        fournisseurs = fournisseurs.filter(
            models.Q(nom__icontains=q) | models.Q(personne_contact__icontains=q) | models.Q(ville__icontains=q)
        )
    return render(request, 'bar/fournisseur_list.html', {
        'fournisseurs': fournisseurs,
        'q': q,
        'nb_grossistes': FournisseurBar.objects.filter(actif=True, type_fournisseur='grossiste').count(),
        'nb_producteurs': FournisseurBar.objects.filter(actif=True, type_fournisseur='producteur').count(),
        'nb_commandes_actives': BonCommandeBar.objects.filter(
            statut__in=['brouillon', 'confirme', 'envoye', 'partiel']
        ).count(),
        'active_nav': 'fournisseurs',
    })


@require_module_access('bar')
def fournisseur_create(request):
    from .models import FournisseurBar
    if request.method == 'POST':
        f = FournisseurBar(
            nom=request.POST.get('nom', '').strip(),
            type_fournisseur=request.POST.get('type_fournisseur', 'grossiste'),
            personne_contact=request.POST.get('personne_contact', ''),
            telephone=request.POST.get('telephone', ''),
            telephone2=request.POST.get('telephone2', ''),
            email=request.POST.get('email', ''),
            adresse=request.POST.get('adresse', ''),
            ville=request.POST.get('ville', ''),
            notes=request.POST.get('notes', ''),
        )
        f.save()
        messages.success(request, f"Fournisseur « {f.nom} » créé.")
        return redirect('bar:fournisseur_list')
    return render(request, 'bar/fournisseur_form.html', {
        'page_title': 'Nouveau Fournisseur',
        'mode': 'create',
        'active_nav': 'fournisseurs',
    })


@require_module_access('bar')
def fournisseur_edit(request, pk):
    from .models import FournisseurBar
    f = get_object_or_404(FournisseurBar, pk=pk)
    if request.method == 'POST':
        f.nom              = request.POST.get('nom', '').strip()
        f.type_fournisseur = request.POST.get('type_fournisseur', 'grossiste')
        f.personne_contact = request.POST.get('personne_contact', '')
        f.telephone        = request.POST.get('telephone', '')
        f.telephone2       = request.POST.get('telephone2', '')
        f.email            = request.POST.get('email', '')
        f.adresse          = request.POST.get('adresse', '')
        f.ville            = request.POST.get('ville', '')
        f.notes            = request.POST.get('notes', '')
        f.save()
        messages.success(request, f"Fournisseur « {f.nom} » modifié.")
        return redirect('bar:fournisseur_list')
    return render(request, 'bar/fournisseur_form.html', {
        'fournisseur': f,
        'page_title': f'Modifier — {f.nom}',
        'mode': 'edit',
        'active_nav': 'fournisseurs',
    })


@require_module_access('bar')
def fournisseur_delete(request, pk):
    from .models import FournisseurBar
    f = get_object_or_404(FournisseurBar, pk=pk)
    if request.method == 'POST':
        f.actif = False
        f.save()
        messages.success(request, f"Fournisseur « {f.nom} » désactivé.")
        return redirect('bar:fournisseur_list')
    return render(request, 'bar/fournisseur_confirm_delete.html', {'fournisseur': f})


@require_module_access('bar')
@require_POST
def bon_commande_annuler(request, pk):
    bon = get_object_or_404(BonCommandeBar, pk=pk)
    bon.statut = 'annule'
    bon.save()
    messages.warning(request, f"Bon {bon.numero} annulé.")
    return redirect(reverse('bar:stock_management') + '?tab=commandes')


@require_module_access('bar')
def bon_commande_changer_statut(request, pk):
    """AJAX : changer le statut d'un bon"""
    if request.method == 'POST':
        bon = get_object_or_404(BonCommandeBar, pk=pk)
        nouveau_statut = request.POST.get('statut')
        statuts_valides = [s[0] for s in BonCommandeBar.STATUT_CHOICES]
        if nouveau_statut in statuts_valides:
            bon.statut = nouveau_statut
            bon.save()
            return JsonResponse({'success': True, 'statut': bon.statut, 'numero': bon.numero})
    return JsonResponse({'success': False}, status=400)


@require_module_access('bar')
def get_article_prix(request, pk):
    """AJAX : récupérer le prix d'achat d'un article"""
    article = get_object_or_404(BoissonBar, pk=pk)
    return JsonResponse({
        'prix_achat': float(article.prix_achat),
        'prix_vente': float(article.prix),
        'unite': article.unite_affichee,
        'reference': article.reference or '',
    })


# ===== BONS DE RÉCEPTION =====

@require_module_access('bar')
@require_bar_gestion
def bon_reception_list(request):
    return redirect(reverse('bar:stock_management') + '?tab=reception')


@require_module_access('bar')
@require_bar_gestion
def bon_reception_create(request):
    fournisseurs = FournisseurBar.objects.filter(actif=True).order_by('nom')
    articles = BoissonBar.objects.exclude(statut='supprime')
    bons_commande = BonCommandeBar.objects.filter(
        type_commande='achat',
        statut__in=['confirme', 'envoye', 'partiel']
    ).order_by('-date_commande')

    if request.method == 'POST':
        from django.db import transaction
        with transaction.atomic():
            bon_commande_id = request.POST.get('bon_commande') or None
            br = BonReceptionBar(
                bon_commande_id=bon_commande_id,
                fournisseur_id=request.POST.get('fournisseur') or None,
                numero_document_fournisseur=request.POST.get('numero_document_fournisseur', ''),
                operateur=request.user,
                date_reception=request.POST.get('date_reception') or timezone.now().date(),
                statut=request.POST.get('statut', 'brouillon'),
                notes=request.POST.get('notes', ''),
            )
            br.save()

            article_ids = request.POST.getlist('article_id[]')
            qtes_commandees = request.POST.getlist('quantite_commandee[]')
            qtes_recues = request.POST.getlist('quantite_recue[]')
            prix_list = request.POST.getlist('prix_unitaire[]')
            notes_list = request.POST.getlist('notes_ligne[]')

            for i, art_id in enumerate(article_ids):
                if art_id and qtes_recues[i]:
                    LigneBonReceptionBar.objects.create(
                        bon=br,
                        article_id=art_id,
                        quantite_commandee=qtes_commandees[i] if qtes_commandees[i] else 0,
                        quantite_recue=qtes_recues[i],
                        prix_unitaire=prix_list[i] if prix_list[i] else 0,
                        notes_ligne=notes_list[i] if i < len(notes_list) else '',
                    )

            # Si statut = validé → mettre à jour le stock
            if br.statut == 'valide':
                _valider_reception(br, request.user)

        messages.success(request, f"Bon de réception {br.numero} créé avec succès.")
        return redirect(reverse('bar:stock_management') + '?tab=reception')

    # Pré-remplir depuis un bon de commande si fourni
    bc_id = request.GET.get('bc', None)
    bc_prefill = None
    if bc_id:
        bc_prefill = BonCommandeBar.objects.filter(pk=bc_id).first()

    art_data_json = json.dumps({
        str(a.pk): {'nom': a.nom, 'ref': a.reference or '', 'prix': float(a.prix_achat or 0)}
        for a in articles
    }, ensure_ascii=False)

    context = {
        'page_title': 'Nouveau Bon de Réception',
        'fournisseurs': fournisseurs,
        'articles': articles,
        'bons_commande': bons_commande,
        'bc_prefill': bc_prefill,
        'mode': 'create',
        'art_data_json': art_data_json,
    }
    return render(request, 'bar/bon_reception_form.html', context)


@require_module_access('bar')
@require_bar_gestion
def bon_reception_detail(request, pk):
    br = get_object_or_404(BonReceptionBar, pk=pk)
    lignes = br.lignes.select_related('article').all()
    # Aligner les champs sur le template partagé cuisine
    br.cree_par = br.operateur
    for l in lignes:
        l.ingredient = l.article
    context = {
        'page_title': f'Réception {br.numero}',
        'br': br,
        'lignes': lignes,
        'dept_label': 'Cave',
    }
    return render(request, 'cuisine/bon_reception_print.html', context)


@require_module_access('bar')
def bon_reception_valider(request, pk):
    """Valider un bon de réception → mise à jour stock + CMUP"""
    br = get_object_or_404(BonReceptionBar, pk=pk)
    if request.method == 'POST':
        if br.statut in ['brouillon', 'en_cours']:
            _valider_reception(br, request.user)
            messages.success(request, f"Bon {br.numero} validé. Stock mis à jour.")
        else:
            messages.warning(request, "Ce bon ne peut pas être validé dans son état actuel.")
    return redirect(reverse('bar:stock_management') + '?tab=reception')


def _valider_reception(br, user):
    """Fonction interne : valider réception et mettre à jour stock + CMUP"""
    for ligne in br.lignes.all():
        article = ligne.article
        qte = ligne.quantite_recue
        prix = ligne.prix_unitaire

        # Calcul CMUP
        if qte > 0 and prix > 0:
            from decimal import Decimal
            valeur_actuelle = article.quantite_stock * article.prix_achat
            valeur_nouvelle = Decimal(str(qte)) * Decimal(str(prix))
            nouvelle_qte = article.quantite_stock + int(qte)
            if nouvelle_qte > 0:
                article.prix_achat = (valeur_actuelle + valeur_nouvelle) / Decimal(str(nouvelle_qte))

        # Mise à jour stock
        MouvementStockBar.objects.create(
            boisson=article,
            type_mouvement='entree',
            quantite=int(qte),
            commentaire=f"Réception {br.numero} — {br.fournisseur or 'Sans fournisseur'}",
            utilisateur=user,
        )

    # Mettre à jour le statut du bon de commande lié si réception totale
    if br.bon_commande:
        bc = br.bon_commande
        total_recu = sum(
            sum(l.quantite_recue for l in reception.lignes.all())
            for reception in bc.receptions.filter(statut='valide')
        )
        total_commande = sum(l.quantite_commandee for l in bc.lignes.all())
        if total_recu >= total_commande:
            bc.statut = 'recu'
        else:
            bc.statut = 'partiel'
        bc.save()

    br.statut = 'valide'
    br.date_validation = timezone.now()
    br.valide_par = user
    br.save()


@require_module_access('bar')
@require_POST
def bon_reception_annuler(request, pk):
    br = get_object_or_404(BonReceptionBar, pk=pk)
    if br.statut != 'valide':
        br.statut = 'annule'
        br.save()
        messages.warning(request, f"Bon {br.numero} annulé.")
    else:
        messages.error(request, "Impossible d'annuler un bon déjà validé.")
    return redirect(reverse('bar:stock_management') + '?tab=reception')


@require_module_access('bar')
def get_bon_commande_lignes(request, pk):
    """AJAX : récupérer les lignes d'un bon de commande pour pré-remplir la réception"""
    bc = get_object_or_404(BonCommandeBar, pk=pk)
    lignes = []
    for l in bc.lignes.select_related('article').all():
        lignes.append({
            'article_id': l.article_id,
            'article_nom': l.article.nom,
            'article_ref': l.article.reference or '',
            'quantite_commandee': float(l.quantite_commandee),
            'reliquat': float(l.reliquat),
            'prix_unitaire': float(l.prix_unitaire),
        })
    return JsonResponse({
        'lignes': lignes,
        'fournisseur_id': bc.fournisseur_id or '',
        'fournisseur_nom': str(bc.fournisseur) if bc.fournisseur else '',
        'numero': bc.numero,
    })


# ===== MOUVEMENTS ENTRÉES / SORTIES =====

@require_module_access('bar')
def mouvement_create(request):
    """Créer un ou plusieurs mouvements manuels depuis le modal multi-articles"""
    if request.method == 'POST':
        boisson_ids = request.POST.getlist('boisson_id[]')
        type_mv = request.POST.get('type_mouvement')
        quantites = request.POST.getlist('quantite[]')
        commentaire = request.POST.get('commentaire', '')

        TYPE_LABELS = {
            'entree': 'Entrée manuelle',
            'sortie': 'Sortie manuelle',
            'casse': 'Casse / Perte',
            'prelevement': 'Prélèvement',
        }

        nb = 0
        for b_id, qte in zip(boisson_ids, quantites):
            if not b_id or not qte:
                continue
            try:
                boisson = BoissonBar.objects.get(pk=b_id)
                MouvementStockBar.objects.create(
                    boisson=boisson,
                    type_mouvement=type_mv,
                    quantite=int(qte),
                    commentaire=commentaire or TYPE_LABELS.get(type_mv, type_mv),
                    utilisateur=request.user,
                )
                nb += 1
            except Exception as e:
                messages.error(request, f"Erreur ({b_id}) : {str(e)}")
        if nb:
            messages.success(request, f"{nb} mouvement(s) enregistré(s).")

    return redirect(reverse('bar:stock_management') + '?tab=mouvements')


# ===== INVENTAIRE =====

@require_module_access('bar')
@require_bar_gestion
def inventaire_list(request):
    return redirect(reverse('bar:stock_management') + '?tab=inventaire')


@require_module_access('bar')
@require_bar_gestion
def inventaire_create(request):
    """Créer un nouvel inventaire avec toutes les boissons actives"""
    articles = BoissonBar.objects.exclude(statut='supprime').select_related('categorie')

    if request.method == 'POST':
        statut_post = request.POST.get('statut', 'brouillon')
        inv = InventaireBar(
            statut='valide' if statut_post == 'valide' else 'brouillon',
            notes=request.POST.get('notes', ''),
            cree_par=request.user,
        )
        inv.save()

        article_ids = request.POST.getlist('article_id[]')
        qtes_comptees = request.POST.getlist('quantite_comptee[]')
        qtes_theoriques = request.POST.getlist('quantite_theorique[]')
        notes_list = request.POST.getlist('notes_ligne[]')

        for i, art_id in enumerate(article_ids):
            if art_id:
                LigneInventaireBar.objects.create(
                    inventaire=inv,
                    article_id=art_id,
                    quantite_theorique=qtes_theoriques[i] if qtes_theoriques[i] else 0,
                    quantite_comptee=qtes_comptees[i] if qtes_comptees[i] else 0,
                    notes_ligne=notes_list[i] if i < len(notes_list) else '',
                )

        if inv.statut == 'valide':
            _valider_inventaire(inv, request.user)

        messages.success(request, f"Inventaire {inv.numero} créé.")
        return redirect(reverse('bar:stock_management') + '?tab=inventaire')

    # Pré-remplir avec tous les articles actifs
    context = {
        'page_title': 'Nouvel Inventaire',
        'articles': articles,
        'mode': 'create',
    }
    return render(request, 'bar/inventaire_form.html', context)


@require_module_access('bar')
@require_bar_gestion
def inventaire_edit(request, pk):
    inv = get_object_or_404(InventaireBar, pk=pk)
    if inv.statut == 'valide':
        messages.error(request, "Un inventaire validé ne peut plus être modifié.")
        return redirect(reverse('bar:inventaire_detail', args=[pk]))

    articles = BoissonBar.objects.exclude(statut='supprime').select_related('categorie')

    if request.method == 'POST':
        statut_post = request.POST.get('statut', 'brouillon')
        inv.statut = 'valide' if statut_post == 'valide' else 'brouillon'
        inv.notes = request.POST.get('notes', inv.notes)
        inv.save()

        article_ids     = request.POST.getlist('article_id[]')
        qtes_comptees   = request.POST.getlist('quantite_comptee[]')
        qtes_theoriques = request.POST.getlist('quantite_theorique[]')
        notes_list      = request.POST.getlist('notes_ligne[]')

        for i, art_id in enumerate(article_ids):
            if art_id:
                LigneInventaireBar.objects.update_or_create(
                    inventaire=inv,
                    article_id=art_id,
                    defaults={
                        'quantite_theorique': qtes_theoriques[i] if qtes_theoriques[i] else 0,
                        'quantite_comptee':   qtes_comptees[i] if qtes_comptees[i] else 0,
                        'notes_ligne':        notes_list[i] if i < len(notes_list) else '',
                    }
                )

        if inv.statut == 'valide':
            _valider_inventaire(inv, request.user)
            messages.success(request, f"Inventaire {inv.numero} validé. Stock ajusté.")
        else:
            messages.success(request, f"Inventaire {inv.numero} mis à jour.")
        return redirect(reverse('bar:stock_management') + '?tab=inventaire')

    import json
    lignes_dict = {l.article_id: l for l in inv.lignes.all()}
    prefill = {
        art_id: {
            'qte':   str(l.quantite_comptee),
            'notes': l.notes_ligne,
        }
        for art_id, l in lignes_dict.items()
    }

    context = {
        'page_title':    f'Continuer {inv.numero}',
        'articles':      articles,
        'mode':          'edit',
        'inv':           inv,
        'prefill_json':  json.dumps(prefill),
    }
    return render(request, 'bar/inventaire_form.html', context)


@require_module_access('bar')
@require_bar_gestion
def inventaire_detail(request, pk):
    inv = get_object_or_404(InventaireBar, pk=pk)
    lignes = inv.lignes.select_related('article').order_by('article__categorie__nom', 'article__nom')
    ecarts = [l for l in lignes if l.ecart_quantite != 0]
    context = {
        'page_title': f'Inventaire {inv.numero}',
        'inv': inv,
        'lignes': lignes,
        'ecarts': ecarts,
        'valeur_ecart_total': sum(abs(l.valeur_ecart) for l in ecarts),
    }
    return render(request, 'bar/inventaire_detail.html', context)


@require_module_access('bar')
@require_POST
def inventaire_valider(request, pk):
    inv = get_object_or_404(InventaireBar, pk=pk)
    if inv.statut == 'brouillon':
        _valider_inventaire(inv, request.user)
        messages.success(request, f"Inventaire {inv.numero} validé. Stock ajusté.")
    return redirect(reverse('bar:stock_management') + '?tab=inventaire')


def _valider_inventaire(inv, user):
    """Valider inventaire → créer mouvements d'ajustement et corriger le stock."""
    for ligne in inv.lignes.all():
        ecart = ligne.ecart_quantite  # positif = excédent, négatif = manquant
        if ecart > 0:
            type_mv = 'inventaire_excedent'
            qte     = int(ecart)
        elif ecart < 0:
            type_mv = 'inventaire_manquant'
            qte     = int(abs(ecart))
        else:
            type_mv = 'inventaire'
            qte     = 0

        MouvementStockBar.objects.create(
            boisson       = ligne.article,
            type_mouvement= type_mv,
            quantite      = qte,
            commentaire   = (
                f"Inventaire {inv.numero} — "
                f"Théorique: {ligne.quantite_theorique} → Compté: {ligne.quantite_comptee}"
                + (f" (écart: {'+' if ecart > 0 else ''}{ecart})" if ecart != 0 else " (conforme)")
            ),
            utilisateur   = user,
        )
        # Forcer le stock à la valeur comptée physiquement
        ligne.article.refresh_from_db()
        ligne.article.quantite_stock = int(ligne.quantite_comptee)
        ligne.article.save()
        # Persister la valeur de l'écart au CMUP (prix_achat) du moment
        ligne.valeur_ecart = ligne.ecart_quantite * ligne.article.prix_achat
        ligne.save()

    inv.statut = 'valide'
    inv.date_validation = timezone.now()
    inv.valide_par = user
    inv.save()


@require_module_access('bar')
@require_POST
def inventaire_annuler(request, pk):
    inv = get_object_or_404(InventaireBar, pk=pk)
    if inv.statut != 'valide':
        inv.statut = 'annule'
        inv.save()
        messages.warning(request, f"Inventaire {inv.numero} annulé.")
    return redirect(reverse('bar:stock_management') + '?tab=inventaire')


# ===== GESTION DES CASSES =====

@require_module_access('bar')
@require_bar_gestion
def casse_list(request):
    return redirect(reverse('bar:stock_management') + '?tab=casses')


@require_module_access('bar')
@require_bar_gestion
def casse_create(request):
    articles = BoissonBar.objects.exclude(statut='supprime').select_related('categorie')

    if request.method == 'POST':
        casse = CasseBar(
            type_casse=request.POST.get('type_casse', 'casse'),
            statut='declare',
            declare_par=request.user,
            date_casse=request.POST.get('date_casse') or timezone.now().date(),
            description=request.POST.get('description', ''),
        )
        casse.save()

        article_ids = request.POST.getlist('article_id[]')
        quantites = request.POST.getlist('quantite[]')
        prix_list = request.POST.getlist('prix_unitaire[]')
        notes_list = request.POST.getlist('notes_ligne[]')

        for i, art_id in enumerate(article_ids):
            if art_id and quantites[i]:
                LigneCasseBar.objects.create(
                    casse=casse,
                    article_id=art_id,
                    quantite=quantites[i],
                    prix_unitaire=prix_list[i] if prix_list[i] else 0,
                    notes_ligne=notes_list[i] if i < len(notes_list) else '',
                )

        # Valider immédiatement si demandé
        if request.POST.get('valider_maintenant') == '1':
            _valider_casse(casse, request.user)
            messages.success(request, f"Casse {casse.numero} déclarée et validée. Stock mis à jour.")
        else:
            messages.success(request, f"Casse {casse.numero} déclarée. En attente de validation.")

        return redirect(reverse('bar:stock_management') + '?tab=casses')

    context = {
        'page_title': 'Déclarer une Casse / Perte',
        'articles': articles,
    }
    return render(request, 'bar/casse_form.html', context)


@require_module_access('bar')
@require_bar_gestion
def casse_detail(request, pk):
    casse = get_object_or_404(CasseBar, pk=pk)
    lignes = casse.lignes.select_related('article').all()
    context = {
        'page_title': f'Casse {casse.numero}',
        'casse': casse,
        'lignes': lignes,
    }
    return render(request, 'bar/casse_detail.html', context)


@require_module_access('bar')
@require_POST
def casse_valider(request, pk):
    casse = get_object_or_404(CasseBar, pk=pk)
    if casse.statut == 'declare':
        _valider_casse(casse, request.user)
        messages.success(request, f"Casse {casse.numero} validée. Stock mis à jour.")
    return redirect(reverse('bar:stock_management') + '?tab=casses')


def _valider_casse(casse, user):
    """Valider une casse → déduire du stock + créer mouvements tracés"""
    for ligne in casse.lignes.all():
        MouvementStockBar.objects.create(
            boisson=ligne.article,
            type_mouvement='casse',
            quantite=int(ligne.quantite),
            commentaire=f"{casse.get_type_casse_display()} — {casse.numero} — {casse.description[:50] if casse.description else ''}",
            utilisateur=user,
        )
    casse.statut = 'valide'
    casse.date_validation = timezone.now()
    casse.valide_par = user
    casse.save()


@require_module_access('bar')
@require_POST
def casse_annuler(request, pk):
    casse = get_object_or_404(CasseBar, pk=pk)
    if casse.statut == 'declare':
        casse.statut = 'annule'
        casse.save()
        messages.warning(request, f"Déclaration {casse.numero} annulée.")
    return redirect(reverse('bar:stock_management') + '?tab=casses')

@require_module_access('bar')
@require_bar_gestion
def rapport_stock_cave(request):
    get = request.GET.copy()
    if 'module' not in get:
        get['module'] = 'cave'
    request.GET = get
    from rapport.views import rapport_stock
    return rapport_stock(request)


def _parse_date(date_str, fallback=None):
    from datetime import datetime
    if fallback is None:
        fallback = timezone.now().date()
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return fallback


def _stock_periode_bar(date_debut_str, date_fin_str, categorie_id, article_id):
    """Calcule le stock Cave sur une période : stock début, mouvements, stock fin."""
    from datetime import datetime, time as dt_time

    today = timezone.now().date()
    date_debut = _parse_date(date_debut_str, today)
    date_fin = _parse_date(date_fin_str, today)
    if date_fin < date_debut:
        date_fin = date_debut

    debut_aware = timezone.make_aware(datetime.combine(date_debut, dt_time(0, 0, 0)))
    fin_aware = timezone.make_aware(datetime.combine(date_fin, dt_time(23, 59, 59)))

    articles = BoissonBar.objects.exclude(statut='supprime').select_related('categorie').order_by('categorie__nom', 'nom')
    if categorie_id:
        articles = articles.filter(categorie_id=categorie_id)
    if article_id:
        articles = articles.filter(pk=article_id)

    resultats = []
    valeur_totale_fin = 0.0
    valeur_totale_debut = 0.0

    for article in articles:
        # Stock à la fin de la période (reconstruit depuis le stock actuel)
        stock_fin = article.quantite_stock
        for mv in MouvementStockBar.objects.filter(boisson=article, date__gt=fin_aware):
            if mv.type_mouvement in ('entree', 'inventaire'):
                stock_fin -= mv.quantite
            else:
                stock_fin += mv.quantite

        # Mouvements durant la période
        mvts = list(MouvementStockBar.objects.filter(boisson=article, date__gte=debut_aware, date__lte=fin_aware))
        nb_entrees = sum(mv.quantite for mv in mvts if mv.type_mouvement == 'entree')
        nb_sorties = sum(mv.quantite for mv in mvts if mv.type_mouvement == 'sortie')
        nb_casses = sum(mv.quantite for mv in mvts if mv.type_mouvement == 'casse')
        nb_inventaires = sum(mv.quantite for mv in mvts if mv.type_mouvement == 'inventaire')

        # Stock au début de la période (avant tout mouvement du date_debut)
        stock_debut = stock_fin - nb_entrees - nb_inventaires + nb_sorties + nb_casses

        valeur_fin = float(stock_fin) * float(article.prix_achat)
        valeur_debut = float(stock_debut) * float(article.prix_achat)
        valeur_totale_fin += valeur_fin
        valeur_totale_debut += valeur_debut

        resultats.append({
            'article': article,
            'stock_debut': stock_debut,
            'nb_entrees': nb_entrees,
            'nb_sorties': nb_sorties,
            'nb_casses': nb_casses,
            'nb_inventaires': nb_inventaires,
            'stock_fin': stock_fin,
            'valeur_fin': valeur_fin,
            'valeur_debut': valeur_debut,
        })

    return date_debut, date_fin, resultats, valeur_totale_fin, valeur_totale_debut


def _mouvements_bar(date_debut_str, date_fin_str, categorie_id, article_id, type_mv):
    """Retourne la liste des mouvements filtrés — Cave."""
    from datetime import datetime
    mvts = MouvementStockBar.objects.select_related('boisson', 'boisson__categorie', 'utilisateur').order_by('date')
    date_debut = date_fin_obj = None
    if date_debut_str:
        try:
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            mvts = mvts.filter(date__date__gte=date_debut)
        except ValueError:
            pass
    if date_fin_str:
        try:
            date_fin_obj = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
            mvts = mvts.filter(date__date__lte=date_fin_obj)
        except ValueError:
            pass
    if categorie_id:
        mvts = mvts.filter(boisson__categorie_id=categorie_id)
    if article_id:
        mvts = mvts.filter(boisson_id=article_id)
    if type_mv:
        mvts = mvts.filter(type_mouvement=type_mv)
    mvts_list = list(mvts)
    return date_debut, date_fin_obj, mvts_list


@require_module_access('bar')
@require_bar_gestion
def etat_stock_print_bar(request):
    qs = BoissonBar.objects.exclude(statut='supprime').select_related('categorie').order_by('categorie__nom', 'nom')
    articles = [{'obj': a, 'valeur': float(a.prix_achat or 0) * float(a.quantite_stock or 0)} for a in qs]
    total_valeur = sum(a['valeur'] for a in articles)
    return render(request, 'bar/etat_stock_print.html', {
        'articles': articles,
        'total_valeur': total_valeur,
    })


@require_module_access('bar')
@require_bar_gestion
def etat_stock_excel_bar(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse as HR

    articles = BoissonBar.objects.exclude(statut='supprime').select_related('categorie').order_by('categorie__nom', 'nom')
    total_valeur = sum(float(a.prix_achat or 0) * float(a.quantite_stock or 0) for a in articles)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "État du Stock"

    header_fill  = PatternFill("solid", fgColor="1a2535")
    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    title_font   = Font(name='Calibri', bold=True, size=14, color='1a2535')
    bold_font    = Font(name='Calibri', bold=True, size=10)
    normal_font  = Font(name='Calibri', size=10)
    ok_fill      = PatternFill("solid", fgColor="dcfce7")
    alerte_fill  = PatternFill("solid", fgColor="fef3c7")
    rupture_fill = PatternFill("solid", fgColor="fee2e2")
    thin   = Side(border_style="thin", color="d4dce8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    right  = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('A1:K1')
    ws['A1'] = "ÉTAT DU STOCK — CAVE BEHANIAN"
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells('A2:K2')
    ws['A2'] = f"Édité le {timezone.now().strftime('%d/%m/%Y à %H:%M')} — {articles.count()} article(s)"
    ws['A2'].font = Font(name='Calibri', size=10, color='7a8b9c', italic=True)
    ws['A2'].alignment = center

    ws.append([])

    headers = ['#', 'Référence', 'Désignation', 'Catégorie', 'Unité', 'Stock actuel', 'Seuil alerte', 'Prix achat (FCFA)', 'Prix vente (FCFA)', 'Valeur stock (FCFA)', 'État']
    ws.append(headers)
    row_h = ws.max_row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_h, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for i, art in enumerate(articles, 1):
        val = float(art.prix_achat or 0) * float(art.quantite_stock or 0)
        if art.quantite_stock <= 0:
            etat = 'Rupture'; fill = rupture_fill
        elif art.seuil_alerte and art.quantite_stock <= art.seuil_alerte:
            etat = 'Alerte'; fill = alerte_fill
        else:
            etat = 'Normal'; fill = ok_fill

        ws.append([
            i, art.reference or '—', art.nom,
            art.categorie.nom if art.categorie else '—',
            art.unite_affichee,
            art.quantite_stock, art.seuil_alerte or 0,
            float(art.prix_achat or 0), float(art.prix or 0), round(val, 0), etat,
        ])
        r = ws.max_row
        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.font = bold_font if col in (3, 10) else normal_font
            cell.border = border
            cell.alignment = right if col in (6, 7, 8, 9, 10) else (center if col in (1, 5, 11) else Alignment(vertical='center'))
            if col == 11:
                cell.fill = fill
                cell.font = Font(name='Calibri', bold=True, size=10,
                    color='15803d' if etat == 'Normal' else ('d97706' if etat == 'Alerte' else 'dc2626'))

    ws.append([])
    r = ws.max_row + 1
    ws.cell(row=r, column=9, value='VALEUR TOTALE').font = Font(bold=True, size=11)
    ws.cell(row=r, column=10, value=round(total_valeur, 0)).font = Font(bold=True, size=11, color='16a34a')
    ws.cell(row=r, column=10).number_format = '#,##0'

    for col, w in enumerate([5, 12, 28, 16, 10, 12, 12, 16, 16, 18, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    response = HR(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fname = f"Stock_Cave_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


@require_module_access('bar')
@require_bar_gestion
def etat_stock_date_bar(request):
    """Page de sélection : état du stock sur une période — Cave"""
    today = timezone.now().date().isoformat()
    date_debut_str = request.GET.get('date_debut', today)
    date_fin_str = request.GET.get('date_fin', today)
    categorie_id = request.GET.get('categorie', '')
    article_id = request.GET.get('article', '')
    date_debut, date_fin, resultats, valeur_totale_fin, valeur_totale_debut = _stock_periode_bar(
        date_debut_str, date_fin_str, categorie_id, article_id)
    context = {
        'page_title': 'État du stock — Cave',
        'date_debut': date_debut,
        'date_fin': date_fin,
        'date_debut_str': date_debut.isoformat(),
        'date_fin_str': date_fin.isoformat(),
        'categories': CategorieBar.objects.all(),
        'boissons_all': BoissonBar.objects.exclude(statut='supprime').select_related('categorie').order_by('categorie__nom', 'nom'),
        'categorie_id': categorie_id,
        'article_id': article_id,
        'resultats': resultats,
        'valeur_totale_fin': valeur_totale_fin,
        'valeur_totale_debut': valeur_totale_debut,
        'nb_articles': len(resultats),
        'is_single_date': date_debut == date_fin,
    }
    return render(request, 'bar/etat_stock_date.html', context)


@require_module_access('bar')
@require_bar_gestion
def etat_stock_date_bar_print(request):
    """Document d'impression : état du stock sur une période — Cave"""
    today = timezone.now().date().isoformat()
    date_debut_str = request.GET.get('date_debut', today)
    date_fin_str = request.GET.get('date_fin', today)
    categorie_id = request.GET.get('categorie', '')
    article_id = request.GET.get('article', '')
    date_debut, date_fin, resultats, valeur_totale_fin, valeur_totale_debut = _stock_periode_bar(
        date_debut_str, date_fin_str, categorie_id, article_id)
    categorie_nom = ''
    if categorie_id:
        try:
            categorie_nom = CategorieBar.objects.get(pk=categorie_id).nom
        except CategorieBar.DoesNotExist:
            pass
    article_nom = ''
    if article_id:
        try:
            article_nom = BoissonBar.objects.get(pk=article_id).nom
        except BoissonBar.DoesNotExist:
            pass
    context = {
        'date_debut': date_debut,
        'date_fin': date_fin,
        'resultats': resultats,
        'valeur_totale_fin': valeur_totale_fin,
        'valeur_totale_debut': valeur_totale_debut,
        'nb_articles': len(resultats),
        'categorie_nom': categorie_nom,
        'article_nom': article_nom,
        'is_single_date': date_debut == date_fin,
    }
    return render(request, 'bar/etat_stock_date_print.html', context)


@require_module_access('bar')
@require_bar_gestion
def etat_stock_date_bar_excel(request):
    """Export Excel : état du stock sur une période — Cave"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    date_debut_str = request.GET.get('date_debut', timezone.now().date().isoformat())
    date_fin_str   = request.GET.get('date_fin',   timezone.now().date().isoformat())
    categorie_id   = request.GET.get('categorie', '')
    article_id     = request.GET.get('article', '')
    date_debut, date_fin, resultats, valeur_totale_fin, valeur_totale_debut = _stock_periode_bar(
        date_debut_str, date_fin_str, categorie_id, article_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Cave"

    hf  = PatternFill("solid", fgColor="1a2535")
    hft = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    tf  = Font(name='Calibri', bold=True, size=14, color='1a2535')
    bf  = Font(name='Calibri', bold=True, size=10)
    nf  = Font(name='Calibri', size=10)
    th  = Side(border_style="thin", color="d4dce8")
    bd  = Border(left=th, right=th, top=th, bottom=th)
    ct  = Alignment(horizontal='center', vertical='center')
    rt  = Alignment(horizontal='right', vertical='center')

    NC = 11
    ws.merge_cells(f'A1:{get_column_letter(NC)}1')
    ws['A1'] = "ÉTAT DU STOCK SUR PÉRIODE — CAVE BEHANIAN"
    ws['A1'].font = tf; ws['A1'].alignment = ct

    ws.merge_cells(f'A2:{get_column_letter(NC)}2')
    prd = (f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
           if date_debut != date_fin else f"Le {date_debut.strftime('%d/%m/%Y')}")
    ws['A2'] = f"{prd} — {len(resultats)} article(s) — Édité le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(name='Calibri', size=10, color='7a8b9c', italic=True)
    ws['A2'].alignment = ct

    ws.append([])
    headers = ['#', 'Désignation', 'Catégorie', 'Unité',
               'Stock début', 'Entrées', 'Sorties', 'Casses', 'Inventaires',
               'Stock fin', 'Val. fin (F)']
    ws.append(headers)
    rh = ws.max_row
    for col in range(1, NC + 1):
        c = ws.cell(row=rh, column=col)
        c.fill = hf; c.font = hft; c.alignment = ct; c.border = bd

    for i, r in enumerate(resultats, 1):
        art = r['article']
        ws.append([
            i, art.nom,
            art.categorie.nom if art.categorie else '—',
            art.unite_affichee,
            float(r['stock_debut']), float(r['nb_entrees']), float(r['nb_sorties']),
            float(r['nb_casses']), float(r['nb_inventaires']),
            float(r['stock_fin']), float(r['valeur_fin']),
        ])
        rw = ws.max_row
        for col in range(1, NC + 1):
            c = ws.cell(row=rw, column=col)
            c.font = bf if col in (2, 10, 11) else nf
            c.border = bd
            c.alignment = rt if col >= 5 else (ct if col == 1 else Alignment(vertical='center'))
        ws.cell(row=rw, column=11).number_format = '#,##0'

    ws.append([])
    tr = ws.max_row + 1
    ws.cell(row=tr, column=10, value='TOTAL').font = Font(name='Calibri', bold=True)
    ws.cell(row=tr, column=11, value=float(valeur_totale_fin)).font = Font(name='Calibri', bold=True, size=11, color='16a34a')
    ws.cell(row=tr, column=11).number_format = '#,##0'

    for col, w in enumerate([5, 28, 16, 9, 11, 9, 9, 9, 11, 11, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    sfx = f"{date_debut.strftime('%Y%m%d')}{'_' + date_fin.strftime('%Y%m%d') if date_debut != date_fin else ''}"
    response['Content-Disposition'] = f'attachment; filename="Stock_Cave_{sfx}.xlsx"'
    wb.save(response)
    return response


@require_module_access('bar')
@require_bar_gestion
def mouvements_print_bar(request):
    """Page de sélection : mouvements de stock — Cave"""
    date_debut_str = request.GET.get('date_debut', '')
    date_fin_str = request.GET.get('date_fin', '')
    categorie_id = request.GET.get('categorie', '')
    article_id = request.GET.get('article', '')
    type_mv = request.GET.get('type_mouvement', '')
    date_debut, date_fin_obj, mvts_list = _mouvements_bar(date_debut_str, date_fin_str, categorie_id, article_id, type_mv)
    nb_entrees = sum(mv.quantite for mv in mvts_list if mv.type_mouvement == 'entree')
    nb_sorties = sum(mv.quantite for mv in mvts_list if mv.type_mouvement in ('sortie', 'casse'))
    nb_inventaires = sum(mv.quantite for mv in mvts_list if mv.type_mouvement == 'inventaire')
    context = {
        'page_title': 'Mouvements de stock — Cave',
        'mouvements': mvts_list,
        'date_debut': date_debut,
        'date_fin': date_fin_obj,
        'categories': CategorieBar.objects.all(),
        'boissons_all': BoissonBar.objects.exclude(statut='supprime').order_by('nom'),
        'categorie_id': categorie_id,
        'article_id': article_id,
        'type_mv': type_mv,
        'nb_entrees': nb_entrees,
        'nb_sorties': nb_sorties,
        'nb_inventaires': nb_inventaires,
        'total_mvts': len(mvts_list),
    }
    return render(request, 'bar/mouvements_print.html', context)


@require_module_access('bar')
@require_bar_gestion
def mouvements_doc_bar(request):
    """Document d'impression : mouvements de stock — Cave"""
    date_debut_str = request.GET.get('date_debut', '')
    date_fin_str = request.GET.get('date_fin', '')
    categorie_id = request.GET.get('categorie', '')
    article_id = request.GET.get('article', '')
    type_mv = request.GET.get('type_mouvement', '')
    date_debut, date_fin_obj, mvts_list = _mouvements_bar(date_debut_str, date_fin_str, categorie_id, article_id, type_mv)
    nb_entrees = sum(mv.quantite for mv in mvts_list if mv.type_mouvement == 'entree')
    nb_sorties = sum(mv.quantite for mv in mvts_list if mv.type_mouvement in ('sortie', 'casse'))
    nb_inventaires = sum(mv.quantite for mv in mvts_list if mv.type_mouvement == 'inventaire')
    categorie_nom = ''
    if categorie_id:
        try:
            categorie_nom = CategorieBar.objects.get(pk=categorie_id).nom
        except CategorieBar.DoesNotExist:
            pass
    article_nom = ''
    if article_id:
        try:
            article_nom = BoissonBar.objects.get(pk=article_id).nom
        except BoissonBar.DoesNotExist:
            pass
    TYPE_LABELS = {'entree': 'Entrées', 'sortie': 'Sorties', 'casse': 'Casses / Pertes', 'inventaire': 'Ajustements'}
    context = {
        'mouvements': mvts_list,
        'date_debut': date_debut,
        'date_fin': date_fin_obj,
        'categorie_nom': categorie_nom,
        'article_nom': article_nom,
        'type_mv_label': TYPE_LABELS.get(type_mv, 'Tous les types'),
        'nb_entrees': nb_entrees,
        'nb_sorties': nb_sorties,
        'nb_inventaires': nb_inventaires,
        'total_mvts': len(mvts_list),
    }
    return render(request, 'bar/mouvements_doc.html', context)


@require_module_access('bar')
@require_bar_gestion
def mouvements_excel_bar(request):
    """Export Excel : mouvements de stock — Cave"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    date_debut_str = request.GET.get('date_debut', '')
    date_fin_str   = request.GET.get('date_fin', '')
    categorie_id   = request.GET.get('categorie', '')
    article_id     = request.GET.get('article', '')
    type_mv        = request.GET.get('type_mouvement', '')
    date_debut, date_fin_obj, mvts_list = _mouvements_bar(
        date_debut_str, date_fin_str, categorie_id, article_id, type_mv)

    TYPE_LABELS = {'entree': 'Entrée', 'sortie': 'Sortie', 'casse': 'Casse / Perte', 'inventaire': 'Inventaire'}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mouvements Cave"

    hf  = PatternFill("solid", fgColor="1a2535")
    hft = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    tf  = Font(name='Calibri', bold=True, size=14, color='1a2535')
    bf  = Font(name='Calibri', bold=True, size=10)
    nf  = Font(name='Calibri', size=10)
    th  = Side(border_style="thin", color="d4dce8")
    bd  = Border(left=th, right=th, top=th, bottom=th)
    ct  = Alignment(horizontal='center', vertical='center')
    rt  = Alignment(horizontal='right', vertical='center')

    NC = 9
    ws.merge_cells(f'A1:{get_column_letter(NC)}1')
    ws['A1'] = "MOUVEMENTS DE STOCK — CAVE BEHANIAN"
    ws['A1'].font = tf; ws['A1'].alignment = ct

    ws.merge_cells(f'A2:{get_column_letter(NC)}2')
    parts = []
    if date_debut: parts.append(f"Du {date_debut.strftime('%d/%m/%Y')}")
    if date_fin_obj: parts.append(f"au {date_fin_obj.strftime('%d/%m/%Y')}")
    ws['A2'] = f"{' '.join(parts) or 'Toutes dates'} — {len(mvts_list)} mouvement(s) — Édité le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(name='Calibri', size=10, color='7a8b9c', italic=True)
    ws['A2'].alignment = ct

    ws.append([])
    headers = ['#', 'Date', 'Heure', 'Désignation', 'Catégorie', 'Type', 'Quantité', 'Utilisateur', 'Commentaire']
    ws.append(headers)
    rh = ws.max_row
    for col in range(1, NC + 1):
        c = ws.cell(row=rh, column=col)
        c.fill = hf; c.font = hft; c.alignment = ct; c.border = bd

    entree_fill = PatternFill("solid", fgColor="f0fdf4")
    sortie_fill = PatternFill("solid", fgColor="fef2f2")
    inv_fill    = PatternFill("solid", fgColor="eff6ff")
    for i, m in enumerate(mvts_list, 1):
        row_fill = (entree_fill if m.type_mouvement == 'entree'
                    else inv_fill if m.type_mouvement == 'inventaire'
                    else sortie_fill)
        ws.append([
            i,
            m.date.strftime('%d/%m/%Y'),
            m.date.strftime('%H:%M'),
            m.boisson.nom if m.boisson else '—',
            m.boisson.categorie.nom if m.boisson and m.boisson.categorie else '—',
            TYPE_LABELS.get(m.type_mouvement, m.type_mouvement),
            float(m.quantite),
            m.utilisateur.get_full_name() or m.utilisateur.username if m.utilisateur else '—',
            m.commentaire or '',
        ])
        rw = ws.max_row
        for col in range(1, NC + 1):
            c = ws.cell(row=rw, column=col)
            c.font = bf if col in (4, 7) else nf
            c.border = bd
            c.fill = row_fill
            c.alignment = rt if col == 7 else (ct if col in (1, 2, 3) else Alignment(vertical='center'))

    ws.append([])
    tr = ws.max_row + 1
    ws.cell(row=tr, column=6, value='TOTAL').font = Font(name='Calibri', bold=True)
    ws.cell(row=tr, column=7, value=len(mvts_list)).font = Font(name='Calibri', bold=True, size=11)

    for col, w in enumerate([5, 12, 8, 28, 16, 16, 10, 18, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    sfx = timezone.now().strftime('%Y%m%d_%H%M')
    response['Content-Disposition'] = f'attachment; filename="Mouvements_Cave_{sfx}.xlsx"'
    wb.save(response)
    return response


# ================================================================
# REMPLACEMENT DE bar_tpe dans bar/views.py
# ================================================================

@require_module_access('bar')
def bar_tpe(request):
    from .models import BoissonBar, CategorieBar
    from django.contrib.auth.models import User, Group
    from dashboard.models import Configuration

    boissons   = BoissonBar.objects.select_related(
        'categorie', 'shot_parent', 'shot_parent__parametrage_shot'
    ).filter(
        statut='actif', disponible=True
    ).order_by('categorie__nom', 'nom')
    categories = CategorieBar.objects.order_by('nom')

    # Groupe exact en base : "Serveuse/Serveur" (id=5)
    try:
        groupe_serveurs = Group.objects.get(name='Serveuse/Serveur')
        serveurs = User.objects.filter(
            groups=groupe_serveurs,
            is_active=True
        ).order_by('first_name', 'last_name', 'username')
    except Group.DoesNotExist:
        serveurs = User.objects.none()

    config = Configuration.load()

    # Chambres occupées pour le mode "Report Chambre"
    from hotel.models import Reservation as HotelReservation
    chambres_occupees = HotelReservation.objects.filter(
        statut='en_cours'
    ).select_related('client', 'chambre').order_by('chambre__numero')

    # Opérateurs Mobile Money actifs
    from parametres.models import OperateurMobileMoney
    operateurs_mobile_money = OperateurMobileMoney.objects.filter(actif=True)

    return render(request, 'bar/index.html', {
        'boissons'                 : boissons,
        'categories'               : categories,
        'serveurs'                 : serveurs,
        'config'                   : config,
        'page_title'               : 'Cave - Vente TPE',
        'chambres_occupees'        : chambres_occupees,
        'operateurs_mobile_money'  : operateurs_mobile_money,
    })


@require_module_access('bar')
def resume_ventes_cave(request):
    """Résumé des ventes Cave (jour ou période) — JSON pour le TPE."""
    from django.utils import timezone
    from datetime import datetime

    aujourd_hui = timezone.now().date()

    def parse_date(param, fallback):
        try:
            return datetime.strptime(request.GET[param], '%Y-%m-%d').date()
        except Exception:
            return fallback

    date_debut = parse_date('date_debut', aujourd_hui)
    date_fin   = parse_date('date_fin',   aujourd_hui)
    if date_fin < date_debut:
        date_fin = date_debut

    try:
        from facturation.models import Ticket as TicketCaisse
        tickets_qs = TicketCaisse.objects.filter(
            module='cave',
            date_creation__date__gte=date_debut,
            date_creation__date__lte=date_fin,
        ).select_related('cree_par').order_by('-date_creation')

        mode_noms = {
            'especes': 'Espèces', 'carte': 'Carte/TPE', 'carte_bancaire': 'Carte/TPE',
            'mobile': 'Mobile Money', 'mobile_money': 'Mobile Money',
            'orange_money': 'Orange Money', 'wave': 'Wave',
            'moov_money': 'Moov Money', 'mtn_money': 'MTN Money',
            'cheque': 'Chèque', 'virement': 'Virement', 'chambre': 'Chambre',
        }

        par_mode    = {}
        par_caissier = {}
        liste_tickets = []
        total_net   = 0

        for tk in tickets_qs:
            # montant_total = vente réelle ; montant_paye = ce que le client donne (peut dépasser en espèces)
            montant = float(tk.montant_total or tk.montant_paye or 0)
            total_net += montant

            m_raw = tk.mode_paiement or 'especes'
            m = mode_noms.get(m_raw, m_raw.replace('_', ' ').capitalize())
            par_mode[m] = par_mode.get(m, 0) + montant

            caissier_nom = tk.cree_par.get_full_name() or tk.cree_par.username if tk.cree_par else ''
            par_caissier.setdefault(caissier_nom or 'Inconnu', {'nb': 0, 'total': 0})
            par_caissier[caissier_nom or 'Inconnu']['nb']    += 1
            par_caissier[caissier_nom or 'Inconnu']['total'] += montant

            # Heure : en mode multi-jour, afficher la date aussi
            if date_debut == date_fin:
                heure_aff = tk.date_creation.strftime('%H:%M')
            else:
                heure_aff = tk.date_creation.strftime('%d/%m %H:%M')

            liste_tickets.append({
                'numero':   tk.numero or '—',
                'heure':    heure_aff,
                'montant':  montant,  # montant_total (vente réelle)
                'mode':     m,
                'caissier': caissier_nom,
                'ref':      getattr(tk, 'reference', '') or '',
            })

        periode = (date_debut.strftime('%d/%m/%Y') if date_debut == date_fin
                   else f"{date_debut.strftime('%d/%m/%Y')} → {date_fin.strftime('%d/%m/%Y')}")

        return JsonResponse({
            'success':      True,
            'date':         periode,
            'nb_tickets':   len(liste_tickets),
            'total_net':    total_net,
            'par_mode':     par_mode,
            'par_caissier': [{'nom': k, **v} for k, v in par_caissier.items()],
            'tickets':      liste_tickets,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_module_access('bar')
@require_POST
def api_ajuster_stock_tpe(request):
    """
    Ajustement de stock en temps réel depuis le TPE Cave (+/- article).
    delta > 0 : sortie (déduction)  — bloque si stock insuffisant
    delta < 0 : entrée (restauration) — toujours autorisé
    Articles shot/tournée : validation uniquement (déduction ml faite à l'enregistrement de la vente).
    """
    try:
        data       = json.loads(request.body)
        boisson_id = int(data['boisson_id'])
        delta      = int(data['delta'])
        session_token = data.get('session_token', '')

        # Articles shot/tournée : stock_disponible calculé en ml, déduction réelle à la vente
        boisson = BoissonBar.objects.select_related('shot_parent__parametrage_shot').get(pk=boisson_id)
        if boisson.est_shot and boisson.shot_parent_id:
            if delta > 0 and boisson.stock_disponible < delta:
                return JsonResponse({
                    'ok': False,
                    'error': f"{boisson.nom} : stock insuffisant ({boisson.stock_disponible} disponible)"
                })
            return JsonResponse({'ok': True, 'stock': boisson.stock_disponible})

        # Articles normaux : ajustement temps réel via MouvementStockBar
        with transaction.atomic():
            boisson = BoissonBar.objects.select_for_update().get(pk=boisson_id)
            if delta > 0:
                if boisson.quantite_stock < delta:
                    return JsonResponse({
                        'ok': False,
                        'error': f"{boisson.nom} : stock insuffisant ({boisson.quantite_stock} disponible)"
                    })
                MouvementStockBar.objects.create(
                    boisson=boisson, type_mouvement='sortie', quantite=delta,
                    commentaire=f'TPE Cave — {session_token}' if session_token else 'TPE Cave — prévisionnel',
                    utilisateur=request.user,
                )
            else:
                MouvementStockBar.objects.create(
                    boisson=boisson, type_mouvement='entree', quantite=abs(delta),
                    commentaire=f'TPE Cave — retrait {session_token}' if session_token else 'TPE Cave — retrait prévisionnel',
                    utilisateur=request.user,
                )
            boisson.refresh_from_db()
        return JsonResponse({'ok': True, 'stock': boisson.quantite_stock})

    except BoissonBar.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Article introuvable'})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@require_module_access('bar')
def _appliquer_deduction_shot(boisson_shot, qty, commentaire, utilisateur, serveur=None):
    """Délègue au modèle BoissonBar.destock() qui gère la logique ml."""
    with transaction.atomic():
        boisson_shot.destock(qty, commentaire, utilisateur, serveur)


@require_POST
def api_vente_create(request):
    """
    API AJAX appelée depuis le TPE Cave.
    POST JSON -> crée Ticket facturation + décrémente stock BoissonBar.
    """
    try:
        data         = json.loads(request.body)
        lignes       = data.get('lignes', [])
        total        = Decimal(str(data.get('total', 0)))
        paiement     = data.get('paiement', 'especes')
        espace       = data.get('espace', '')
        ref          = data.get('ref', '')
        ticket_nom   = data.get('ticket_nom', 'T-???')
        montant_recu    = Decimal(str(data.get('montant_recu', total)))
        montant_especes = Decimal(str(data.get('montant_especes', 0) or 0))
        serveur_nom     = data.get('serveur', '')
        serveur_id   = data.get('serveur_id')
        stock_live   = data.get('stock_live', False)  # True = stock déjà prélevé en temps réel
        session_token = data.get('session_token', '')
        serveur_obj  = None
        if serveur_id:
            from django.contrib.auth.models import User as AuthUser
            try:
                serveur_obj = AuthUser.objects.get(pk=int(serveur_id), is_active=True)
            except (AuthUser.DoesNotExist, ValueError):
                pass

        if not lignes:
            return JsonResponse({'ok': False, 'error': 'Ticket vide'}, status=400)

        # Vérification stock avant toute transaction
        # Pour les articles normaux, ignorée si déjà prélevée en temps réel (stock_live=True)
        # Pour les shots/tournées, toujours vérifiée (api_ajuster_stock_tpe ne décompte pas le ml)
        manques = []
        for l in lignes:
            try:
                b = BoissonBar.objects.select_related('shot_parent__parametrage_shot').get(pk=int(l['id']))
                qty = int(l['qty'])
                if b.est_shot and b.shot_parent_id:
                    # Vérification en ml pour les shots (toujours, même en mode stock_live)
                    try:
                        param = b.shot_parent.parametrage_shot
                        ml_total = Decimal(b.shot_parent.quantite_stock) * param.volume_contenant_ml - param.ml_en_cours
                        ml_requis = Decimal(b.shot_ml or 30) * qty
                        if ml_requis > ml_total:
                            manques.append(f"• {b.nom} : {int(ml_total)} ml disponibles / {int(ml_requis)} ml requis")
                    except ParametrageShot.DoesNotExist:
                        manques.append(f"• {b.nom} : paramétrage shot introuvable")
                elif not stock_live and b.quantite_stock < qty:
                    manques.append(f"• {b.nom} : {b.quantite_stock} en stock / {qty} demandé")
            except BoissonBar.DoesNotExist:
                manques.append(f"• Article introuvable (id={l['id']})")
        if manques:
            return JsonResponse({
                'ok': False,
                'error': "Vente bloquée — stock insuffisant :",
                'details': manques,
            }, status=400)

        # Map mode paiement TPE -> choix Ticket facturation
        operateur_mobile = data.get('operateur_mobile', '')
        def _map_mode(mode):
            if mode == 'mobile':
                op = (operateur_mobile or '').lower()
                if 'wave' in op:    return 'wave'
                if 'orange' in op:  return 'orange_money'
                if 'mtn' in op:     return 'mtn_money'
                if 'moov' in op:    return 'moov_money'
                return 'mobile_money'
            if mode == 'carte':   return 'carte_bancaire'
            if mode == 'chambre': return 'autre'
            return mode
        mode_fact = _map_mode(paiement)

        ESPACE_LABELS = {
            'restaurant'  : 'Restaurant',
            'piscine'     : 'Piscine',
            'evenementiel': 'Espace evenementiel',
            'hotel'       : 'Hotel',
            'comptoir'    : 'Bar comptoir',
            'plage'       : 'Plage',
            'visite'      : 'Visiteur',
        }
        espace_label = ESPACE_LABELS.get(espace, espace or 'Cave')
        rendu        = max(Decimal('0'), montant_recu - total)

        # Construire le contenu texte lisible du ticket
        from django.utils import timezone as tz
        date_str  = tz.now().strftime('%d/%m/%Y %H:%M')
        lignes_txt = "\n".join(
            f"  {l['nom']} x{l['qty']}  {int(Decimal(str(l['prix'])) * int(l['qty'])):,} F"
            for l in lignes
        )
        # Label règlement lisible pour le ticket texte
        _REGLEMENT_LABELS = {'especes': 'Espèces', 'carte': 'Carte / TPE', 'chambre': 'Report Chambre'}
        if paiement == 'mobile' and operateur_mobile:
            reglement_label = operateur_mobile
        else:
            reglement_label = _REGLEMENT_LABELS.get(paiement, 'Mobile Money')

        # Serveur sélectionné ou à défaut le caissier connecté
        serveur_cave = serveur_nom or request.user.get_full_name() or request.user.username
        contenu = (
            f'<span class="ticket-meta" data-serveur="{serveur_cave}"></span>'
            f"COMPLEXE BEHANIAN - Cave\n"
            f"Ticket  : {ticket_nom}\n"
            f"Date    : {date_str}\n"
            f"Espace  : {espace_label}\n"
            f"Ref     : {ref or '-'}\n"
            f"{'='*32}\n"
            f"{lignes_txt}\n"
            f"{'='*32}\n"
            f"TOTAL   : {int(total):,} FCFA\n"
            f"Reglement : {reglement_label}\n"
        )
        if paiement in ('especes', 'mobile') and rendu > 0:
            contenu += f"Recu    : {int(montant_recu):,} F\n"
            contenu += f"Rendu   : {int(rendu):,} F\n"
        if montant_especes > 0 and paiement not in ('especes', 'chambre'):
            montant_mobile = max(Decimal('0'), total - montant_especes)
            contenu += f"Especes : {int(montant_especes):,} F\n"
            contenu += f"Mobile  : {int(montant_mobile):,} F\n"

        # Si sur_chambre=True → lier sans créer de ticket facturation séparé
        sur_chambre = data.get('sur_chambre', False)
        reservation_id = data.get('reservation_id')

        if sur_chambre and reservation_id and paiement == 'chambre':
            try:
                from hotel.models import Reservation as HotelRes, Consommation as HotelConso
                reservation = HotelRes.objects.get(id=reservation_id, statut='en_cours')
                for l in lignes:
                    boisson_obj = None
                    try:
                        boisson_obj = BoissonBar.objects.get(pk=int(l['id']))
                    except Exception:
                        pass
                    HotelConso.objects.create(
                        reservation=reservation,
                        type_service='bar',
                        boisson=boisson_obj,
                        nom=f"[Cave] {l['nom']}",
                        quantite=int(l['qty']),
                        prix_unitaire=Decimal(str(l['prix'])),
                        serveur=serveur_obj,
                    )
                    # Décrémenter stock
                    # Shots/tournées : toujours décompté (ml_en_cours non géré en temps réel)
                    # Articles normaux : ignoré si déjà prélevé en temps réel (stock_live=True)
                    if boisson_obj:
                        qty = int(l['qty'])
                        cmt_ch = f'Cave → Chambre {reservation.chambre.numero}'
                        if boisson_obj.est_shot and boisson_obj.shot_parent_id:
                            _appliquer_deduction_shot(boisson_obj, qty, cmt_ch, request.user, serveur_obj)
                        elif not stock_live:
                            MouvementStockBar.objects.create(
                                boisson=boisson_obj, type_mouvement='sortie',
                                quantite=qty, commentaire=cmt_ch,
                                utilisateur=request.user, serveur=serveur_obj,
                            )
                # Lier mouvements provisoires TPE (sur chambre)
                if session_token and stock_live:
                    try:
                        MouvementStockBar.objects.filter(
                            commentaire__in=[
                                f'TPE Cave — {session_token}',
                                f'TPE Cave — retrait {session_token}',
                            ]
                        ).update(commentaire=f'Cave → Chambre {reservation.chambre.numero}')
                    except Exception:
                        pass
                return JsonResponse({'ok': True, 'ticket_numero': 'CHAMBRE', 'total': float(total), 'erreurs_stock': []})
            except Exception as e:
                return JsonResponse({'ok': False, 'error': str(e)})

        # Enregistrer la vente (traçabilité + restauration du stock si le ticket est supprimé)
        vente_cave = VenteCave.objects.create(
            espace=espace, reference=ref, total=total,
            serveur=serveur_obj, utilisateur=request.user,
        )
        for l in lignes:
            boisson_ligne = None
            try:
                boisson_ligne = BoissonBar.objects.get(pk=int(l['id']))
            except (BoissonBar.DoesNotExist, ValueError, KeyError):
                pass
            LigneVenteCave.objects.create(
                vente=vente_cave, boisson=boisson_ligne,
                nom_article=l.get('nom', ''), quantite=int(l['qty']),
                prix_unitaire=Decimal(str(l['prix'])),
            )

        # Creer le Ticket dans facturation
        from facturation.models import Ticket, generate_ticket_numero
        ticket = Ticket.objects.create(
            numero        = generate_ticket_numero(),
            module        = 'cave',
            objet_id      = vente_cave.id,
            montant_total = total,
            mode_paiement = mode_fact,
            montant_paye  = montant_recu,
            contenu       = contenu,
            cree_par      = request.user,
            imprime       = True,
            date_impression = tz.now(),
        )

        # Si mode chambre : lier les articles à la réservation hôtel
        reservation_id = data.get('reservation_id')
        if paiement == 'chambre' and reservation_id:
            try:
                from hotel.models import Reservation as HotelRes, Consommation as HotelConso
                reservation = HotelRes.objects.get(id=reservation_id, statut='en_cours')
                for l in lignes:
                    boisson_obj = None
                    try:
                        boisson_obj = BoissonBar.objects.get(pk=int(l['id']))
                    except Exception:
                        pass
                    HotelConso.objects.create(
                        reservation=reservation,
                        type_service='bar',
                        boisson=boisson_obj,
                        nom=f"[Cave] {l['nom']}",
                        quantite=int(l['qty']),
                        prix_unitaire=Decimal(str(l['prix'])),
                        serveur=serveur_obj,
                    )
            except Exception as e:
                pass  # Ne pas bloquer la vente si erreur liaison

        # Lier mouvements provisoires TPE au ticket
        if session_token and stock_live:
            _cmt = f'Vente Cave — {ticket.numero}'
            if espace:
                _cmt += f' [{espace_label}]'
            MouvementStockBar.objects.filter(
                commentaire__in=[
                    f'TPE Cave — {session_token}',
                    f'TPE Cave — retrait {session_token}',
                ]
            ).update(commentaire=_cmt)

        # Decrementer stock + tracer mouvements
        # Articles normaux : ignoré si déjà prélevé en temps réel (stock_live=True)
        # Articles shot/tournée : toujours décompté ici (ml_en_cours non géré en temps réel)
        erreurs_stock = []
        for l in lignes:
            try:
                boisson = BoissonBar.objects.select_related('shot_parent__parametrage_shot').get(pk=int(l['id']))
                qty     = int(l['qty'])
                cmt     = f"Vente TPE - {ticket.numero} - {espace_label} {ref}".strip(' -')
                if boisson.est_shot and boisson.shot_parent_id:
                    # Logique ml pour articles shot/tournée (jamais prélevé en temps réel)
                    _appliquer_deduction_shot(boisson, qty, cmt, request.user, serveur_obj)
                elif not stock_live:
                    MouvementStockBar.objects.create(
                        boisson        = boisson,
                        type_mouvement = 'sortie',
                        quantite       = qty,
                        commentaire    = cmt,
                        utilisateur    = request.user,
                        serveur        = serveur_obj,
                    )
            except BoissonBar.DoesNotExist:
                erreurs_stock.append(f"Article id={l['id']} introuvable")

        return JsonResponse({
            'ok'            : True,
            'ticket_id'     : ticket.pk,
            'ticket_numero' : ticket.numero,
            'total'         : int(total),
            'rendu'         : int(rendu),
            'erreurs_stock' : erreurs_stock,
        })

    except Exception as e:
        import traceback
        return JsonResponse({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}, status=500)


# ─────────────────────────────────────────────────────────
#  SHOTS / LIQUEURS AU VERRE
# ─────────────────────────────────────────────────────────

@require_module_access('bar')
@require_bar_gestion
def shot_parametrage(request):
    """Liste des articles configurés pour la vente au shot + formulaire d'ajout/modif."""
    articles_avec_config = BoissonBar.objects.filter(
        statut='actif', disponible=True
    ).select_related('parametrage_shot').order_by('categorie__nom', 'nom')

    if request.method == 'POST':
        boisson_id  = request.POST.get('boisson_id')
        vol_cont    = request.POST.get('volume_contenant_ml', 700)
        vol_shot    = request.POST.get('volume_shot_ml', 30)
        prix_shot   = request.POST.get('prix_shot', 0)
        prix_tournee = request.POST.get('prix_tournee', 0)
        actif        = request.POST.get('actif') == '1'

        boisson = get_object_or_404(BoissonBar, pk=boisson_id)
        param, created = ParametrageShot.objects.update_or_create(
            boisson=boisson,
            defaults={
                'volume_contenant_ml': int(vol_cont),
                'volume_shot_ml':      int(vol_shot),
                'prix_shot':           Decimal(str(prix_shot)),
                'prix_tournee':        Decimal(str(prix_tournee)),
                'actif':               actif,
            }
        )
        action = "créé" if created else "mis à jour"
        messages.success(request, f"Paramétrage shot pour « {boisson.nom} » {action}.")
        return redirect('bar:shot_parametrage')

    context = {
        'articles': articles_avec_config,
        'page_title': 'Paramétrage — Vente au Shot',
    }
    return render(request, 'bar/shot_parametrage.html', context)


@require_module_access('bar')
def shot_vente(request):
    """Page de vente shot + historique du jour."""
    parametrages = ParametrageShot.objects.filter(
        actif=True, boisson__statut='actif', boisson__disponible=True
    ).select_related('boisson').order_by('boisson__nom')

    if request.method == 'POST':
        param_id   = request.POST.get('parametrage_id')
        type_vente = request.POST.get('type_vente', 'shot')

        param = get_object_or_404(ParametrageShot, pk=param_id, actif=True)
        boisson = param.boisson

        nb_shots   = 2 if type_vente == 'tournee' else 1
        ml_a_debiter = Decimal(nb_shots * param.volume_shot_ml)
        prix_encaisse = param.prix_tournee if type_vente == 'tournee' else param.prix_shot

        # Vérification stock
        ml_total_stock = Decimal(boisson.quantite_stock) * param.volume_contenant_ml - param.ml_en_cours
        if ml_a_debiter > ml_total_stock:
            messages.error(request, f"Stock insuffisant — {boisson.nom} : seulement {int(ml_total_stock)} ml restants.")
            return redirect('bar:shot_vente')

        with transaction.atomic():
            # Mise à jour bouteille en cours
            param.ml_en_cours += ml_a_debiter
            bouteilles_consommees = int(param.ml_en_cours // param.volume_contenant_ml)
            param.ml_en_cours = param.ml_en_cours % param.volume_contenant_ml
            param.save()

            # Décrémenter le stock bouteilles si nécessaire
            if bouteilles_consommees > 0:
                from django.db.models import F
                BoissonBar.objects.filter(pk=boisson.pk).update(
                    quantite_stock=F('quantite_stock') - bouteilles_consommees
                )
                MouvementStockBar.objects.create(
                    boisson=boisson,
                    type_mouvement='sortie',
                    quantite=bouteilles_consommees,
                    commentaire=f"Vente shot — {bouteilles_consommees} bouteille(s) épuisée(s)",
                    utilisateur=request.user,
                )

            VenteShot.objects.create(
                boisson=boisson,
                type_vente=type_vente,
                nb_shots=nb_shots,
                ml_debites=ml_a_debiter,
                prix_encaisse=prix_encaisse,
                operateur=request.user,
            )

        label = "Tournée (2 shots)" if type_vente == 'tournee' else "Shot"
        messages.success(request, f"{label} de {boisson.nom} enregistré — {int(ml_a_debiter)} ml débités.")
        return redirect('bar:shot_vente')

    # Historique du jour
    today = timezone.now().date()
    historique = VenteShot.objects.filter(
        date__date=today
    ).select_related('boisson', 'operateur').order_by('-date')

    context = {
        'parametrages': parametrages,
        'historique':   historique,
        'page_title':   'Vente au Shot',
    }
    return render(request, 'bar/shot_vente.html', context)


@require_module_access('bar')
@require_bar_gestion
def shot_historique(request):
    """Historique complet des ventes shot avec filtres date."""
    from django.db.models import Sum as DSum
    date_debut = request.GET.get('date_debut')
    date_fin   = request.GET.get('date_fin')

    qs = VenteShot.objects.select_related('boisson', 'operateur').order_by('-date')
    if date_debut:
        qs = qs.filter(date__date__gte=date_debut)
    if date_fin:
        qs = qs.filter(date__date__lte=date_fin)

    totaux = qs.aggregate(
        total_shots=DSum('nb_shots'),
        total_ml=DSum('ml_debites'),
        total_ca=DSum('prix_encaisse'),
    )

    context = {
        'ventes':      qs[:200],
        'totaux':      totaux,
        'date_debut':  date_debut or '',
        'date_fin':    date_fin or '',
        'page_title':  'Historique Ventes Shot',
    }
    return render(request, 'bar/shot_historique.html', context)

