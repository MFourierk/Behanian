from utils.permissions import require_module_access
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, F
from django.utils import timezone
from django.http import JsonResponse
from decimal import Decimal

from .models import (
    Fournisseur,
    CategorieIngredient,
    UniteIngredient,
    Ingredient,
    MouvementStockCuisine,
    BonCommandeCuisine,
    LigneBonCommandeCuisine,
    BonReceptionCuisine,
    LigneBonReceptionCuisine,
    CategoriePlat,
    FicheTechnique,
    LigneFicheTechnique,
    Plat,
    InventaireCuisine,
    LigneInventaireCuisine,
    CasseCuisine,
    LigneCasseCuisine,
)


# ==============================================================================
# TABLEAU DE BORD CUISINE
# ==============================================================================

@require_module_access('cuisine')
def index(request):
    ingredients = Ingredient.objects.filter(statut=True)

    total_articles  = ingredients.count()
    valeur_stock    = sum(i.valeur_stock for i in ingredients)
    stock_bas       = ingredients.filter(quantite_stock__gt=0, quantite_stock__lte=F('seuil_alerte')).count()
    ruptures        = ingredients.filter(quantite_stock__lte=0).count()
    bc_en_cours     = BonCommandeCuisine.objects.filter(statut__in=['brouillon', 'confirme', 'envoye', 'partiel']).count()
    derniers_mvts   = MouvementStockCuisine.objects.select_related('ingredient', 'utilisateur').order_by('-date')[:10]

    context = {
        'page_title':     'Cuisine — Tableau de bord',
        'total_articles': total_articles,
        'valeur_stock':   valeur_stock,
        'stock_bas':      stock_bas,
        'ruptures':       ruptures,
        'bc_en_cours':    bc_en_cours,
        'derniers_mvts':  derniers_mvts,
    }
    return render(request, 'cuisine/index.html', context)


# ==============================================================================
# STOCK MANAGEMENT (page principale avec onglets)
# ==============================================================================

@require_module_access('cuisine')
def stock_management(request):
    ingredients = Ingredient.objects.select_related('categorie', 'unite_stock').filter(statut=True)

    # Stats
    total_articles = ingredients.count()
    valeur_stock   = sum(i.valeur_stock for i in ingredients)
    stock_bas      = ingredients.filter(quantite_stock__gt=0, quantite_stock__lte=F('seuil_alerte')).count()
    ruptures       = ingredients.filter(quantite_stock__lte=0).count()
    bc_en_cours    = BonCommandeCuisine.objects.filter(statut__in=['brouillon', 'confirme', 'envoye', 'partiel']).count()

    # Bons de commande
    bons = BonCommandeCuisine.objects.select_related('fournisseur', 'cree_par').all()
    bc_statut = request.GET.get('bc_statut', '')
    bc_q      = request.GET.get('bc_q', '')
    if bc_statut:
        bons = bons.filter(statut=bc_statut)
    if bc_q:
        bons = bons.filter(Q(numero__icontains=bc_q) | Q(fournisseur__nom__icontains=bc_q))

    bc_total      = BonCommandeCuisine.objects.count()
    bc_brouillons = BonCommandeCuisine.objects.filter(statut='brouillon').count()
    bc_en_retard  = sum(1 for b in BonCommandeCuisine.objects.filter(statut__in=['confirme', 'envoye', 'partiel']) if b.est_en_retard)

    # Réceptions
    receptions    = BonReceptionCuisine.objects.select_related('fournisseur', 'cree_par').all()
    br_total      = receptions.count()
    br_brouillons = receptions.filter(statut='brouillon').count()

    # Mouvements
    mouvements = MouvementStockCuisine.objects.select_related('ingredient', 'utilisateur').order_by('-date')
    mvt_type = request.GET.get('mvt_type', '')
    if mvt_type:
        mouvements = mouvements.filter(type_mouvement=mvt_type)

    # Inventaires
    inventaires = InventaireCuisine.objects.select_related('cree_par').all()

    # Casses
    casses = CasseCuisine.objects.select_related('cree_par').all()

    context = {
        'page_title':     'Stock Cuisine',
        'ingredients':    ingredients,
        'categories':     CategorieIngredient.objects.all(),
        'fournisseurs':   Fournisseur.objects.filter(actif=True),
        'total_articles': total_articles,
        'valeur_stock':   valeur_stock,
        'stock_bas':      stock_bas,
        'ruptures':       ruptures,
        'bc_en_cours':    bc_en_cours,
        # BC
        'bons':           bons,
        'bc_total':       bc_total,
        'bc_brouillons':  bc_brouillons,
        'bc_en_retard':   bc_en_retard,
        'bc_statut':      bc_statut,
        'bc_q':           bc_q,
        # BR
        'receptions':     receptions,
        'br_total':       br_total,
        'br_brouillons':  br_brouillons,
        # Mouvements
        'mouvements':     mouvements[:50],
        'mvt_type':       mvt_type,
        # Inventaires & Casses
        'inventaires':    inventaires,
        'casses':         casses,
    }
    return render(request, 'cuisine/stock_management.html', context)


# ==============================================================================
# INGRÉDIENTS
# ==============================================================================

@require_module_access('cuisine')
def ingredient_list(request):
    ingredients = Ingredient.objects.select_related('categorie', 'unite_stock').filter(statut=True)
    q = request.GET.get('q', '')
    cat = request.GET.get('categorie', '')
    if q:
        ingredients = ingredients.filter(Q(nom__icontains=q) | Q(reference__icontains=q))
    if cat:
        ingredients = ingredients.filter(categorie_id=cat)
    context = {
        'page_title':  'Ingrédients',
        'ingredients': ingredients,
        'categories':  CategorieIngredient.objects.all(),
        'q': q, 'categorie_id': cat,
    }
    return render(request, 'cuisine/ingredient_list.html', context)


@require_module_access('cuisine')
def ingredient_create(request):
    categories = CategorieIngredient.objects.all()
    unites     = UniteIngredient.objects.all()
    fournisseurs = Fournisseur.objects.filter(actif=True)
    if request.method == 'POST':
        ing = Ingredient(
            nom=request.POST.get('nom'),
            reference=request.POST.get('reference') or None,
            categorie_id=request.POST.get('categorie') or None,
            description=request.POST.get('description', ''),
            unite_stock_id=request.POST.get('unite_stock') or None,
            unite_recette_id=request.POST.get('unite_recette') or None,
            facteur_conversion=request.POST.get('facteur_conversion') or 1,
            prix_achat=request.POST.get('prix_achat') or 0,
            seuil_alerte=request.POST.get('seuil_alerte') or 0,
            stock_max=request.POST.get('stock_max') or 0,
            fournisseur_principal_id=request.POST.get('fournisseur_principal') or None,
        )
        ing.save()
        messages.success(request, f"Ingrédient '{ing.nom}' créé.")
        return redirect('/cuisine/stock/?tab=ingredients')
    context = {
        'page_title':    'Nouvel Ingrédient',
        'categories':    categories,
        'unites':        unites,
        'fournisseurs':  fournisseurs,
        'mode':          'create',
    }
    return render(request, 'cuisine/ingredient_form.html', context)


@require_module_access('cuisine')
def ingredient_edit(request, pk):
    ing        = get_object_or_404(Ingredient, pk=pk)
    categories = CategorieIngredient.objects.all()
    unites     = UniteIngredient.objects.all()
    fournisseurs = Fournisseur.objects.filter(actif=True)
    if request.method == 'POST':
        ing.nom = request.POST.get('nom')
        ing.reference = request.POST.get('reference') or None
        ing.categorie_id = request.POST.get('categorie') or None
        ing.description = request.POST.get('description', '')
        ing.unite_stock_id = request.POST.get('unite_stock') or None
        ing.unite_recette_id = request.POST.get('unite_recette') or None
        ing.facteur_conversion = request.POST.get('facteur_conversion') or 1
        ing.prix_achat = request.POST.get('prix_achat') or 0
        ing.seuil_alerte = request.POST.get('seuil_alerte') or 0
        ing.stock_max = request.POST.get('stock_max') or 0
        ing.fournisseur_principal_id = request.POST.get('fournisseur_principal') or None
        ing.save()
        messages.success(request, f"Ingrédient '{ing.nom}' modifié.")
        return redirect('/cuisine/stock/?tab=ingredients')
    context = {
        'page_title':    'Modifier Ingrédient',
        'ingredient':    ing,
        'categories':    categories,
        'unites':        unites,
        'fournisseurs':  fournisseurs,
        'mode':          'edit',
    }
    return render(request, 'cuisine/ingredient_form.html', context)


@require_module_access('cuisine')
def ingredient_delete(request, pk):
    ing = get_object_or_404(Ingredient, pk=pk)
    if request.method == 'POST':
        ing.statut = False
        ing.save()
        messages.success(request, f"Ingrédient '{ing.nom}' désactivé.")
        return redirect('/cuisine/stock/?tab=ingredients')
    return render(request, 'cuisine/ingredient_confirm_delete.html', {'ingredient': ing})


# ==============================================================================
# MOUVEMENTS DE STOCK
# ==============================================================================

@require_module_access('cuisine')
def mouvement_create(request):
    if request.method == 'POST':
        ing_id  = request.POST.get('ingredient')
        type_m  = request.POST.get('type_mouvement')
        quantite = request.POST.get('quantite')
        commentaire = request.POST.get('commentaire', '')
        if ing_id and type_m and quantite:
            MouvementStockCuisine.objects.create(
                ingredient_id  = ing_id,
                type_mouvement = type_m,
                quantite       = quantite,
                commentaire    = commentaire,
                utilisateur    = request.user,
            )
            messages.success(request, "Mouvement enregistré.")
    return redirect('/cuisine/stock/?tab=mouvements')


# ==============================================================================
# BONS DE COMMANDE CUISINE
# ==============================================================================

@require_module_access('cuisine')
def bon_commande_list(request):
    return redirect('/cuisine/stock/?tab=commandes')


@require_module_access('cuisine')
def bon_commande_create(request):
    fournisseurs = Fournisseur.objects.filter(actif=True)
    ingredients  = Ingredient.objects.filter(statut=True)
    if request.method == 'POST':
        bon = BonCommandeCuisine(
            fournisseur_id=request.POST.get('fournisseur') or None,
            statut=request.POST.get('statut', 'brouillon'),
            date_commande=request.POST.get('date_commande') or timezone.now().date(),
            date_livraison_prevue=request.POST.get('date_livraison_prevue') or None,
            notes=request.POST.get('notes', ''),
            cree_par=request.user,
        )
        bon.save()
        ing_ids  = request.POST.getlist('ingredient_id[]')
        qtes     = request.POST.getlist('quantite[]')
        prix_l   = request.POST.getlist('prix_unitaire[]')
        notes_l  = request.POST.getlist('notes_ligne[]')
        for i, ing_id in enumerate(ing_ids):
            if ing_id and qtes[i]:
                LigneBonCommandeCuisine.objects.create(
                    bon=bon,
                    ingredient_id=ing_id,
                    quantite_commandee=qtes[i],
                    prix_unitaire=prix_l[i] if prix_l[i] else 0,
                    notes_ligne=notes_l[i] if i < len(notes_l) else '',
                )
        messages.success(request, f"Bon {bon.numero} créé.")
        return redirect('/cuisine/stock/?tab=commandes')
    context = {
        'page_title':  'Nouveau Bon de Commande — Cuisine',
        'fournisseurs': fournisseurs,
        'ingredients':  ingredients,
        'mode':         'create',
    }
    return render(request, 'cuisine/bon_commande_form.html', context)


@require_module_access('cuisine')
def bon_commande_detail(request, pk):
    bon   = get_object_or_404(BonCommandeCuisine, pk=pk)
    lignes = bon.lignes.select_related('ingredient').all()
    return render(request, 'cuisine/bon_commande_detail.html', {
        'page_title': f'Bon {bon.numero}', 'bon': bon, 'lignes': lignes
    })


@require_module_access('cuisine')
def bon_commande_print(request, pk):
    bon    = get_object_or_404(BonCommandeCuisine, pk=pk)
    lignes = bon.lignes.select_related('ingredient').all()
    return render(request, 'cuisine/bon_commande_print.html', {
        'bon': bon, 'lignes': lignes
    })


@require_module_access('cuisine')
def bon_commande_annuler(request, pk):
    bon = get_object_or_404(BonCommandeCuisine, pk=pk)
    if request.method == 'POST':
        bon.statut = 'annule'
        bon.save()
        messages.warning(request, f"Bon {bon.numero} annulé.")
    return redirect('/cuisine/stock/?tab=commandes')


# ==============================================================================
# BONS DE RÉCEPTION CUISINE
# ==============================================================================

@require_module_access('cuisine')
def bon_reception_list(request):
    return redirect('/cuisine/stock/?tab=receptions')


@require_module_access('cuisine')
def bon_reception_create(request):
    fournisseurs = Fournisseur.objects.filter(actif=True)
    bons_cmd     = BonCommandeCuisine.objects.filter(statut__in=['confirme', 'envoye', 'partiel'])
    ingredients  = Ingredient.objects.filter(statut=True)
    if request.method == 'POST':
        br = BonReceptionCuisine(
            fournisseur_id=request.POST.get('fournisseur') or None,
            bon_commande_id=request.POST.get('bon_commande') or None,
            date_reception=request.POST.get('date_reception') or timezone.now().date(),
            notes=request.POST.get('notes', ''),
            cree_par=request.user,
        )
        br.save()
        ing_ids = request.POST.getlist('ingredient_id[]')
        qtes    = request.POST.getlist('quantite_recue[]')
        prix_l  = request.POST.getlist('prix_unitaire[]')
        notes_l = request.POST.getlist('notes_ligne[]')
        for i, ing_id in enumerate(ing_ids):
            if ing_id and qtes[i]:
                LigneBonReceptionCuisine.objects.create(
                    bon=br,
                    ingredient_id=ing_id,
                    quantite_recue=qtes[i],
                    prix_unitaire=prix_l[i] if prix_l[i] else 0,
                    notes_ligne=notes_l[i] if i < len(notes_l) else '',
                )
        if request.POST.get('valider') == '1':
            ok = br.valider(request.user)
            if ok:
                messages.success(request, f"Réception {br.numero} validée — stock mis à jour.")
            else:
                messages.error(request, "Erreur lors de la validation.")
        else:
            messages.success(request, f"Réception {br.numero} enregistrée en brouillon.")
        return redirect('/cuisine/stock/?tab=receptions')
    import json
    ing_data = {
        str(ing.pk): {
            'nom':   ing.nom,
            'unite': ing.unite_stock.abreviation if ing.unite_stock else '-',
            'cmup':  float(ing.cmup or 0),
            'stock': float(ing.quantite_stock or 0),
        }
        for ing in ingredients
    }
    bc_data = {
        str(bc.pk): {
            'numero':          bc.numero,
            'fournisseur':     bc.fournisseur.nom if bc.fournisseur else '-',
            'fournisseur_id':  str(bc.fournisseur_id or ''),
        }
        for bc in bons_cmd
    }
    context = {
        'page_title':    'Nouvelle Réception — Cuisine',
        'fournisseurs':  fournisseurs,
        'bons_cmd':      bons_cmd,
        'ingredients':   ingredients,
        'ing_data_json': json.dumps(ing_data, ensure_ascii=False),
        'bc_data_json':  json.dumps(bc_data, ensure_ascii=False),
        'mode':          'create',
    }
    return render(request, 'cuisine/bon_reception_form.html', context)


@require_module_access('cuisine')
def bon_reception_detail(request, pk):
    br     = get_object_or_404(BonReceptionCuisine, pk=pk)
    lignes = br.lignes.select_related('ingredient').all()
    return render(request, 'cuisine/bon_reception_detail.html', {
        'page_title': f'Réception {br.numero}', 'br': br, 'lignes': lignes
    })
@require_module_access('cuisine')
def bon_reception_print(request, pk):
    """Version imprimable du bon de réception."""
    br     = get_object_or_404(BonReceptionCuisine, pk=pk)
    lignes = br.lignes.select_related('ingredient').all()
    return render(request, 'cuisine/bon_reception_print.html', {
        'br': br, 'lignes': lignes
    })




@require_module_access('cuisine')
def bon_reception_valider(request, pk):
    br = get_object_or_404(BonReceptionCuisine, pk=pk)
    if request.method == 'POST':
        ok = br.valider(request.user)
        if ok:
            messages.success(request, f"Réception {br.numero} validée — stock mis à jour.")
        else:
            messages.error(request, "Impossible de valider (déjà validée ou annulée).")
    return redirect('/cuisine/stock/?tab=receptions')


@require_module_access('cuisine')
def bon_reception_annuler(request, pk):
    br = get_object_or_404(BonReceptionCuisine, pk=pk)
    if request.method == 'POST':
        br.statut = 'annule'
        br.save()
        messages.warning(request, f"Réception {br.numero} annulée.")
    return redirect('/cuisine/stock/?tab=receptions')


# ==============================================================================
# FICHES TECHNIQUES
# ==============================================================================

@require_module_access('cuisine')
def fiche_list(request):
    fiches = FicheTechnique.objects.select_related('categorie').exclude(statut='archive')
    q = request.GET.get('q', '')
    cat = request.GET.get('categorie', '')
    if q:
        fiches = fiches.filter(nom__icontains=q)
    if cat:
        fiches = fiches.filter(categorie_id=cat)
    context = {
        'page_title': 'Fiches Techniques',
        'fiches':     fiches,
        'categories': CategoriePlat.objects.all(),
        'q': q, 'categorie_id': cat,
    }
    return render(request, 'cuisine/fiche_list.html', context)



def _ing_data_json_fiche(ingredients):
    import json
    return json.dumps({
        str(ing.pk): {
            'nom':   ing.nom,
            'unite': ing.unite_recette.abreviation if ing.unite_recette else '',
            'cmup':  float(ing.cmup or 0),
        }
        for ing in ingredients
    }, ensure_ascii=False)

@require_module_access('cuisine')
def fiche_create(request):
    categories   = CategoriePlat.objects.all()
    ingredients  = Ingredient.objects.filter(statut=True).select_related('unite_recette')
    if request.method == 'POST':
        fiche = FicheTechnique(
            nom=request.POST.get('nom'),
            categorie_id=request.POST.get('categorie') or None,
            description=request.POST.get('description', ''),
            nb_portions=request.POST.get('nb_portions') or 1,
            temps_preparation=request.POST.get('temps_preparation') or 0,
            temps_cuisson=request.POST.get('temps_cuisson') or 0,
            statut=request.POST.get('statut', 'actif'),
            cree_par=request.user,
        )
        fiche.save()
        # Lignes
        ing_ids  = request.POST.getlist('ingredient_id[]')
        qtes     = request.POST.getlist('quantite[]')
        notes_l  = request.POST.getlist('notes_ligne[]')
        for i, ing_id in enumerate(ing_ids):
            if ing_id and qtes[i]:
                LigneFicheTechnique.objects.create(
                    fiche=fiche,
                    ingredient_id=ing_id,
                    quantite=qtes[i],
                    notes_ligne=notes_l[i] if i < len(notes_l) else '',
                )
        messages.success(request, f"Fiche technique '{fiche.nom}' créée.")
        return redirect('/cuisine/fiches/')
    context = {
        'page_title':    'Nouvelle Fiche Technique',
        'categories':    categories,
        'ingredients':   ingredients,
        'ing_data_json': _ing_data_json_fiche(ingredients),
        'mode':          'create',
    }
    return render(request, 'cuisine/fiche_form.html', context)


@require_module_access('cuisine')
def fiche_edit(request, pk):
    fiche        = get_object_or_404(FicheTechnique, pk=pk)
    categories   = CategoriePlat.objects.all()
    ingredients  = Ingredient.objects.filter(statut=True).select_related('unite_recette')
    if request.method == 'POST':
        fiche.nom               = request.POST.get('nom')
        fiche.categorie_id      = request.POST.get('categorie') or None
        fiche.description       = request.POST.get('description', '')
        fiche.nb_portions       = request.POST.get('nb_portions') or 1
        fiche.temps_preparation = request.POST.get('temps_preparation') or 0
        fiche.temps_cuisson     = request.POST.get('temps_cuisson') or 0
        fiche.statut            = request.POST.get('statut', 'actif')
        fiche.save()
        # Reconstruire les lignes
        fiche.lignes.all().delete()
        ing_ids = request.POST.getlist('ingredient_id[]')
        qtes    = request.POST.getlist('quantite[]')
        notes_l = request.POST.getlist('notes_ligne[]')
        for i, ing_id in enumerate(ing_ids):
            if ing_id and qtes[i]:
                LigneFicheTechnique.objects.create(
                    fiche=fiche,
                    ingredient_id=ing_id,
                    quantite=qtes[i],
                    notes_ligne=notes_l[i] if i < len(notes_l) else '',
                )
        messages.success(request, f"Fiche technique '{fiche.nom}' modifiée.")
        return redirect('/cuisine/fiches/')
    context = {
        'page_title':    f'Modifier : {fiche.nom}',
        'fiche':         fiche,
        'lignes':        fiche.lignes.select_related('ingredient').all(),
        'categories':    categories,
        'ingredients':   ingredients,
        'ing_data_json': _ing_data_json_fiche(ingredients),
        'mode':          'edit',
    }
    return render(request, 'cuisine/fiche_form.html', context)


@require_module_access('cuisine')
def fiche_detail(request, pk):
    fiche  = get_object_or_404(FicheTechnique, pk=pk)
    lignes = fiche.lignes.select_related('ingredient__unite_recette').all()
    return render(request, 'cuisine/fiche_detail.html', {
        'page_title': fiche.nom, 'fiche': fiche, 'lignes': lignes
    })


@require_module_access('cuisine')
def fiche_print(request, pk):
    fiche  = get_object_or_404(FicheTechnique, pk=pk)
    lignes = fiche.lignes.select_related('ingredient__unite_recette').all()
    return render(request, 'cuisine/fiche_print.html', {'fiche': fiche, 'lignes': lignes})


@require_module_access('cuisine')
def fiche_delete(request, pk):
    fiche = get_object_or_404(FicheTechnique, pk=pk)
    if request.method == 'POST':
        nom = fiche.nom
        fiche.statut = 'archive'
        fiche.save()
        messages.success(request, f"Fiche '{nom}' archivée.")
        return redirect('/cuisine/fiches/')
    return render(request, 'cuisine/fiche_confirm_delete.html', {'fiche': fiche})


# ==============================================================================
# PLATS / CARTE
# ==============================================================================

@require_module_access('cuisine')
def plat_list(request):
    plats = Plat.objects.select_related('categorie', 'fiche_technique').exclude(statut='archive')
    q   = request.GET.get('q', '')
    cat = request.GET.get('categorie', '')
    if q:
        plats = plats.filter(nom__icontains=q)
    if cat:
        plats = plats.filter(categorie_id=cat)
    context = {
        'page_title': 'Plats / Carte',
        'plats':      plats,
        'categories': CategoriePlat.objects.all(),
        'q': q, 'categorie_id': cat,
    }
    return render(request, 'cuisine/plat_list.html', context)



def _save_fiche_from_plat(request, plat):
    """Crée ou met à jour la FicheTechnique liée au plat depuis le formulaire."""
    ing_ids  = request.POST.getlist('ingredient_id[]')
    quantites = request.POST.getlist('quantite[]')

    # Créer ou récupérer la fiche liée
    if plat.fiche_technique:
        fiche = plat.fiche_technique
    else:
        fiche = FicheTechnique(
            nom=plat.nom,
            cree_par=request.user,
        )

    fiche.nom                = plat.nom
    fiche.nb_portions        = request.POST.get('nb_portions') or 1
    fiche.temps_preparation  = request.POST.get('temps_preparation') or 15
    fiche.temps_cuisson      = request.POST.get('temps_cuisson') or 0
    fiche.description        = request.POST.get('description_technique', '')
    fiche.statut             = 'actif'
    if plat.categorie:
        from cuisine.models import CategoriePlat
        fiche.categorie = plat.categorie
    fiche.save()

    # Lier la fiche au plat
    if not plat.fiche_technique:
        plat.fiche_technique = fiche
        plat.save(update_fields=['fiche_technique'])

    # Recréer les lignes
    fiche.lignes.all().delete()
    for ing_id, qte in zip(ing_ids, quantites):
        if ing_id and qte:
            try:
                LigneFicheTechnique.objects.create(
                    fiche=fiche,
                    ingredient_id=int(ing_id),
                    quantite=float(qte),
                )
            except Exception:
                pass


def _sync_plat_to_restaurant(plat, ancien_nom=None):
    """Crée ou met à jour le PlatMenu correspondant dans le restaurant.
    ancien_nom : nom avant renommage, pour retrouver le bon PlatMenu existant.
    """
    try:
        from restaurant.models import PlatMenu, CategorieMenu
        # Trouver ou créer la catégorie restaurant correspondante
        cat_nom = plat.categorie.nom if plat.categorie else 'Plats'
        cat_resto, _ = CategorieMenu.objects.get_or_create(
            nom=cat_nom,
            defaults={'ordre': 1}
        )
        # Chercher d'abord par ancien nom (si renommage), puis par nom actuel
        nom_recherche = ancien_nom if ancien_nom else plat.nom
        plat_resto = PlatMenu.objects.filter(nom__iexact=nom_recherche).first()
        created = False
        if not plat_resto:
            # Pas trouvé non plus par le nom actuel → création
            plat_resto = PlatMenu(
                nom=plat.nom,
                categorie=cat_resto,
                prix=plat.prix_vente,
                temps_preparation=15,
                disponible=plat.statut == 'disponible',
                description=plat.description_carte,
            )
            created = True
        if not created:
            # Mettre à jour : nom (peut avoir changé), catégorie, prix, dispo, description
            plat_resto.nom         = plat.nom
            plat_resto.categorie   = cat_resto
            plat_resto.prix        = plat.prix_vente
            plat_resto.disponible  = plat.statut == 'disponible'
            plat_resto.description = plat.description_carte
        # is_accompagnement stocké dans le contexte de l'appel
        if hasattr(plat, '_is_accompagnement'):
            plat_resto.is_accompagnement = plat._is_accompagnement
        # Synchroniser la photo (convertir en JPEG RGB si nécessaire)
        if plat.image:
            try:
                from PIL import Image as PILImage
                from io import BytesIO
                from django.core.files.base import ContentFile
                plat.image.open('rb')
                img = PILImage.open(plat.image)
                img = img.convert('RGB')
                buffer = BytesIO()
                img.save(buffer, format='JPEG', quality=88)
                buffer.seek(0)
                img_name = plat.nom.replace(' ', '_') + '.jpg'
                plat_resto.image.save(img_name, ContentFile(buffer.read()), save=False)
                plat.image.close()
            except Exception:
                pass
        plat_resto.save()
    except Exception as e:
        pass  # Ne pas bloquer la sauvegarde si la synchro échoue

@require_module_access('cuisine')
def plat_create(request):
    categories   = CategoriePlat.objects.all()
    ingredients  = Ingredient.objects.filter(statut=True).select_related('unite_recette')
    if request.method == 'POST':
        plat = Plat(
            nom=request.POST.get('nom'),
            categorie_id=request.POST.get('categorie') or None,
            description_carte=request.POST.get('description_carte', ''),
            prix_vente=request.POST.get('prix_vente') or 0,
            statut=request.POST.get('statut', 'disponible'),
        )
        if request.FILES.get('image'):
            plat.image = request.FILES['image']
        # Mémoriser l'ancien nom AVANT save pour la synchro restaurant
        ancien_nom = Plat.objects.get(pk=plat.pk).nom if plat.pk else None
        plat.save()
        _save_fiche_from_plat(request, plat)
        plat._is_accompagnement = request.POST.get('is_accompagnement') == '1'
        _sync_plat_to_restaurant(plat, ancien_nom=ancien_nom)
        messages.success(request, f"Plat '{plat.nom}' créé avec sa fiche technique.")
        return redirect('/cuisine/plats/')
    import json
    ing_data = _ing_data_json_fiche(ingredients)
    context = {
        'page_title':    'Nouveau Plat',
        'categories':    categories,
        'ingredients':   ingredients,
        'ing_data_json': ing_data,
        'mode':          'create',
    }
    return render(request, 'cuisine/plat_form.html', context)


@require_module_access('cuisine')




@require_module_access('cuisine')
def sync_plats_restaurant(request):
    """Synchronise tous les plats cuisine vers le restaurant."""
    from cuisine.models import Plat
    plats = Plat.objects.exclude(statut='archive').select_related('categorie','fiche_technique')
    count = 0
    for plat in plats:
        _sync_plat_to_restaurant(plat)
        count += 1
    messages.success(request, f"{count} plat(s) synchronisé(s) avec le menu restaurant.")
    return redirect('/cuisine/plats/')


@require_module_access('cuisine')
def plat_edit(request, pk):
    plat        = get_object_or_404(Plat, pk=pk)
    categories  = CategoriePlat.objects.all()
    ingredients = Ingredient.objects.filter(statut=True).select_related('unite_recette')
    if request.method == 'POST':
        plat.nom               = request.POST.get('nom')
        plat.categorie_id      = request.POST.get('categorie') or None
        plat.description_carte = request.POST.get('description_carte', '')
        plat.prix_vente        = request.POST.get('prix_vente') or 0
        plat.statut            = request.POST.get('statut', 'disponible')
        if request.FILES.get('image'):
            plat.image = request.FILES['image']
        elif request.POST.get('image-clear'):
            plat.image = None
        # Mémoriser l'ancien nom AVANT save pour la synchro restaurant
        ancien_nom = Plat.objects.get(pk=plat.pk).nom if plat.pk else None
        plat.save()
        _save_fiche_from_plat(request, plat)
        plat._is_accompagnement = request.POST.get('is_accompagnement') == '1'
        _sync_plat_to_restaurant(plat, ancien_nom=ancien_nom)
        messages.success(request, f"Plat '{plat.nom}' modifié et synchronisé.")
        return redirect('/cuisine/plats/')
    from restaurant.models import PlatMenu
    plat_menu = PlatMenu.objects.filter(nom=plat.nom).first()
    ing_data = _ing_data_json_fiche(ingredients)
    context = {
        'page_title':         f'Modifier : {plat.nom}',
        'plat':               plat,
        'categories':         categories,
        'ingredients':        ingredients,
        'ing_data_json':      ing_data,
        'is_accompagnement':  plat_menu.is_accompagnement if plat_menu else False,
        'mode':               'edit',
    }
    return render(request, 'cuisine/plat_form.html', context)

@require_module_access('cuisine')
def plat_delete(request, pk):
    plat = get_object_or_404(Plat, pk=pk)
    if request.method == 'POST':
        plat.statut = 'archive'
        plat.save()
        messages.success(request, f"Plat '{plat.nom}' archivé.")
        return redirect('/cuisine/plats/')
    return render(request, 'cuisine/plat_confirm_delete.html', {'plat': plat})


# ==============================================================================
# FOURNISSEURS
# ==============================================================================

@require_module_access('cuisine')
def fournisseur_list(request):
    fournisseurs = Fournisseur.objects.filter(actif=True)
    q = request.GET.get('q', '')
    if q:
        fournisseurs = fournisseurs.filter(Q(nom__icontains=q) | Q(telephone__icontains=q))
    context = {
        'page_title':    'Fournisseurs',
        'fournisseurs':  fournisseurs,
        'q':             q,
    }
    return render(request, 'cuisine/fournisseur_list.html', context)


@require_module_access('cuisine')
def fournisseur_create(request):
    if request.method == 'POST':
        f = Fournisseur(
            nom=request.POST.get('nom'),
            type_fournisseur=request.POST.get('type_fournisseur', 'grossiste'),
            personne_contact=request.POST.get('personne_contact', ''),
            telephone=request.POST.get('telephone', ''),
            telephone2=request.POST.get('telephone2', ''),
            email=request.POST.get('email', ''),
            adresse=request.POST.get('adresse', ''),
            ville=request.POST.get('ville', ''),
            notes=request.POST.get('notes', ''),
        )
        f.save()
        messages.success(request, f"Fournisseur '{f.nom}' créé.")
        return redirect('/cuisine/fournisseurs/')
    return render(request, 'cuisine/fournisseur_form.html', {'page_title': 'Nouveau Fournisseur', 'mode': 'create'})


@require_module_access('cuisine')
def fournisseur_edit(request, pk):
    f = get_object_or_404(Fournisseur, pk=pk)
    if request.method == 'POST':
        f.nom              = request.POST.get('nom')
        f.type_fournisseur = request.POST.get('type_fournisseur', 'grossiste')
        f.personne_contact = request.POST.get('personne_contact', '')
        f.telephone        = request.POST.get('telephone', '')
        f.telephone2       = request.POST.get('telephone2', '')
        f.email            = request.POST.get('email', '')
        f.adresse          = request.POST.get('adresse', '')
        f.ville            = request.POST.get('ville', '')
        f.notes            = request.POST.get('notes', '')
        f.save()
        messages.success(request, f"Fournisseur '{f.nom}' modifié.")
        return redirect('/cuisine/fournisseurs/')
    return render(request, 'cuisine/fournisseur_form.html', {
        'page_title': f'Modifier : {f.nom}', 'fournisseur': f, 'mode': 'edit'
    })


@require_module_access('cuisine')
def fournisseur_delete(request, pk):
    f = get_object_or_404(Fournisseur, pk=pk)
    if request.method == 'POST':
        f.actif = False
        f.save()
        messages.success(request, f"Fournisseur '{f.nom}' désactivé.")
        return redirect('/cuisine/fournisseurs/')
    return render(request, 'cuisine/fournisseur_confirm_delete.html', {'fournisseur': f})


# ==============================================================================
# INVENTAIRE
# ==============================================================================

@require_module_access('cuisine')
def inventaire_create(request):
    ingredients = Ingredient.objects.filter(statut=True).select_related('categorie', 'unite_stock')
    if request.method == 'POST':
        inv = InventaireCuisine(
            date_inventaire=request.POST.get('date_inventaire') or timezone.now().date(),
            notes=request.POST.get('notes', ''),
            cree_par=request.user,
        )
        inv.save()
        ing_ids  = request.POST.getlist('ingredient_id[]')
        th_list  = request.POST.getlist('quantite_theorique[]')
        ph_list  = request.POST.getlist('quantite_physique[]')
        for i, ing_id in enumerate(ing_ids):
            if ing_id:
                LigneInventaireCuisine.objects.create(
                    inventaire=inv,
                    ingredient_id=ing_id,
                    quantite_theorique=th_list[i] if th_list[i] else 0,
                    quantite_physique=ph_list[i] if ph_list[i] else 0,
                )
        if request.POST.get('valider') == '1':
            inv.valider(request.user)
            messages.success(request, f"Inventaire {inv.numero} validé.")
        else:
            messages.success(request, f"Inventaire {inv.numero} enregistré.")
        return redirect('/cuisine/stock/?tab=inventaire')
    context = {
        'page_title':   'Nouvel Inventaire — Cuisine',
        'ingredients':  ingredients,
    }
    return render(request, 'cuisine/inventaire_form.html', context)


@require_module_access('cuisine')
def inventaire_valider(request, pk):
    inv = get_object_or_404(InventaireCuisine, pk=pk)
    if request.method == 'POST':
        ok = inv.valider(request.user)
        if ok:
            messages.success(request, f"Inventaire {inv.numero} validé.")
        else:
            messages.error(request, "Impossible de valider.")
    return redirect('/cuisine/stock/?tab=inventaire')


# ==============================================================================
# CASSES
# ==============================================================================

@require_module_access('cuisine')
def casse_create(request):
    ingredients = Ingredient.objects.filter(statut=True)
    if request.method == 'POST':
        casse = CasseCuisine(
            type_casse=request.POST.get('type_casse', 'casse'),
            date_casse=request.POST.get('date_casse') or timezone.now().date(),
            description=request.POST.get('description', ''),
            cree_par=request.user,
        )
        casse.save()
        ing_ids = request.POST.getlist('ingredient_id[]')
        qtes    = request.POST.getlist('quantite[]')
        notes_l = request.POST.getlist('notes_ligne[]')
        for i, ing_id in enumerate(ing_ids):
            if ing_id and qtes[i]:
                LigneCasseCuisine.objects.create(
                    casse=casse,
                    ingredient_id=ing_id,
                    quantite=qtes[i],
                    notes_ligne=notes_l[i] if i < len(notes_l) else '',
                )
        messages.success(request, f"Déclaration {casse.numero} créée.")
        return redirect('/cuisine/stock/?tab=casses')
    context = {
        'page_title':   'Déclarer une casse — Cuisine',
        'ingredients':  ingredients,
    }
    return render(request, 'cuisine/casse_form.html', context)


@require_module_access('cuisine')
def casse_valider(request, pk):
    casse = get_object_or_404(CasseCuisine, pk=pk)
    if request.method == 'POST':
        ok = casse.valider(request.user)
        if ok:
            messages.success(request, f"Casse {casse.numero} validée — stock déduit.")
        else:
            messages.error(request, "Impossible de valider.")
    return redirect('/cuisine/stock/?tab=casses')


# ==============================================================================
# AJAX HELPERS
# ==============================================================================

@require_module_access('cuisine')
def get_ingredient_prix(request, pk):
    ing = get_object_or_404(Ingredient, pk=pk)
    return JsonResponse({
        'prix_achat': float(ing.prix_achat),
        'cmup':       float(ing.cmup),
        'unite':      str(ing.unite_stock) if ing.unite_stock else '',
        'reference':  ing.reference or '',
        'stock':      float(ing.quantite_stock),
    })


@require_module_access('cuisine')
def get_bc_lignes(request, pk):
    """AJAX : lignes d'un BC pour pré-remplir un bon de réception"""
    bon    = get_object_or_404(BonCommandeCuisine, pk=pk)
    lignes = []
    for l in bon.lignes.select_related('ingredient__unite_stock').all():
        lignes.append({
            'ingredient_id':   l.ingredient.pk,
            'ingredient_nom':  l.ingredient.nom,
            'quantite':        float(l.reliquat),
            'prix_unitaire':   float(l.prix_unitaire),
        })
    return JsonResponse({'lignes': lignes, 'fournisseur_id': bon.fournisseur_id or ''})

@require_module_access('cuisine')
def etat_stock_print(request):
    from .models import Ingredient
    ingredients = Ingredient.objects.filter(statut=True).select_related('categorie','unite_stock').order_by('categorie__nom','nom')
    from django.db.models import Sum, F, ExpressionWrapper, DecimalField
    total_valeur = sum(float(i.cmup or 0) * float(i.quantite_stock or 0) for i in ingredients)
    return render(request, 'cuisine/etat_stock_print.html', {
        'ingredients': ingredients,
        'total_valeur': total_valeur,
    })


def etat_stock_excel(request):
    """Export Excel du stock cuisine."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from .models import Ingredient
    from django.utils import timezone

    ingredients = Ingredient.objects.filter(statut=True).select_related('categorie','unite_stock').order_by('categorie__nom','nom')
    total_valeur = sum(float(i.cmup or 0) * float(i.quantite_stock or 0) for i in ingredients)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "État du Stock"

    # Styles
    header_fill = PatternFill("solid", fgColor="1a2535")
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    title_font  = Font(name='Calibri', bold=True, size=14, color='1a2535')
    bold_font   = Font(name='Calibri', bold=True, size=10)
    normal_font = Font(name='Calibri', size=10)
    ok_fill     = PatternFill("solid", fgColor="dcfce7")
    alerte_fill = PatternFill("solid", fgColor="fef3c7")
    rupture_fill= PatternFill("solid", fgColor="fee2e2")
    thin = Side(border_style="thin", color="d4dce8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')

    # Titre
    ws.merge_cells('A1:I1')
    ws['A1'] = f"ÉTAT DU STOCK — CUISINE BEHANIAN"
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells('A2:I2')
    ws['A2'] = f"Édité le {timezone.now().strftime('%d/%m/%Y à %H:%M')} — {ingredients.count()} article(s)"
    ws['A2'].font = Font(name='Calibri', size=10, color='7a8b9c', italic=True)
    ws['A2'].alignment = center

    ws.append([])

    # En-têtes
    headers = ['#', 'Ingrédient', 'Catégorie', 'Unité', 'Stock actuel', 'Seuil alerte', 'CMUP (FCFA)', 'Valeur stock (FCFA)', 'État']
    ws.append(headers)
    row_h = ws.max_row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_h, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Données
    for i, ing in enumerate(ingredients, 1):
        val = float(ing.cmup or 0) * float(ing.quantite_stock or 0)
        if ing.quantite_stock <= 0:
            etat = 'Rupture'; fill = rupture_fill
        elif ing.seuil_alerte and ing.quantite_stock <= ing.seuil_alerte:
            etat = 'Alerte'; fill = alerte_fill
        else:
            etat = 'Normal'; fill = ok_fill

        row = [
            i,
            ing.nom,
            ing.categorie.nom if ing.categorie else '—',
            ing.unite_stock.abreviation if ing.unite_stock else '—',
            float(ing.quantite_stock or 0),
            float(ing.seuil_alerte or 0),
            float(ing.cmup or 0),
            round(val, 0),
            etat,
        ]
        ws.append(row)
        r = ws.max_row
        for col in range(1, 10):
            cell = ws.cell(row=r, column=col)
            cell.font = bold_font if col in (2, 8) else normal_font
            cell.border = border
            cell.alignment = right if col in (5,6,7,8) else (center if col in (1,4,9) else Alignment(vertical='center'))
            if col == 9:
                cell.fill = fill
                cell.font = Font(name='Calibri', bold=True, size=10,
                    color='15803d' if etat=='Normal' else ('d97706' if etat=='Alerte' else 'dc2626'))

    # Total
    ws.append([])
    r = ws.max_row + 1
    ws.cell(row=r, column=7, value='VALEUR TOTALE').font = Font(bold=True, size=11)
    ws.cell(row=r, column=8, value=round(total_valeur, 0)).font = Font(bold=True, size=11, color='16a34a')
    ws.cell(row=r, column=8).number_format = '#,##0'

    # Largeurs colonnes
    for col, w in enumerate([5, 28, 16, 10, 13, 13, 13, 18, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    from django.utils import timezone as tz
    fname = f"Stock_Cuisine_{tz.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


@require_module_access('cuisine')
def rapport_stock_cuisine(request):
    get = request.GET.copy()
    if 'module' not in get:
        get['module'] = 'cuisine'
    request.GET = get
    from rapport.views import rapport_stock
    return rapport_stock(request)
