from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from decimal import Decimal
import datetime
from django.utils import timezone

from .models import Ingredient, MouvementStock, CategorieIngredient

@login_required
def export_stock_excel(request):
    """Exporte l'état du stock valorisé en fichier Excel."""

    # --- 1. Récupérer les filtres et calculer les données (logique similaire à etat_stock)
    date_str = request.GET.get('date')
    categorie_ids = request.GET.getlist('categorie')

    try:
        date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        target_date = datetime.datetime.combine(date_obj, datetime.time.max).replace(tzinfo=timezone.get_current_timezone())
    except (ValueError, TypeError):
        target_date = timezone.now()

    ingredients_qs = Ingredient.objects.select_related('unite').all()
    if categorie_ids:
        ingredients_qs = ingredients_qs.filter(categorie_id__in=categorie_ids)

    stock_data = []
    total_valeur_cmup = Decimal('0.0')
    total_valeur_vente = Decimal('0.0')

    for ing in ingredients_qs.order_by('nom'):
        stock_at_date = ing.quantite_stock
        mouvements_apres = MouvementStock.objects.filter(ingredient=ing, date__gt=target_date)
        for mvt in mouvements_apres:
            stock_at_date -= mvt.quantite

        valeur_cmup = stock_at_date * ing.prix_moyen
        valeur_vente = stock_at_date * ing.prix_vente
        marge = valeur_vente - valeur_cmup

        total_valeur_cmup += valeur_cmup
        total_valeur_vente += valeur_vente
        
        stock_data.append({
            'nom': ing.nom,
            'quantite': stock_at_date,
            'unite': ing.unite.abreviation if ing.unite else '',
            'prix_moyen': ing.prix_moyen,
            'valeur_cmup': valeur_cmup,
            'prix_vente': ing.prix_vente,
            'valeur_vente': valeur_vente,
            'marge': marge,
        })
    
    total_marge = total_valeur_vente - total_valeur_cmup

    # --- 2. Création du fichier Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "État du Stock"

    # --- 3. Styles ---
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- 4. Titre et En-tête ---
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"État du Stock Valorisé au {target_date.strftime('%d/%m/%Y')}"
    title_cell.font = Font(name='Calibri', size=16, bold=True)
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 30

    headers = ["Article", "Qté", "Unité", "CMUP", "Valeur Stock", "Prix Vente", "Valeur Vente", "Marge"]
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- 5. Remplissage des données ---
    row_num = 4
    for item in stock_data:
        ws.cell(row=row_num, column=1, value=item['nom']).border = thin_border
        ws.cell(row=row_num, column=2, value=item['quantite']).number_format = '0.00'
        ws.cell(row=row_num, column=3, value=item['unite']).alignment = center_align
        ws.cell(row=row_num, column=4, value=item['prix_moyen']).number_format = '#,##0'
        ws.cell(row=row_num, column=5, value=item['valeur_cmup']).number_format = '#,##0'
        ws.cell(row=row_num, column=6, value=item['prix_vente']).number_format = '#,##0'
        ws.cell(row=row_num, column=7, value=item['valeur_vente']).number_format = '#,##0'
        ws.cell(row=row_num, column=8, value=item['marge']).number_format = '#,##0'
        # Appliquer les bordures et l'alignement
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = thin_border
            if col in [2, 4, 5, 6, 7, 8]:
                ws.cell(row=row_num, column=col).alignment = right_align
        row_num += 1

    # --- 6. Ligne des totaux ---
    ws.cell(row=row_num, column=4, value="TOTAUX").font = Font(bold=True)
    ws.cell(row=row_num, column=5, value=total_valeur_cmup).number_format = '#,##0 FCFA'
    ws.cell(row=row_num, column=7, value=total_valeur_vente).number_format = '#,##0 FCFA'
    ws.cell(row=row_num, column=8, value=total_marge).number_format = '#,##0 FCFA'
    for col in [4, 5, 7, 8]:
        ws.cell(row=row_num, column=col).font = Font(bold=True)
        ws.cell(row=row_num, column=col).alignment = right_align
        ws.cell(row=row_num, column=col).border = thin_border

    # --- 7. Ajuster la largeur des colonnes ---
    column_widths = {'A': 40, 'B': 10, 'C': 10, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # --- 8. Créer la réponse HTTP ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=etat_stock_{target_date.strftime("%Y-%m-%d")}.xlsx'
    wb.save(response)

    return response
