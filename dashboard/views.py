from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.http import JsonResponse
from utils.permissions import get_accessible_modules
from datetime import timedelta


def _get_dashboard_stats(user, modules):
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)

    stats = {
        'date_ref': None,
        'ca_jour': 0, 'ca_hier': 0, 'ca_mois': 0,
        'nb_tickets_jour': 0,
        'taux_occupation': 0, 'chambres_occupees': 0, 'total_chambres': 0,
        'reservations_actives': 0, 'reservations_attente': 0,
        'alertes_stock': 0,
        'caisse_ouverte': False, 'caisse_type': '',
        'tickets_recents': [],
        'ca_par_module': {},
        'ca_7_jours': [],
        'piscine_entrees': 0,
        'espaces_reservations': 0,
        'commandes_restaurant': 0,
    }

    try:
        from facturation.models import Ticket
        tickets_jour = Ticket.objects.filter(date_creation__date=today)
        tickets_hier = Ticket.objects.filter(date_creation__date=yesterday)
        tickets_mois = Ticket.objects.filter(
            date_creation__month=today.month,
            date_creation__year=today.year
        )

        stats['ca_jour'] = int(tickets_jour.aggregate(s=Sum('montant_total'))['s'] or 0)
        stats['ca_hier'] = int(tickets_hier.aggregate(s=Sum('montant_total'))['s'] or 0)
        stats['ca_mois'] = int(tickets_mois.aggregate(s=Sum('montant_total'))['s'] or 0)
        stats['nb_tickets_jour'] = tickets_jour.count()

        # Si pas de tickets aujourd'hui → utiliser la dernière journée active
        tickets_module_ref = tickets_jour
        stats['date_ref'] = today
        if not tickets_jour.exists():
            last = Ticket.objects.order_by('-date_creation').first()
            if last:
                last_date = last.date_creation.date()
                tickets_module_ref = Ticket.objects.filter(date_creation__date=last_date)
                stats['date_ref'] = last_date
                stats['ca_jour'] = int(tickets_module_ref.aggregate(s=Sum('montant_total'))['s'] or 0)
                stats['nb_tickets_jour'] = tickets_module_ref.count()

        # CA par module
        for mod, label in [('hotel','Hôtel'),('restaurant','Restaurant'),('cave','Cave'),('piscine','Piscine'),('espace','Espaces')]:
            ca = tickets_module_ref.filter(module__startswith=mod).aggregate(s=Sum('montant_total'))['s'] or 0
            if ca: stats['ca_par_module'][label] = int(ca)

        # CA 7 derniers jours
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            ca = Ticket.objects.filter(date_creation__date=d).aggregate(s=Sum('montant_total'))['s'] or 0
            stats['ca_7_jours'].append({'jour': d.strftime('%a'), 'ca': int(ca)})

        # Tickets récents
        stats['tickets_recents'] = list(
            tickets_jour.select_related('client','cree_par').order_by('-date_creation')[:8].values(
                'numero','module','montant_total','mode_paiement',
                'date_creation','cree_par__first_name','cree_par__last_name','cree_par__username'
            )
        )
    except Exception:
        pass

    # Hôtel
    if 'hotel' in modules or '*' in modules:
        try:
            from hotel.models import Chambre, Reservation
            stats['total_chambres'] = Chambre.objects.count()
            stats['chambres_occupees'] = Chambre.objects.filter(statut='occupee').count()
            stats['taux_occupation'] = round(
                (stats['chambres_occupees'] / stats['total_chambres'] * 100)
                if stats['total_chambres'] else 0
            )
            stats['reservations_actives'] = Reservation.objects.filter(statut='en_cours').count()
            stats['reservations_attente'] = Reservation.objects.filter(statut__in=['en_attente','confirmee']).count()
        except Exception:
            pass

    # Cuisine stock
    if 'cuisine' in modules or '*' in modules:
        try:
            from cuisine.models import Ingredient
            stats['alertes_stock'] = Ingredient.objects.filter(quantite_stock__lte=5).count()
        except Exception:
            pass

    # Caisse ouverte
    try:
        from caisse.models import CaisseSession
        # La caisse centrale est ouverte uniquement si une session de type 'centrale' existe
        session_centrale = CaisseSession.objects.filter(is_open=True, type_caisse='centrale').first()
        session_any = CaisseSession.objects.filter(is_open=True).first()
        stats['caisse_ouverte'] = session_centrale is not None
        stats['caisse_type'] = session_centrale.get_type_caisse_display() if session_centrale else (
            session_any.get_type_caisse_display() if session_any else ''
        )
    except Exception:
        pass

    # Piscine entrées du jour
    try:
        from facturation.models import Ticket
        stats['piscine_entrees'] = Ticket.objects.filter(
            date_creation__date=today, module='piscine'
        ).count()
    except Exception:
        pass

    # Espaces réservations actives
    try:
        from espaces_evenementiels.models import ReservationEspace
        stats['espaces_reservations'] = ReservationEspace.objects.filter(statut='confirmee').count()
    except Exception:
        pass

    # Restaurant commandes en cours
    try:
        from restaurant.models import Commande
        stats['commandes_restaurant'] = Commande.objects.filter(statut='en_attente').count()
    except Exception:
        pass

    return stats


@login_required
def dashboard_view(request):
    today = timezone.now().date()
    modules = get_accessible_modules(request.user)
    stats = _get_dashboard_stats(request.user, modules)

    context = {
        **stats,
        'date_ref': stats.get('date_ref', today),
        'user': request.user,
        'today': today,
        'accessible_modules': modules,
        'variation': f"+{round(((stats["ca_jour"]-stats["ca_hier"])/stats["ca_hier"]*100) if stats["ca_hier"] else 0)}%",
    }
    return render(request, 'dashboard/dashboard.html', context)


@login_required
def direction_view(request):
    """Vue consolidée Direction — stocks, mouvements et stats de tous les modules."""
    from utils.permissions import _is_manager
    from django.contrib import messages
    from datetime import date as date_type
    if not (_is_manager(request.user) or request.user.is_superuser):
        messages.error(request, "Accès réservé à la Direction et aux Managers.")
        return redirect('dashboard:index')

    today = timezone.now().date()
    modules = get_accessible_modules(request.user)
    stats = _get_dashboard_stats(request.user, modules)

    # Filtre période mouvements — défaut = aujourd'hui
    active_tab = request.GET.get('tab', 'global')
    date_debut_str = request.GET.get('date_debut', today.isoformat())
    date_fin_str   = request.GET.get('date_fin',   today.isoformat())
    try:
        date_debut = date_type.fromisoformat(date_debut_str)
    except ValueError:
        date_debut = today
    try:
        date_fin = date_type.fromisoformat(date_fin_str)
    except ValueError:
        date_fin = today
    if date_fin < date_debut:
        date_fin = date_debut

    bar_ruptures, bar_alertes = 0, 0
    cuisine_ruptures, cuisine_alertes = 0, 0
    mouvements_combines = []

    try:
        from bar.models import BoissonBar, MouvementStockBar
        bar_qs = BoissonBar.objects.filter(statut='actif')
        bar_ruptures = sum(1 for a in bar_qs if a.est_en_rupture)
        bar_alertes  = sum(1 for a in bar_qs if a.est_stock_bas)
        qs_bar = MouvementStockBar.objects.filter(
            date__date__gte=date_debut, date__date__lte=date_fin
        ).select_related('boisson', 'utilisateur').order_by('-date')
        for m in qs_bar:
            mouvements_combines.append({
                'source': 'cave', 'nom': m.boisson.nom,
                'type': m.get_type_mouvement_display(), 'type_code': m.type_mouvement,
                'quantite': m.quantite, 'date': m.date,
                'user': m.utilisateur, 'commentaire': m.commentaire or '',
            })
    except Exception:
        pass

    try:
        from cuisine.models import Ingredient, MouvementStockCuisine
        cuisine_qs = Ingredient.objects.filter(statut=True)
        cuisine_ruptures = sum(1 for i in cuisine_qs if i.est_en_rupture)
        cuisine_alertes  = sum(1 for i in cuisine_qs if i.est_stock_bas)
        qs_cuisine = MouvementStockCuisine.objects.filter(
            date__date__gte=date_debut, date__date__lte=date_fin
        ).select_related('ingredient', 'utilisateur').order_by('-date')
        for m in qs_cuisine:
            mouvements_combines.append({
                'source': 'cuisine', 'nom': m.ingredient.nom,
                'type': m.get_type_mouvement_display(), 'type_code': m.type_mouvement,
                'quantite': m.quantite, 'date': m.date,
                'user': m.utilisateur, 'commentaire': m.commentaire or '',
            })
    except Exception:
        pass

    mouvements_combines.sort(key=lambda x: x['date'], reverse=True)

    # ── Fond de caisse & sessions ──────────────────────────────
    sessions_jour = []
    stats_caisse_jour = {}
    solde_veille_dir = 0
    try:
        from caisse.models import CaisseSession
        from caisse.views import get_stats_jour, get_solde_veille
        today_local = timezone.localdate()
        sessions_jour = list(
            CaisseSession.objects.filter(date_session=today_local)
            .select_related('user').order_by('-opened_at')
        )
        stats_caisse_jour = get_stats_jour(today_local)
        solde_veille_dir, _ = get_solde_veille()
    except Exception:
        pass

    context = {
        **stats,
        'today': today,
        'accessible_modules': modules,
        'bar_ruptures': bar_ruptures,
        'bar_alertes': bar_alertes,
        'cuisine_ruptures': cuisine_ruptures,
        'cuisine_alertes': cuisine_alertes,
        'mouvements_combines': mouvements_combines,
        'sessions_jour': sessions_jour,
        'stats_caisse_jour': stats_caisse_jour,
        'solde_veille_dir': solde_veille_dir,
        'active_tab': active_tab,
        'date_debut': date_debut.isoformat(),
        'date_fin':   date_fin.isoformat(),
    }
    return render(request, 'dashboard/direction.html', context)


@login_required
def mouvements_print(request):
    """Page d'impression des mouvements de stock — modèle print_base.html."""
    from utils.permissions import _is_manager
    from django.contrib import messages
    from datetime import date as date_type

    if not (_is_manager(request.user) or request.user.is_superuser):
        messages.error(request, "Accès réservé à la Direction.")
        return redirect('dashboard:index')

    today_p = timezone.now().date()
    date_debut_str = request.GET.get('date_debut', today_p.isoformat())
    date_fin_str   = request.GET.get('date_fin',   today_p.isoformat())
    try:
        date_debut = date_type.fromisoformat(date_debut_str)
    except ValueError:
        date_debut = today_p
    try:
        date_fin = date_type.fromisoformat(date_fin_str)
    except ValueError:
        date_fin = today_p
    if date_fin < date_debut:
        date_fin = date_debut

    mouvements = []
    try:
        from bar.models import MouvementStockBar
        for m in MouvementStockBar.objects.filter(
            date__date__gte=date_debut, date__date__lte=date_fin
        ).select_related('boisson', 'utilisateur').order_by('-date'):
            mouvements.append({
                'source': 'cave', 'source_label': 'Cave',
                'nom': m.boisson.nom,
                'type': m.get_type_mouvement_display(), 'type_code': m.type_mouvement,
                'quantite': m.quantite, 'unite': 'unité(s)',
                'date': m.date,
                'user': m.utilisateur.get_full_name() or m.utilisateur.username if m.utilisateur else '—',
                'commentaire': m.commentaire or '—',
            })
    except Exception:
        pass

    try:
        from cuisine.models import MouvementStockCuisine
        for m in MouvementStockCuisine.objects.filter(
            date__date__gte=date_debut, date__date__lte=date_fin
        ).select_related('ingredient', 'utilisateur').order_by('-date'):
            mouvements.append({
                'source': 'cuisine', 'source_label': 'Cuisine',
                'nom': m.ingredient.nom,
                'type': m.get_type_mouvement_display(), 'type_code': m.type_mouvement,
                'quantite': m.quantite, 'unite': str(m.ingredient.unite_stock) if m.ingredient.unite_stock else '—',
                'date': m.date,
                'user': m.utilisateur.get_full_name() or m.utilisateur.username if m.utilisateur else '—',
                'commentaire': m.commentaire or '—',
            })
    except Exception:
        pass

    mouvements.sort(key=lambda x: x['date'], reverse=True)

    return render(request, 'dashboard/mouvements_print.html', {
        'date_debut': date_debut,
        'date_fin':   date_fin,
        'mouvements': mouvements,
        'nb_cave': sum(1 for m in mouvements if m['source'] == 'cave'),
        'nb_cuisine': sum(1 for m in mouvements if m['source'] == 'cuisine'),
        'generated_at': timezone.now(),
    })


@login_required
def mouvements_excel(request):
    """Export Excel : mouvements de stock globaux (Cave + Cuisine) — Vue Direction"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from utils.permissions import _is_manager
    from datetime import date as date_type

    if not (_is_manager(request.user) or request.user.is_superuser):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    today_x = timezone.now().date()
    date_debut_str = request.GET.get('date_debut', today_x.isoformat())
    date_fin_str   = request.GET.get('date_fin',   today_x.isoformat())
    try:
        date_debut = date_type.fromisoformat(date_debut_str)
    except ValueError:
        date_debut = today_x
    try:
        date_fin = date_type.fromisoformat(date_fin_str)
    except ValueError:
        date_fin = today_x
    if date_fin < date_debut:
        date_fin = date_debut

    mouvements = []
    try:
        from bar.models import MouvementStockBar
        for m in MouvementStockBar.objects.filter(
            date__date__gte=date_debut, date__date__lte=date_fin
        ).select_related('boisson', 'utilisateur').order_by('-date'):
            mouvements.append({
                'source': 'Cave', 'nom': m.boisson.nom,
                'type': m.get_type_mouvement_display(), 'type_code': m.type_mouvement,
                'quantite': float(m.quantite), 'unite': 'unité(s)',
                'date': m.date,
                'user': m.utilisateur.get_full_name() or m.utilisateur.username if m.utilisateur else '—',
                'commentaire': m.commentaire or '',
            })
    except Exception:
        pass
    try:
        from cuisine.models import MouvementStockCuisine
        for m in MouvementStockCuisine.objects.filter(
            date__date__gte=date_debut, date__date__lte=date_fin
        ).select_related('ingredient', 'utilisateur').order_by('-date'):
            mouvements.append({
                'source': 'Cuisine', 'nom': m.ingredient.nom,
                'type': m.get_type_mouvement_display(), 'type_code': m.type_mouvement,
                'quantite': float(m.quantite),
                'unite': str(m.ingredient.unite_stock) if m.ingredient.unite_stock else '—',
                'date': m.date,
                'user': m.utilisateur.get_full_name() or m.utilisateur.username if m.utilisateur else '—',
                'commentaire': m.commentaire or '',
            })
    except Exception:
        pass
    mouvements.sort(key=lambda x: x['date'], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mouvements"

    hf  = PatternFill("solid", fgColor="1a2535")
    hft = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    tf  = Font(name='Calibri', bold=True, size=14, color='1a2535')
    bf  = Font(name='Calibri', bold=True, size=10)
    nf  = Font(name='Calibri', size=10)
    th  = Side(border_style="thin", color="d4dce8")
    bd  = Border(left=th, right=th, top=th, bottom=th)
    ct  = Alignment(horizontal='center', vertical='center')
    rt  = Alignment(horizontal='right', vertical='center')

    NC = 9
    ws.merge_cells(f'A1:{get_column_letter(NC)}1')
    periode = f"{date_debut.strftime('%d/%m/%Y')}" if date_debut == date_fin else f"{date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    ws['A1'] = f"MOUVEMENTS DE STOCK — {periode.upper()} — COMPLEXE BEHANIAN"
    ws['A1'].font = tf; ws['A1'].alignment = ct

    ws.merge_cells(f'A2:{get_column_letter(NC)}2')
    ws['A2'] = f"{len(mouvements)} mouvement(s) — Édité le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(name='Calibri', size=10, color='7a8b9c', italic=True)
    ws['A2'].alignment = ct

    ws.append([])
    headers = ['#', 'Module', 'Article', 'Type de mouvement', 'Quantité', 'Unité', 'Heure', 'Utilisateur', 'Commentaire']
    ws.append(headers)
    rh = ws.max_row
    for col in range(1, NC + 1):
        c = ws.cell(row=rh, column=col)
        c.fill = hf; c.font = hft; c.alignment = ct; c.border = bd

    cave_fill = PatternFill("solid", fgColor="ede9fe")
    cuis_fill = PatternFill("solid", fgColor="fff7ed")
    for i, m in enumerate(mouvements, 1):
        ws.append([
            i, m['source'], m['nom'], m['type'],
            m['quantite'], m['unite'],
            m['date'].strftime('%H:%M'), m['user'], m['commentaire'],
        ])
        rw = ws.max_row
        row_fill = cave_fill if m['source'] == 'Cave' else cuis_fill
        for col in range(1, NC + 1):
            c = ws.cell(row=rw, column=col)
            c.font = bf if col in (3, 5) else nf
            c.border = bd
            c.fill = row_fill
            c.alignment = rt if col == 5 else (ct if col in (1, 2, 7) else Alignment(vertical='center'))

    ws.append([])
    tr = ws.max_row + 1
    ws.cell(row=tr, column=4, value='TOTAL').font = Font(name='Calibri', bold=True)
    ws.cell(row=tr, column=5, value=len(mouvements)).font = Font(name='Calibri', bold=True, size=11)

    for col, w in enumerate([5, 12, 28, 20, 10, 9, 8, 18, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Mouvements_{date_mvt.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def api_stats(request):
    """API temps réel — appelée toutes les 30s par le dashboard."""
    modules = get_accessible_modules(request.user)
    stats = _get_dashboard_stats(request.user, modules)
    # Retirer les objets non-sérialisables
    stats.pop('tickets_recents', None)
    return JsonResponse(stats)


@login_required
def keep_alive(request):
    """Rafraîchit la session Django (appelé par le bouton 'Je suis là')."""
    request.session.modified = True
    return JsonResponse({'ok': True})


# ═══════════════════════════════════════════════════════════════
#  RÉSUMÉ VENTES — Multi-module (Vue Direction + exports)
# ═══════════════════════════════════════════════════════════════

def _parse_resume_ventes_data(modules_filter, date_debut, date_fin):
    """Helper partagé : requête Ticket et agrégation pour le résumé ventes."""
    MODULE_LABELS = {
        'restaurant': 'Restaurant', 'cave': 'Cave / Bar',
        'hotel': 'Hôtel', 'piscine': 'Piscine',
        'espace': 'Espaces', 'autre': 'Autre',
    }
    MODE_LABELS = {
        'especes': 'Espèces', 'carte': 'Carte/TPE', 'carte_bancaire': 'Carte/TPE',
        'mobile': 'Mobile Money', 'mobile_money': 'Mobile Money',
        'orange_money': 'Orange Money', 'wave': 'Wave',
        'moov_money': 'Moov Money', 'mtn_money': 'MTN Money',
        'cheque': 'Chèque', 'virement': 'Virement', 'chambre': 'Chambre',
    }
    from facturation.models import Ticket
    qs = list(Ticket.objects.filter(
        module__in=modules_filter,
        date_creation__date__gte=date_debut,
        date_creation__date__lte=date_fin,
    ).select_related('cree_par').order_by('-date_creation'))

    # Serveur mapping: restaurant ticket → Commande.serveur
    from restaurant.models import Commande as RestaurantCommande
    rest_ids = [tk.objet_id for tk in qs if tk.module == 'restaurant' and tk.objet_id]
    serveur_map = {}
    if rest_ids:
        for cmd in RestaurantCommande.objects.filter(pk__in=rest_ids).select_related('serveur'):
            nom = (cmd.serveur.get_full_name() or cmd.serveur.username) if cmd.serveur else 'Non assigné'
            serveur_map[cmd.pk] = nom

    total_net, par_module, par_mode, par_caissier, par_serveur, tickets = 0, {}, {}, {}, {}, []
    same_day = (date_debut == date_fin)
    for tk in qs:
        montant = float(tk.montant_paye or 0)
        total_net += montant
        mod = MODULE_LABELS.get(tk.module, tk.module)
        par_module.setdefault(mod, {'nb': 0, 'total': 0})
        par_module[mod]['nb']    += 1
        par_module[mod]['total'] += montant
        m = MODE_LABELS.get(tk.mode_paiement or 'especes',
                            (tk.mode_paiement or 'especes').replace('_', ' ').capitalize())
        par_mode[m] = par_mode.get(m, 0) + montant
        caissier = (tk.cree_par.get_full_name() or tk.cree_par.username) if tk.cree_par else 'Inconnu'
        par_caissier.setdefault(caissier, {'nb': 0, 'total': 0})
        par_caissier[caissier]['nb']    += 1
        par_caissier[caissier]['total'] += montant
        if tk.module == 'restaurant':
            serveur = serveur_map.get(tk.objet_id, 'Non assigné')
            par_serveur.setdefault(serveur, {'nb': 0, 'total': 0})
            par_serveur[serveur]['nb']    += 1
            par_serveur[serveur]['total'] += montant
        tickets.append({
            'numero':   tk.numero or '—',
            'module':   mod,
            'date':     tk.date_creation,
            'heure':    tk.date_creation.strftime('%H:%M') if same_day else tk.date_creation.strftime('%d/%m %H:%M'),
            'mode':     m,
            'caissier': caissier,
            'montant':  montant,
        })

    periode = (date_debut.strftime('%d/%m/%Y') if same_day
               else f"{date_debut.strftime('%d/%m/%Y')} → {date_fin.strftime('%d/%m/%Y')}")
    return {
        'periode': periode, 'nb_tickets': len(tickets), 'total_net': total_net,
        'par_module': par_module, 'par_mode': par_mode,
        'par_caissier': par_caissier, 'par_serveur': par_serveur, 'tickets': tickets,
    }


@login_required
def resume_ventes_direction(request):
    """Page Résumé Ventes — Vue Direction (multi-module, toute période)."""
    from utils.permissions import _is_manager
    from django.contrib import messages as _msg
    if not (_is_manager(request.user) or request.user.is_superuser):
        _msg.error(request, "Accès réservé à la Direction.")
        return redirect('dashboard:index')
    today = timezone.now().date()
    return render(request, 'dashboard/resume_ventes.html', {
        'page_title': 'Résumé Ventes — Vue Direction',
        'date_debut_default': today.replace(day=1).strftime('%Y-%m-%d'),
        'date_fin_default':   today.strftime('%Y-%m-%d'),
    })


@login_required
def api_resume_ventes(request):
    """JSON : résumé des ventes multi-module sur une période."""
    from utils.permissions import _is_manager
    from datetime import datetime
    if not (_is_manager(request.user) or request.user.is_superuser):
        return JsonResponse({'success': False, 'error': 'Accès refusé'}, status=403)

    today = timezone.now().date()
    def parse_date(p, fb):
        try: return datetime.strptime(request.GET[p], '%Y-%m-%d').date()
        except: return fb

    date_debut = parse_date('date_debut', today)
    date_fin   = parse_date('date_fin', today)
    if date_fin < date_debut: date_fin = date_debut

    ALL_MODULES = ['restaurant', 'cave', 'hotel', 'piscine', 'espace', 'autre']
    mp = request.GET.get('modules', '')
    modules_filter = [m for m in mp.split(',') if m in ALL_MODULES] if mp else ALL_MODULES

    try:
        data = _parse_resume_ventes_data(modules_filter, date_debut, date_fin)
        return JsonResponse({
            'success': True,
            'periode':     data['periode'],
            'nb_tickets':  data['nb_tickets'],
            'total_net':   data['total_net'],
            'par_module':  [{'nom': k, **v} for k, v in data['par_module'].items()],
            'par_mode':    data['par_mode'],
            'par_caissier': [{'nom': k, **v} for k, v in data['par_caissier'].items()],
            'par_serveur':  [{'nom': k, **v} for k, v in data['par_serveur'].items()],
            'tickets':     [{k: v for k, v in t.items() if k != 'date'} for t in data['tickets']],
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def resume_ventes_excel(request):
    """Export Excel : résumé ventes multi-module."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from utils.permissions import _is_manager
    from datetime import datetime

    if not (_is_manager(request.user) or request.user.is_superuser):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    today = timezone.now().date()
    def parse_date(p, fb):
        try: return datetime.strptime(request.GET[p], '%Y-%m-%d').date()
        except: return fb
    date_debut = parse_date('date_debut', today)
    date_fin   = parse_date('date_fin', today)
    if date_fin < date_debut: date_fin = date_debut

    ALL_MODULES = ['restaurant', 'cave', 'hotel', 'piscine', 'espace', 'autre']
    mp = request.GET.get('modules', '')
    modules_filter = [m for m in mp.split(',') if m in ALL_MODULES] if mp else ALL_MODULES

    try:
        data = _parse_resume_ventes_data(modules_filter, date_debut, date_fin)
    except Exception:
        data = {'periode': '', 'nb_tickets': 0, 'total_net': 0,
                'par_module': {}, 'par_mode': {}, 'par_caissier': {}, 'tickets': []}

    GOLD, DARK, WHITE = 'C9A84C', '1A2535', 'FFFFFF'
    thin = Border(**{s: Side(style='thin', color='D4DCE8') for s in ['left','right','top','bottom']})

    def hf(bg=DARK, fg=WHITE, sz=10, bold=True):
        return Font(name='Calibri', bold=bold, size=sz, color=fg), PatternFill('solid', fgColor=bg)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Résumé Ventes'
    period_str = data['periode']

    def title_row(ws, row, text, cols, bg=DARK, fg=WHITE, sz=12):
        ws.merge_cells(f'A{row}:{get_column_letter(cols)}{row}')
        c = ws.cell(row, 1, text)
        c.font = Font(name='Calibri', bold=True, size=sz, color=fg)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 24

    def hdr_row(ws, row, labels, bg=DARK, fg=WHITE):
        for i, lbl in enumerate(labels, 1):
            c = ws.cell(row, i, lbl)
            c.font = Font(name='Calibri', bold=True, size=9.5, color=fg)
            c.fill = PatternFill('solid', fgColor=bg)
            c.border = thin
            c.alignment = Alignment(horizontal='center', vertical='center')

    def data_row(ws, row, values, bold_cols=()):
        for i, v in enumerate(values, 1):
            c = ws.cell(row, i, v)
            c.font = Font(name='Calibri', bold=(i in bold_cols), size=9.5)
            c.border = thin
            if row % 2 == 0:
                c.fill = PatternFill('solid', fgColor='F8F9FB')

    # ── Feuille 1 : Résumé global
    title_row(ws, 1, 'COMPLEXE HÔTELIER BEHANIAN — RÉSUMÉ DES VENTES', 5, GOLD, DARK, 13)
    title_row(ws, 2, f'Période : {period_str}   |   {data["nb_tickets"]} ticket(s)   |   Total : {int(data["total_net"]):,} F CFA'.replace(',', ' '), 5)
    r = 4
    hdr_row(ws, r, ['Module', 'Tickets', 'Total (F CFA)', '% total', '']); r += 1
    total = data['total_net']
    for mod, d in data['par_module'].items():
        pct = f"{round(d['total']/total*100, 1)}%" if total else '—'
        data_row(ws, r, [mod, d['nb'], int(d['total']), pct, ''], bold_cols=(3,)); r += 1

    r += 1
    hdr_row(ws, r, ['Mode de paiement', 'Total (F CFA)', '', '', '']); r += 1
    for mode, t in data['par_mode'].items():
        data_row(ws, r, [mode, int(t), '', '', ''], bold_cols=(2,)); r += 1

    r += 1
    hdr_row(ws, r, ['Caissier', 'Tickets', 'Total (F CFA)', '', '']); r += 1
    for caissier, d in data['par_caissier'].items():
        data_row(ws, r, [caissier, d['nb'], int(d['total']), '', ''], bold_cols=(3,)); r += 1

    for i, w in enumerate([28, 14, 18, 12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Feuille 2 : Détail tickets
    ws2 = wb.create_sheet('Détail Tickets')
    title_row(ws2, 1, f'Détail Tickets — {period_str}', 6, GOLD, DARK, 11)
    hdr_row(ws2, 2, ['N° Ticket', 'Module', 'Date / Heure', 'Mode', 'Caissier', 'Montant (F)'])
    for r, tk in enumerate(data['tickets'], 3):
        data_row(ws2, r, [
            tk['numero'], tk['module'],
            tk['date'].strftime('%d/%m/%Y %H:%M'),
            tk['mode'], tk['caissier'], int(tk['montant'])
        ], bold_cols=(6,))
    for i, w in enumerate([16, 14, 18, 16, 22, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fname = f"resume_ventes_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp


@login_required
def resume_ventes_print(request):
    """Page impression résumé ventes — modèle print_base.html."""
    from utils.permissions import _is_manager
    from django.contrib import messages as _msg
    from datetime import datetime

    if not (_is_manager(request.user) or request.user.is_superuser):
        _msg.error(request, "Accès réservé à la Direction.")
        return redirect('dashboard:index')

    today = timezone.now().date()
    def parse_date(p, fb):
        try: return datetime.strptime(request.GET[p], '%Y-%m-%d').date()
        except: return fb
    date_debut = parse_date('date_debut', today)
    date_fin   = parse_date('date_fin', today)
    if date_fin < date_debut: date_fin = date_debut

    ALL_MODULES = ['restaurant', 'cave', 'hotel', 'piscine', 'espace', 'autre']
    MODULE_LABELS = {'restaurant': 'Restaurant', 'cave': 'Cave / Bar',
                     'hotel': 'Hôtel', 'piscine': 'Piscine',
                     'espace': 'Espaces', 'autre': 'Autre'}
    mp = request.GET.get('modules', '')
    modules_filter = [m for m in mp.split(',') if m in ALL_MODULES] if mp else ALL_MODULES
    modules_display = ', '.join(MODULE_LABELS.get(m, m) for m in modules_filter)

    try:
        data = _parse_resume_ventes_data(modules_filter, date_debut, date_fin)
    except Exception as e:
        data = {'periode': '—', 'nb_tickets': 0, 'total_net': 0,
                'par_module': {}, 'par_mode': {}, 'par_caissier': {}, 'tickets': []}

    return render(request, 'dashboard/resume_ventes_print.html', {
        'date_debut':      date_debut,
        'date_fin':        date_fin,
        'modules_display': modules_display,
        'modules_param':   mp or ','.join(ALL_MODULES),
        'periode':         data['periode'],
        'nb_tickets':      data['nb_tickets'],
        'total_net':       data['total_net'],
        'par_module':      data['par_module'],
        'par_mode':        data['par_mode'],
        'par_caissier':    data['par_caissier'],
        'tickets':         data['tickets'],
        'generated_at':    timezone.now(),
    })
