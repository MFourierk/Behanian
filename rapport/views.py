# ==============================================================
# RAPPORT DE STOCK — COMPLEXE BEHANIAN
# ==============================================================
# Fichier : rapport/views.py  (ou ajouter dans dashboard/views.py)
#
# Accessible depuis :
#   /rapport/stock/              → rapport complet
#   /cuisine/rapport/stock/      → filtré sur Cuisine par défaut
#   /cave/rapport/stock/         → filtré sur Cave par défaut
# ==============================================================

from django.shortcuts import render
from django.utils import timezone
from decimal import Decimal

# Imports Cuisine
from cuisine.models import Ingredient, CategorieIngredient

# Imports Cave / Bar
from bar.models import BoissonBar, CategorieBar


def rapport_stock(request):
    """
    Rapport de stock unifié — Cuisine + Cave.
    Filtres : module, catégorie, date (snapshot), état.
    """

    # --- Paramètres GET ---
    module_filtre    = request.GET.get('module', 'tous')      # cuisine / cave / tous
    categorie_filtre = request.GET.get('categorie', '')       # pk de la catégorie
    etat_filtre      = request.GET.get('etat', 'tous')        # tous / bas / rupture
    date_rapport     = request.GET.get('date', timezone.now().date().isoformat())

    # Pour l'instant le stock est en temps réel (snapshot live).
    # Une future évolution pourra interroger MouvementStock pour reconstruire
    # le stock à une date donnée — la vue est déjà prête pour ça.

    # -------------------------------------------------------
    # CUISINE — Ingrédients
    # -------------------------------------------------------
    ingredients = []
    total_valeur_cuisine = Decimal('0')

    if module_filtre in ('cuisine', 'tous'):
        qs_cuisine = Ingredient.objects.select_related(
            'categorie', 'unite_stock', 'fournisseur_principal'
        ).filter(statut=True).order_by('categorie__ordre', 'categorie__nom', 'nom')

        if categorie_filtre and module_filtre == 'cuisine':
            qs_cuisine = qs_cuisine.filter(categorie__pk=categorie_filtre)

        for ing in qs_cuisine:
            valeur = ing.quantite_stock * ing.cmup

            if etat_filtre == 'rupture' and not ing.est_en_rupture:
                continue
            if etat_filtre == 'bas' and not (ing.est_stock_bas or ing.est_en_rupture):
                continue

            etat = 'rupture' if ing.est_en_rupture else ('bas' if ing.est_stock_bas else 'normal')
            total_valeur_cuisine += valeur

            ingredients.append({
                'module'     : 'cuisine',
                'reference'  : ing.reference or '—',
                'nom'        : ing.nom,
                'categorie'  : ing.categorie.nom if ing.categorie else 'Sans catégorie',
                'categorie_ordre': ing.categorie.ordre if ing.categorie else 999,
                'quantite'   : ing.quantite_stock,
                'unite'      : str(ing.unite_stock) if ing.unite_stock else '—',
                'cmup'       : ing.cmup,
                'valeur'     : valeur,
                'seuil'      : ing.seuil_alerte,
                'stock_max'  : ing.stock_max,
                'fournisseur': str(ing.fournisseur_principal) if ing.fournisseur_principal else '—',
                'etat'       : etat,
            })

    # -------------------------------------------------------
    # CAVE — Boissons
    # -------------------------------------------------------
    total_valeur_cave = Decimal('0')
    boissons = []

    if module_filtre in ('cave', 'tous'):
        qs_cave = BoissonBar.objects.select_related(
            'categorie'
        ).filter(statut='actif').order_by('categorie__nom', 'nom')

        if categorie_filtre and module_filtre == 'cave':
            qs_cave = qs_cave.filter(categorie__pk=categorie_filtre)

        for b in qs_cave:
            valeur = Decimal(b.quantite_stock) * b.prix_achat

            if etat_filtre == 'rupture' and not b.est_en_rupture:
                continue
            if etat_filtre == 'bas' and not (b.est_stock_bas or b.est_en_rupture):
                continue

            etat = 'rupture' if b.est_en_rupture else ('bas' if b.est_stock_bas else 'normal')
            total_valeur_cave += valeur

            boissons.append({
                'module'     : 'cave',
                'reference'  : b.reference or '—',
                'nom'        : b.nom,
                'categorie'  : b.categorie.nom if b.categorie else 'Sans catégorie',
                'categorie_ordre': 0,
                'quantite'   : b.quantite_stock,
                'unite'      : b.unite_affichee,
                'cmup'       : b.prix_achat,   # Pour la cave : prix achat = coût unitaire
                'valeur'     : valeur,
                'seuil'      : b.seuil_alerte,
                'stock_max'  : 0,
                'fournisseur': '—',
                'etat'       : etat,
            })

    # -------------------------------------------------------
    # Fusion + regroupement par module > catégorie
    # -------------------------------------------------------
    tous_articles = ingredients + boissons

    # Regroupement pour le template
    groupes = {}
    for art in tous_articles:
        cle = (art['module'], art['categorie'])
        if cle not in groupes:
            groupes[cle] = {
                'module'    : art['module'],
                'categorie' : art['categorie'],
                'articles'  : [],
                'valeur'    : Decimal('0'),
                'nb_rupture': 0,
                'nb_bas'    : 0,
            }
        groupes[cle]['articles'].append(art)
        groupes[cle]['valeur'] += art['valeur']
        if art['etat'] == 'rupture': groupes[cle]['nb_rupture'] += 1
        if art['etat'] == 'bas':     groupes[cle]['nb_bas'] += 1

    groupes_list = sorted(groupes.values(), key=lambda g: (g['module'], g['categorie']))

    # -------------------------------------------------------
    # Statistiques globales
    # -------------------------------------------------------
    total_articles  = len(tous_articles)
    nb_rupture      = sum(1 for a in tous_articles if a['etat'] == 'rupture')
    nb_bas          = sum(1 for a in tous_articles if a['etat'] == 'bas')
    nb_normal       = total_articles - nb_rupture - nb_bas
    total_valeur    = total_valeur_cuisine + total_valeur_cave

    # -------------------------------------------------------
    # Listes pour les filtres
    # -------------------------------------------------------
    categories_cuisine = CategorieIngredient.objects.order_by('ordre', 'nom')
    categories_cave    = CategorieBar.objects.order_by('nom')

    context = {
        'page_title'          : 'Rapport de Stock',

        # Données
        'groupes'             : groupes_list,

        # Stats
        'total_articles'      : total_articles,
        'nb_rupture'          : nb_rupture,
        'nb_bas'              : nb_bas,
        'nb_normal'           : nb_normal,
        'total_valeur'        : total_valeur,
        'total_valeur_cuisine': total_valeur_cuisine,
        'total_valeur_cave'   : total_valeur_cave,

        # Filtres actifs (pour pré-remplir les selects)
        'module_filtre'       : module_filtre,
        'categorie_filtre'    : categorie_filtre,
        'etat_filtre'         : etat_filtre,
        'date_rapport'        : date_rapport,

        # Listes filtres
        'categories_cuisine'  : categories_cuisine,
        'categories_cave'     : categories_cave,

        # Date/heure du rapport
        'generated_at'        : timezone.now(),
    }

    return render(request, 'rapport/stock.html', context)


def _calcul_marges(date_rapport):
    """Calcule les données de marges pour une date — partagé entre vue et impression."""
    from django.db.models import Sum
    from facturation.models import Ticket
    from restaurant.models import Commande, LigneCommande

    MODULES = [
        {'key': 'restaurant', 'label': 'Restaurant',            'icon': '🍽️', 'color': '#c0392b', 'css': 'restaurant'},
        {'key': 'cave',       'label': 'Cave & Bar',            'icon': '🍷', 'color': '#1a5276', 'css': 'cave'},
        {'key': 'hotel',      'label': 'Hôtel',                 'icon': '🏨', 'color': '#27ae60', 'css': 'hotel'},
        {'key': 'piscine',    'label': 'Piscine',               'icon': '🏊', 'color': '#2980b9', 'css': 'piscine'},
        {'key': 'espace',     'label': 'Espaces Événementiels', 'icon': '🎪', 'color': '#8e44ad', 'css': 'espace'},
    ]

    modules_data = []
    total_ca   = Decimal('0')
    total_cout = Decimal('0')
    has_any_cout = False

    for mod in MODULES:
        tickets = Ticket.objects.filter(module=mod['key'], date_creation__date=date_rapport)
        ca = tickets.aggregate(s=Sum('montant_paye'))['s'] or Decimal('0')
        nb_tickets = tickets.count()

        cout = Decimal('0')
        cout_disponible = False
        nb_lignes_avec_cout = 0
        nb_lignes_total = 0

        if mod['key'] == 'restaurant':
            try:
                from cuisine.models import Plat as PlatCuisine
                commandes = Commande.objects.filter(
                    statut='payee', date_creation__date=date_rapport
                ).prefetch_related('lignes__plat', 'lignes__boisson')
                plat_cache = {}
                for cmd in commandes:
                    for ligne in cmd.lignes.all():
                        nb_lignes_total += 1
                        if ligne.plat and ligne.plat.cuisine_plat_id:
                            pid = ligne.plat.cuisine_plat_id
                            if pid not in plat_cache:
                                try:
                                    pc = PlatCuisine.objects.select_related('fiche_technique').get(pk=pid)
                                    plat_cache[pid] = pc
                                except PlatCuisine.DoesNotExist:
                                    plat_cache[pid] = None
                            pc = plat_cache.get(pid)
                            if pc and getattr(pc, 'fiche_technique', None):
                                cout += pc.fiche_technique.cout_par_portion * ligne.quantite
                                nb_lignes_avec_cout += 1
                        if ligne.boisson and ligne.boisson.prix_achat:
                            cout += Decimal(str(ligne.boisson.prix_achat)) * ligne.quantite
                            nb_lignes_avec_cout += 1
                cout_disponible = nb_lignes_avec_cout > 0
            except Exception:
                pass

        elif mod['key'] == 'cave':
            try:
                lignes = LigneCommande.objects.filter(
                    commande__date_creation__date=date_rapport,
                    boisson__isnull=False
                ).select_related('boisson')
                nb_lignes_total = lignes.count()
                for l in lignes:
                    if l.boisson.prix_achat:
                        cout += Decimal(str(l.boisson.prix_achat)) * l.quantite
                        nb_lignes_avec_cout += 1
                cout_disponible = nb_lignes_avec_cout > 0
            except Exception:
                pass

        marge_brute = ca - cout if cout_disponible else None
        taux_marge  = float(marge_brute / ca * 100) if (marge_brute is not None and ca > 0) else None
        coeff_mult  = float(ca / cout) if (cout_disponible and cout > 0) else None

        total_ca += ca
        if cout_disponible:
            total_cout += cout
            has_any_cout = True

        modules_data.append({
            **mod,
            'ca': ca, 'nb_tickets': nb_tickets,
            'cout': cout if cout_disponible else None,
            'cout_disponible': cout_disponible,
            'marge_brute': marge_brute,
            'taux_marge': taux_marge,
            'coeff_mult': coeff_mult,
            'nb_lignes_cout': nb_lignes_avec_cout,
            'nb_lignes_total': nb_lignes_total,
        })

    total_marge       = total_ca - total_cout if has_any_cout else None
    taux_marge_global = float(total_marge / total_ca * 100) if (total_marge and total_ca > 0) else None

    return {
        'modules_data': modules_data,
        'total_ca': total_ca,
        'total_cout': total_cout if has_any_cout else None,
        'total_marge': total_marge,
        'taux_marge_global': taux_marge_global,
    }


def _parse_date(request, param='date'):
    from datetime import date as date_type
    date_str = request.GET.get(param, timezone.now().date().isoformat())
    try:
        return date_type.fromisoformat(date_str)
    except ValueError:
        return timezone.now().date()


def rapport_marges(request):
    """Rapport Marges par Module — vue dashboard avec filtre date."""
    date_rapport = _parse_date(request)
    data = _calcul_marges(date_rapport)
    return render(request, 'rapport/marges.html', {
        'date_rapport': date_rapport,
        'generated_at': timezone.now(),
        **data,
    })


def rapport_marges_print(request):
    """Page d'impression Marges par Module — modèle print_base.html."""
    date_rapport = _parse_date(request)
    data = _calcul_marges(date_rapport)
    return render(request, 'rapport/marges_print.html', {
        'date_rapport': date_rapport,
        'generated_at': timezone.now(),
        **data,
    })


def rapport_marges_excel(request):
    """Export Excel : Rapport Marges par Module."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    date_rapport = _parse_date(request)
    data = _calcul_marges(date_rapport)
    modules_data = data['modules_data']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marges par Module"

    hf  = PatternFill("solid", fgColor="1a2535")
    hft = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    tf  = Font(name='Calibri', bold=True, size=14, color='1a2535')
    bf  = Font(name='Calibri', bold=True, size=10)
    nf  = Font(name='Calibri', size=10)
    th  = Side(border_style="thin", color="d4dce8")
    bd  = Border(left=th, right=th, top=th, bottom=th)
    ct  = Alignment(horizontal='center', vertical='center')
    rt  = Alignment(horizontal='right', vertical='center')

    NC = 8
    ws.merge_cells(f'A1:{get_column_letter(NC)}1')
    ws['A1'] = f"RAPPORT MARGES PAR MODULE — {date_rapport.strftime('%d/%m/%Y').upper()} — COMPLEXE BEHANIAN"
    ws['A1'].font = tf; ws['A1'].alignment = ct

    ws.merge_cells(f'A2:{get_column_letter(NC)}2')
    ws['A2'] = f"Édité le {timezone.now().strftime('%d/%m/%Y à %H:%M')} — {len(modules_data)} module(s) analysé(s)"
    ws['A2'].font = Font(name='Calibri', size=10, color='7a8b9c', italic=True)
    ws['A2'].alignment = ct

    ws.append([])
    headers = ['Module', 'Tickets', 'CA Encaissé (F)', 'Coût Estimé (F)', 'Marge Brute (F)', 'Taux Marge %', 'Coefficient ×', 'Remarque']
    ws.append(headers)
    rh = ws.max_row
    for col in range(1, NC + 1):
        c = ws.cell(row=rh, column=col)
        c.fill = hf; c.font = hft; c.alignment = ct; c.border = bd

    for mod in modules_data:
        cout_val  = float(mod['cout']) if mod['cout'] is not None else None
        marge_val = float(mod['marge_brute']) if mod['marge_brute'] is not None else None
        taux_val  = mod['taux_marge']
        coeff_val = mod['coeff_mult']
        if mod['ca'] == 0:
            remarque = 'Aucune vente'
        elif not mod['cout_disponible']:
            remarque = 'Service pur — pas de coût matière'
        else:
            remarque = 'Coût partiel' if mod['nb_lignes_cout'] < mod['nb_lignes_total'] else 'Coût complet'

        ws.append([
            mod['label'],
            mod['nb_tickets'],
            float(mod['ca']),
            cout_val,
            marge_val,
            taux_val,
            coeff_val,
            remarque,
        ])
        rw = ws.max_row
        for col in range(1, NC + 1):
            c = ws.cell(row=rw, column=col)
            c.font = bf if col in (1, 3, 5) else nf
            c.border = bd
            c.alignment = rt if col in (2, 3, 4, 5, 6, 7) else Alignment(vertical='center')
        for col in (3, 4, 5):
            ws.cell(row=rw, column=col).number_format = '#,##0'
        ws.cell(row=rw, column=6).number_format = '0.00'
        ws.cell(row=rw, column=7).number_format = '0.00'

    ws.append([])
    tr = ws.max_row + 1
    ws.cell(row=tr, column=1, value='TOTAL GÉNÉRAL').font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=tr, column=3, value=float(data['total_ca'])).font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=tr, column=3).number_format = '#,##0'
    if data['total_cout'] is not None:
        ws.cell(row=tr, column=4, value=float(data['total_cout'])).font = Font(name='Calibri', bold=True, size=11, color='c0392b')
        ws.cell(row=tr, column=4).number_format = '#,##0'
    if data['total_marge'] is not None:
        ws.cell(row=tr, column=5, value=float(data['total_marge'])).font = Font(name='Calibri', bold=True, size=11, color='16a34a')
        ws.cell(row=tr, column=5).number_format = '#,##0'
    if data['taux_marge_global'] is not None:
        ws.cell(row=tr, column=6, value=data['taux_marge_global']).font = Font(name='Calibri', bold=True, size=11, color='c9a84c')
        ws.cell(row=tr, column=6).number_format = '0.00'

    for col, w in enumerate([24, 10, 18, 18, 18, 14, 14, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Marges_{date_rapport.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response
