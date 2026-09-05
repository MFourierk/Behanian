from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum, Q, Count
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import json
from decimal import Decimal, InvalidOperation

from utils.permissions import require_module_access, require_manager, GROUPE_MANAGER_GENERAL
from facturation.models import Ticket
from .models import CaisseSession, MouvementCaisse, PrelevementBanque


_MODE_LABELS = {
    'especes':       ('💵 Espèces',      '#059669'),
    'wave':          ('📱 Wave',          '#1d4ed8'),
    'orange_money':  ('📱 Orange Money',  '#c2410c'),
    'mtn_money':     ('📱 MTN MoMo',     '#854d0e'),
    'moov_money':    ('📱 Moov Money',   '#0f766e'),
    'mobile_money':  ('📱 Mobile Money', '#7c3aed'),
    'carte_bancaire':('💳 Carte',         '#7c3aed'),
    'virement':      ('🏦 Virement',      '#d97706'),
    'cheque':        ('📃 Chèque',        '#475569'),
}

_TYPE_CONFIG = [
    # (type_key, label, is_entree)
    ('fond_caisse',  'Fond de caisse (ouverture)',  True),
    ('versement',    'Versements reçus',             True),
    ('encaissement', 'Encaissements divers',         True),
    ('ajustement',   'Ajustements',                  True),
    ('depense',      'Dépenses / Décaissements',     False),
    ('prelevement',  'Prélèvements banque',          False),
    ('remboursement','Remboursements clients',        False),
]


def get_caisse_flux(qs):
    """Résume les MouvementCaisse par type et mode de paiement.
    Retourne entrees, sorties, totaux pour le panneau Flux de caisse.
    """
    # Agrégation par type + mode
    rows = list(qs.values('type', 'mode_paiement').annotate(s=Sum('montant')))
    raw = {}
    for r in rows:
        t = r['type']
        m = r['mode_paiement']
        s = int(r['s'] or 0)
        if t not in raw:
            raw[t] = {}
        raw[t][m] = raw[t].get(m, 0) + s

    # Détail individuel des versements pour audit
    detail_versements = list(
        qs.filter(type='versement')
          .values('id', 'date', 'montant', 'mode_paiement', 'module', 'description')
          .order_by('date')
    )
    for dv in detail_versements:
        dv['montant'] = int(dv['montant'])
        m = dv['mode_paiement']
        dv['mode_lbl'] = _MODE_LABELS[m][0] if m in _MODE_LABELS else m

    entrees, sorties = [], []
    total_entrees = total_sorties = 0

    for type_key, label, is_entree in _TYPE_CONFIG:
        modes_data = raw.get(type_key, {})
        total = sum(modes_data.values())
        if total == 0:
            continue
        par_mode = [
            {'lbl': _MODE_LABELS[m][0], 'clr': _MODE_LABELS[m][1], 'mt': v}
            for m, v in sorted(modes_data.items(), key=lambda x: -x[1])
            if m in _MODE_LABELS and v > 0
        ]
        entry = {'label': label, 'total': total, 'par_mode': par_mode}
        if type_key == 'versement':
            entry['detail'] = detail_versements
        if is_entree:
            entrees.append(entry)
            total_entrees += total
        else:
            sorties.append(entry)
            total_sorties += total

    return {
        'entrees':       entrees,
        'sorties':       sorties,
        'total_entrees': total_entrees,
        'total_sorties': total_sorties,
        'net':           total_entrees - total_sorties,
    }


# ── Modules à réconcilier (ticket_module, caisse_module, label, emoji) ─────
MODULES_RECONCILIATION = [
    ('hotel',      'hotel',    'Hôtel',        '🏨'),
    ('restaurant', 'restaurant','Restaurant',  '🍽️'),
    ('cave',       'cave',     'Cave / Bar',   '🍷'),
    ('piscine',    'piscine',  'Piscine',      '🏊'),
    ('espace',     'espaces',  'Espaces',      '🎪'),
]


def get_reconciliation_jour(date=None):
    """
    Retourne par module : total transactions du jour, total versé manuellement,
    solde restant à verser.
    """
    if date is None:
        date = timezone.localdate()

    lignes = []
    grand_total_tx    = 0
    grand_total_verse = 0
    grand_especes     = 0
    grand_mobile      = 0
    grand_wave        = 0
    grand_orange      = 0
    grand_mtn         = 0
    grand_moov        = 0
    grand_carte       = 0
    grand_virement    = 0
    grand_mixte       = 0

    for ticket_mod, caisse_mod, label, emoji in MODULES_RECONCILIATION:
        qs = Ticket.objects.filter(date_creation__date=date, module=ticket_mod)
        total_tx = int(qs.aggregate(s=Sum('montant_total'))['s'] or 0)

        def _sum(modes):
            return int(qs.filter(mode_paiement__in=modes).aggregate(s=Sum('montant_total'))['s'] or 0)

        especes  = _sum(['especes'])
        wave     = _sum(['wave'])
        orange   = _sum(['orange_money'])
        mtn      = _sum(['mtn_money'])
        moov     = _sum(['moov_money'])
        mobile   = wave + orange + mtn + moov + _sum(['mobile_money', 'mobile'])
        carte    = _sum(['carte_bancaire', 'carte'])
        virement = _sum(['virement'])
        mixte    = _sum(['mixte'])
        autres   = total_tx - especes - mobile - carte - virement - mixte

        vs_qs = MouvementCaisse.objects.filter(
            date__date=date,
            type='versement',
            module=caisse_mod,
            valide=True,
        ).exclude(reference__startswith='CONSOLIDATION')

        def _vsum(modes):
            return int(vs_qs.filter(mode_paiement__in=modes).aggregate(s=Sum('montant'))['s'] or 0)

        verse_especes  = _vsum(['especes'])
        verse_wave     = _vsum(['wave'])
        verse_orange   = _vsum(['orange_money'])
        verse_mtn      = _vsum(['mtn_money'])
        verse_moov     = _vsum(['moov_money'])
        verse_mobile   = verse_wave + verse_orange + verse_mtn + verse_moov + _vsum(['mobile_money', 'mobile'])
        verse_carte    = _vsum(['carte_bancaire', 'carte'])
        verse_virement = _vsum(['virement'])
        total_verse    = int(vs_qs.aggregate(s=Sum('montant'))['s'] or 0)

        solde = total_tx - total_verse
        lignes.append({
            'label':          label,
            'emoji':          emoji,
            'total_tx':       total_tx,
            'especes':        especes,
            'mobile':         mobile,
            'wave':           wave,
            'orange':         orange,
            'mtn':            mtn,
            'moov':           moov,
            'carte':          carte,
            'virement':       virement,
            'mixte':          mixte,
            'autres':         autres if autres > 0 else 0,
            'total_verse':    total_verse,
            'verse_especes':  verse_especes,
            'verse_wave':     verse_wave,
            'verse_orange':   verse_orange,
            'verse_mtn':      verse_mtn,
            'verse_moov':     verse_moov,
            'verse_mobile':   verse_mobile,
            'verse_carte':    verse_carte,
            'verse_virement': verse_virement,
            'solde':          solde,
            'complet':        solde <= 0,
        })
        grand_total_tx    += total_tx
        grand_total_verse += total_verse
        grand_especes     += especes
        grand_mobile      += mobile
        grand_wave        += wave
        grand_orange      += orange
        grand_mtn         += mtn
        grand_moov        += moov
        grand_carte       += carte
        grand_virement    += virement
        grand_mixte       += mixte

    return {
        'lignes':             lignes,
        'grand_total_tx':     grand_total_tx,
        'grand_total_verse':  grand_total_verse,
        'grand_solde':        grand_total_tx - grand_total_verse,
        'grand_especes':      grand_especes,
        'grand_mobile':       grand_mobile,
        'grand_wave':         grand_wave,
        'grand_orange':       grand_orange,
        'grand_mtn':          grand_mtn,
        'grand_moov':         grand_moov,
        'grand_carte':        grand_carte,
        'grand_virement':     grand_virement,
        'grand_mixte':        grand_mixte,
    }


# ── Helpers ────────────────────────────────────────────────────────────────

def _dec(val, default=0):
    try:
        return Decimal(str(val or default))
    except (InvalidOperation, TypeError):
        return Decimal(str(default))


def get_stats_jour(date=None, type_caisse=None):
    """Stats complètes d'une journée.
    - type_caisse=None ou 'centrale' : toutes les transactions
    - type_caisse='hotel'            : tickets hotel uniquement
    - type_caisse='module'           : tickets hors hotel
    """
    if date is None:
        date = timezone.now().date()

    tickets = Ticket.objects.filter(date_creation__date=date)

    if type_caisse == 'hotel':
        tickets = tickets.filter(module__in=['hotel'])
    elif type_caisse == 'module':
        tickets = tickets.exclude(module__in=['hotel'])

    total        = tickets.aggregate(s=Sum('montant_total'))['s'] or 0
    especes      = tickets.filter(mode_paiement='especes').aggregate(s=Sum('montant_total'))['s'] or 0
    carte        = tickets.filter(mode_paiement__in=['carte_bancaire', 'carte']).aggregate(s=Sum('montant_total'))['s'] or 0
    virement     = tickets.filter(mode_paiement='virement').aggregate(s=Sum('montant_total'))['s'] or 0

    # Mobile money — détail par opérateur
    _wave        = int(tickets.filter(mode_paiement='wave').aggregate(s=Sum('montant_total'))['s'] or 0)
    _orange      = int(tickets.filter(mode_paiement='orange_money').aggregate(s=Sum('montant_total'))['s'] or 0)
    _mtn         = int(tickets.filter(mode_paiement='mtn_money').aggregate(s=Sum('montant_total'))['s'] or 0)
    _moov        = int(tickets.filter(mode_paiement='moov_money').aggregate(s=Sum('montant_total'))['s'] or 0)
    _mobile_gen  = int(tickets.filter(mode_paiement__in=['mobile', 'mobile_money']).aggregate(s=Sum('montant_total'))['s'] or 0)
    mobile       = _wave + _orange + _mtn + _moov + _mobile_gen

    par_mobile = []
    if _wave:       par_mobile.append(('Wave',             _wave,   '#1d4ed8', 'W'))
    if _orange:     par_mobile.append(('Orange Money',     _orange, '#c2410c', 'O'))
    if _mtn:        par_mobile.append(('MTN Mobile Money', _mtn,    '#854d0e', 'M'))
    if _moov:       par_mobile.append(('Moov Money',       _moov,   '#0f766e', 'V'))
    if _mobile_gen: par_mobile.append(('Mobile Money',     _mobile_gen, '#7c3aed', '📱'))

    par_module = {}
    for mod, label in [('hotel', 'Hôtel'), ('restaurant', 'Restaurant'), ('cave', 'Cave'),
                        ('piscine', 'Piscine'), ('espace', 'Espaces'), ('caisse', 'Caisse')]:
        t = tickets.filter(module__startswith=mod).aggregate(s=Sum('montant_total'))['s'] or 0
        if t:
            par_module[label] = int(t)

    prelevements   = PrelevementBanque.objects.filter(date__date=date, valide=True)
    total_prelev   = prelevements.aggregate(s=Sum('montant'))['s'] or 0
    depenses       = MouvementCaisse.objects.filter(date__date=date, type='depense', valide=True)
    total_depenses = depenses.aggregate(s=Sum('montant'))['s'] or 0

    return {
        'date':         date,
        'total':        int(total),
        'nb_tickets':   tickets.count(),
        'especes':      int(especes),
        'mobile':       mobile,       # total agrégé (backward compat)
        'par_mobile':   par_mobile,   # [(label, montant, couleur, badge), ...]
        'carte':        int(carte),
        'virement':     int(virement),
        'par_module':   par_module,
        'prelevements': int(total_prelev),
        'depenses':     int(total_depenses),
        'net':          int(total) - int(total_prelev) - int(total_depenses),
        'tickets':      tickets.select_related('client', 'cree_par').order_by('-date_creation'),
    }


def get_stats_session(session):
    """Stats des tickets encaissés pendant la fenêtre horaire d'une session.
    Filtre par opened_at → closed_at (ou now() si session encore ouverte).
    Gère les shifts à cheval sur minuit naturellement.
    Retourne la même structure que get_stats_jour pour compatibilité.
    """
    date_fin = session.closed_at or timezone.now()
    tickets  = Ticket.objects.filter(
        date_creation__gte=session.opened_at,
        date_creation__lt=date_fin,
    )

    total    = tickets.aggregate(s=Sum('montant_total'))['s'] or 0
    especes  = tickets.filter(mode_paiement='especes').aggregate(s=Sum('montant_total'))['s'] or 0
    carte    = tickets.filter(mode_paiement__in=['carte_bancaire', 'carte']).aggregate(s=Sum('montant_total'))['s'] or 0
    virement = tickets.filter(mode_paiement='virement').aggregate(s=Sum('montant_total'))['s'] or 0

    _wave       = int(tickets.filter(mode_paiement='wave').aggregate(s=Sum('montant_total'))['s'] or 0)
    _orange     = int(tickets.filter(mode_paiement='orange_money').aggregate(s=Sum('montant_total'))['s'] or 0)
    _mtn        = int(tickets.filter(mode_paiement='mtn_money').aggregate(s=Sum('montant_total'))['s'] or 0)
    _moov       = int(tickets.filter(mode_paiement='moov_money').aggregate(s=Sum('montant_total'))['s'] or 0)
    _mobile_gen = int(tickets.filter(mode_paiement__in=['mobile', 'mobile_money']).aggregate(s=Sum('montant_total'))['s'] or 0)
    mobile      = _wave + _orange + _mtn + _moov + _mobile_gen

    par_mobile = []
    if _wave:       par_mobile.append(('Wave',             _wave,       '#1d4ed8', 'W'))
    if _orange:     par_mobile.append(('Orange Money',     _orange,     '#c2410c', 'O'))
    if _mtn:        par_mobile.append(('MTN Mobile Money', _mtn,        '#854d0e', 'M'))
    if _moov:       par_mobile.append(('Moov Money',       _moov,       '#0f766e', 'V'))
    if _mobile_gen: par_mobile.append(('Mobile Money',     _mobile_gen, '#7c3aed', '📱'))

    par_module = {}
    for mod, label in [('hotel', 'Hôtel'), ('restaurant', 'Restaurant'), ('cave', 'Cave'),
                        ('piscine', 'Piscine'), ('espace', 'Espaces'), ('caisse', 'Caisse')]:
        t = tickets.filter(module__startswith=mod).aggregate(s=Sum('montant_total'))['s'] or 0
        if t:
            par_module[label] = int(t)

    total_prelev   = MouvementCaisse.objects.filter(session=session, type='prelevement_banque', valide=True).aggregate(s=Sum('montant'))['s'] or 0
    total_depenses = MouvementCaisse.objects.filter(session=session, type='depense', valide=True).aggregate(s=Sum('montant'))['s'] or 0

    return {
        'date':         session.opened_at.date(),
        'total':        int(total),
        'nb_tickets':   tickets.count(),
        'especes':      int(especes),
        'mobile':       mobile,
        'par_mobile':   par_mobile,
        'carte':        int(carte),
        'virement':     int(virement),
        'par_module':   par_module,
        'prelevements': int(total_prelev),
        'depenses':     int(total_depenses),
        'net':          int(total) - int(total_prelev) - int(total_depenses),
        'tickets':      tickets.select_related('client', 'cree_par').order_by('-date_creation'),
    }


def get_reconciliation_session(session):
    """Réconciliation tickets vs versements pour la fenêtre horaire d'une session.
    - Tickets : filtrés par opened_at → closed_at (ou now())
    - Versements : filtrés par session FK (MouvementCaisse.session)
    """
    date_fin = session.closed_at or timezone.now()

    lignes            = []
    grand_total_tx    = 0
    grand_total_verse = 0
    grand_especes     = 0
    grand_mobile      = 0
    grand_wave        = 0
    grand_orange      = 0
    grand_mtn         = 0
    grand_moov        = 0
    grand_carte       = 0
    grand_virement    = 0
    grand_mixte       = 0

    for ticket_mod, caisse_mod, label, emoji in MODULES_RECONCILIATION:
        qs = Ticket.objects.filter(
            date_creation__gte=session.opened_at,
            date_creation__lt=date_fin,
            module=ticket_mod,
        )
        total_tx = int(qs.aggregate(s=Sum('montant_total'))['s'] or 0)

        def _sum(modes):
            return int(qs.filter(mode_paiement__in=modes).aggregate(s=Sum('montant_total'))['s'] or 0)

        especes  = _sum(['especes'])
        wave     = _sum(['wave'])
        orange   = _sum(['orange_money'])
        mtn      = _sum(['mtn_money'])
        moov     = _sum(['moov_money'])
        mobile   = wave + orange + mtn + moov + _sum(['mobile_money', 'mobile'])
        carte    = _sum(['carte_bancaire', 'carte'])
        virement = _sum(['virement'])
        mixte    = _sum(['mixte'])

        vs_qs = MouvementCaisse.objects.filter(
            session=session,
            type='versement',
            module=caisse_mod,
            valide=True,
        ).exclude(reference__startswith='CONSOLIDATION')

        def _vsum(modes):
            return int(vs_qs.filter(mode_paiement__in=modes).aggregate(s=Sum('montant'))['s'] or 0)

        verse_especes  = _vsum(['especes'])
        verse_wave     = _vsum(['wave'])
        verse_orange   = _vsum(['orange_money'])
        verse_mtn      = _vsum(['mtn_money'])
        verse_moov     = _vsum(['moov_money'])
        verse_mobile   = verse_wave + verse_orange + verse_mtn + verse_moov + _vsum(['mobile_money', 'mobile'])
        verse_carte    = _vsum(['carte_bancaire', 'carte'])
        verse_virement = _vsum(['virement'])
        total_verse    = int(vs_qs.aggregate(s=Sum('montant'))['s'] or 0)

        solde = total_tx - total_verse
        lignes.append({
            'label': label, 'emoji': emoji,
            'total_tx': total_tx, 'especes': especes,
            'mobile': mobile, 'wave': wave, 'orange': orange, 'mtn': mtn, 'moov': moov,
            'carte': carte, 'virement': virement, 'mixte': mixte,
            'total_verse':    total_verse,
            'verse_especes':  verse_especes,
            'verse_wave':     verse_wave,
            'verse_orange':   verse_orange,
            'verse_mtn':      verse_mtn,
            'verse_moov':     verse_moov,
            'verse_mobile':   verse_mobile,
            'verse_carte':    verse_carte,
            'verse_virement': verse_virement,
            'solde': solde, 'complet': solde <= 0,
        })
        grand_total_tx    += total_tx
        grand_total_verse += total_verse
        grand_especes     += especes
        grand_mobile      += mobile
        grand_wave        += wave
        grand_orange      += orange
        grand_mtn         += mtn
        grand_moov        += moov
        grand_carte       += carte
        grand_virement    += virement
        grand_mixte       += mixte

    return {
        'lignes':            lignes,
        'grand_total_tx':    grand_total_tx,
        'grand_total_verse': grand_total_verse,
        'grand_solde':       grand_total_tx - grand_total_verse,
        'grand_especes':     grand_especes,
        'grand_mobile':      grand_mobile,
        'grand_wave':        grand_wave,
        'grand_orange':      grand_orange,
        'grand_mtn':         grand_mtn,
        'grand_moov':        grand_moov,
        'grand_carte':       grand_carte,
        'grand_virement':    grand_virement,
        'grand_mixte':       grand_mixte,
    }


def get_solde_veille():
    """Retourne le fond_caisse_reel de la dernière clôture (cash physiquement compté = solde reporté)."""
    last = CaisseSession.objects.filter(is_open=False, type_caisse='centrale').order_by('-closed_at').first()
    if not last:
        return 0, None
    return int(last.fond_caisse_reel), last


def _session_centrale_non_cloturee():
    """Retourne la session centrale ouverte d'un jour antérieur, ou None."""
    today = timezone.localdate()
    return CaisseSession.objects.filter(
        is_open=True,
        type_caisse='centrale',
        date_session__lt=today,
    ).order_by('-date_session').first()


# ── Vues principales ───────────────────────────────────────────────────────

@require_module_access('caisse')
def index(request):
    today = timezone.localdate()
    from utils.permissions import _is_manager as _chk_manager
    is_manager = _chk_manager(request.user)
    # Session d'aujourd'hui (ou session oubliée d'une journée précédente)
    session_oubliee = CaisseSession.objects.filter(
        user=request.user, is_open=True, date_session__lt=today
    ).order_by('-date_session').first()
    session_active = CaisseSession.objects.filter(user=request.user, is_open=True, date_session=today).first()
    session_ouverte_par = None if (session_active or session_oubliee) else CaisseSession.objects.filter(is_open=True, date_session=today).select_related('user').first()

    sessions_jour = CaisseSession.objects.filter(
        opened_at__date=today
    ).select_related('user').order_by('opened_at')

    # Manager peut filtrer sur une session précise via ?session_id=X
    session_filtre = None
    if is_manager and not session_active:
        sid = request.GET.get('session_id')
        if sid:
            session_filtre = CaisseSession.objects.filter(pk=sid, opened_at__date=today).select_related('user').first()

    if session_active:
        # Stats limitées au shift de la caissière active (fenêtre horaire)
        stats = get_stats_session(session_active)
        attente_session = False
    elif is_manager:
        if session_filtre:
            stats = get_stats_session(session_filtre)
        else:
            # Vue journée complète (tous shifts confondus)
            stats = get_stats_jour(today, type_caisse=None)
        attente_session = False
    else:
        # Caissière sans session ouverte : zéros — ne pas montrer les shifts des collègues
        stats = {
            'date': today, 'total': 0, 'nb_tickets': 0,
            'especes': 0, 'mobile': 0, 'par_mobile': [],
            'carte': 0, 'virement': 0, 'par_module': {},
            'prelevements': 0, 'depenses': 0, 'net': 0,
            'tickets': Ticket.objects.none(),
        }
        attente_session = True

    # Mouvements, prélèvements, réconciliation — isolés par scope
    if session_active:
        date_fin_session = session_active.closed_at or timezone.now()
        mouvements = MouvementCaisse.objects.filter(
            session=session_active, valide=True
        ).select_related('cree_par').order_by('-date')
        prelevements = PrelevementBanque.objects.filter(
            date__gte=session_active.opened_at,
            date__lt=date_fin_session,
            valide=True,
        ).select_related('cree_par').order_by('-date')
        reconciliation = get_reconciliation_session(session_active)
        vue_session = True
    elif is_manager:
        if session_filtre:
            mouvements = MouvementCaisse.objects.filter(
                session=session_filtre, valide=True
            ).select_related('cree_par').order_by('-date')
            date_fin_sf = session_filtre.closed_at or timezone.now()
            prelevements = PrelevementBanque.objects.filter(
                date__gte=session_filtre.opened_at,
                date__lt=date_fin_sf,
                valide=True,
            ).select_related('cree_par').order_by('-date')
            reconciliation = get_reconciliation_session(session_filtre)
        else:
            mouvements = MouvementCaisse.objects.filter(
                date__date=today, valide=True
            ).select_related('cree_par').order_by('-date')
            prelevements = PrelevementBanque.objects.filter(
                date__date=today, valide=True
            ).select_related('cree_par').order_by('-date')
            reconciliation = get_reconciliation_jour(today)
        vue_session = False
    else:
        # Caissière sans session : aucune donnée visible
        mouvements    = MouvementCaisse.objects.none()
        prelevements  = PrelevementBanque.objects.none()
        reconciliation = None
        vue_session = False

    solde_veille, last_session = get_solde_veille()

    # Sessions centrales non clôturées des jours précédents (alerte manager)
    sessions_bloquantes = CaisseSession.objects.filter(
        is_open=True,
        type_caisse='centrale',
        date_session__lt=today,
    ).select_related('user').order_by('-date_session')

    can_open_caisse = (
        request.user.is_superuser or
        any(g in list(request.user.groups.values_list('name', flat=True)) for g in [
            'Chef caissier(e)', 'Caissier(ère) Principal(e)', 'Caissier(ere) Principal(e)',
            'Manager Général(e)', 'Manager General(e)',
            'Réceptionniste', 'Receptionniste', 'Responsable Hôtel',
            'Caissière / Caissier', 'Caissiere / Caissier',
        ])
    )

    # Flux de caisse (mouvements enregistrés par la caissière, par type et mode)
    # Exclure les versements CONSOLIDATION (synthétiques, mode especes en dur, non représentatifs)
    caisse_flux = get_caisse_flux(mouvements.exclude(reference__startswith='CONSOLIDATION'))

    # Net disponible en caisse centrale = fond initial + versements reçus - prélèvements - dépenses
    # verse_recu = tous les versements de la session (quel que soit le module enregistré),
    # hors CONSOLIDATION — cohérent avec le Flux de caisse.
    fond_initial = int(session_active.fond_caisse) if session_active else 0
    verse_recu   = int(
        mouvements.filter(type='versement')
                  .exclude(reference__startswith='CONSOLIDATION')
                  .aggregate(s=Sum('montant'))['s'] or 0
    )
    net_caisse_central = fond_initial + verse_recu - stats['prelevements'] - stats['depenses']

    context = {
        'billetage_vals': [10000, 5000, 2000, 1000, 500, 250, 200, 100, 50, 25, 10, 5],
        'today': today,
        'session_active': session_active,
        'is_manager': is_manager,
        'can_open_caisse': can_open_caisse,
        'stats': stats,
        'sessions_jour': sessions_jour,
        'mouvements': mouvements,
        'prelevements': prelevements,
        'solde_veille': solde_veille,
        'last_session': last_session,
        'sessions_bloquantes': sessions_bloquantes,
        'reconciliation': reconciliation,
        'session_ouverte_par': session_ouverte_par,
        'session_oubliee': session_oubliee,
        'vue_session': vue_session,
        'attente_session': attente_session,
        'net_caisse_central': net_caisse_central,
        'caisse_flux': caisse_flux,
        'session_filtre': session_filtre,
    }
    return render(request, 'caisse/index.html', context)


# Groupes autorisés à ouvrir la caisse centrale
GROUPES_CAISSE_CENTRALE = [
    'Chef caissier(e)',
    'Caissier(ère) Principal(e)',
    'Caissier(ere) Principal(e)',
    'Manager Général(e)',
    'Manager General(e)',
]


def _get_type_caisse(user):
    """Seule la caisse centrale est ouverte/clôturée. Tous les utilisateurs autorisés ouvrent la centrale."""
    return 'centrale'


@require_module_access('caisse')
@require_POST
def ouvrir_caisse(request):
    today = timezone.localdate()
    type_attendu = _get_type_caisse(request.user)

    # 1a. Vérifier si cet utilisateur a déjà une session ouverte
    session_existante = CaisseSession.objects.filter(
        user=request.user, is_open=True, type_caisse=type_attendu
    ).first()
    if session_existante:
        return JsonResponse({
            'success': False,
            'error': f'Votre caisse {session_existante.get_type_caisse_display()} est déjà ouverte (depuis {session_existante.opened_at.strftime("%H:%M")})',
        })

    # 1b-bis. Bloquer si l'utilisateur a déjà clôturé une session aujourd'hui
    # sans qu'un manager ait accordé une réouverture
    session_cloturee_aujourd_hui = CaisseSession.objects.filter(
        user=request.user, is_open=False, date_session=today
    ).order_by('-closed_at').first()
    if session_cloturee_aujourd_hui and not session_cloturee_aujourd_hui.reouverture_autorisee:
        return JsonResponse({
            'success': False,
            'error': (
                f'⛔ Vous avez déjà clôturé votre session de la journée '
                f'({session_cloturee_aujourd_hui.numero_session}, clôturée à '
                f'{session_cloturee_aujourd_hui.closed_at.strftime("%H:%M")}). '
                f'Contactez un responsable pour autoriser une réouverture.'
            ),
            'bloquee_cloture': True,
        })

    # 1b. Bloquer si une AUTRE caissière a déjà une session ouverte (shifts non-chevauchants)
    session_autre = CaisseSession.objects.filter(is_open=True).exclude(user=request.user).select_related('user').first()
    if session_autre:
        nom = session_autre.user.get_full_name() or session_autre.user.username
        return JsonResponse({
            'success': False,
            'error': f'⛔ Caisse déjà ouverte par {nom} depuis {session_autre.opened_at.strftime("%H:%M")}. Elle doit clôturer sa session avant que vous puissiez ouvrir.',
        })

    # 2. Bloquer si la caisse centrale du jour précédent n'a pas été clôturée
    session_ancienne = _session_centrale_non_cloturee()
    if session_ancienne:
        return JsonResponse({
            'success': False,
            'error': (
                f'⛔ Ouverture impossible : la caisse centrale du {session_ancienne.date_session.strftime("%d/%m/%Y")} '
                f'n\'a pas été clôturée. La clôture de fin de journée est obligatoire pour '
                f'positionner le solde de veille. Veuillez clôturer cette session avant d\'ouvrir une nouvelle journée.'
            ),
            'session_bloquante_id': session_ancienne.pk,
            'bloquee': True,
        })

    try:
        data  = json.loads(request.body)
        fond  = _dec(data.get('fond_caisse', 0))
        notes = data.get('notes', '')

        session = CaisseSession.objects.create(
            user=request.user,
            type_caisse=type_attendu,
            date_session=today,
            fond_caisse=fond,
            notes=notes,
        )

        # Fond de caisse comme premier mouvement
        if fond > 0:
            MouvementCaisse.objects.create(
                session=session,
                type='fond_caisse',
                module='caisse',
                montant=fond,
                mode_paiement='especes',
                description=f'Fond de caisse — ouverture {session.opened_at.strftime("%d/%m/%Y %H:%M")}',
                cree_par=request.user,
            )

        # ── Consolidation automatique pour la caisse centrale ──────────────
        msg_consolidation = ''
        if type_attendu == 'centrale':
            sessions_autres = CaisseSession.objects.filter(
                opened_at__date=today,
                type_caisse__in=['hotel', 'module']
            ).exclude(id=session.id)

            total_consolide = _dec(0)
            nb_sessions = sessions_autres.count()

            for s in sessions_autres:
                stats_s  = get_stats_jour(today, type_caisse=s.type_caisse)
                montant_s = _dec(stats_s['total'])
                if montant_s > 0:
                    total_consolide += montant_s
                    MouvementCaisse.objects.get_or_create(
                        session=session,
                        type='versement',
                        module='caisse',
                        reference=f'CONSOLIDATION-{s.pk}',
                        defaults={
                            'montant': montant_s,
                            'mode_paiement': 'especes',
                            'description': f'Consolidation auto — {s.get_type_caisse_display()} ({s.user.get_full_name() or s.user.username})',
                            'cree_par': request.user,
                        }
                    )

            if nb_sessions > 0:
                msg_consolidation = f' | {nb_sessions} caisse(s) consolidée(s) : {int(total_consolide):,} F'
        # ──────────────────────────────────────────────────────────────────

        return JsonResponse({
            'success': True,
            'message': f'Caisse ouverte — {session.numero_session} — Fond: {int(fond):,} F{msg_consolidation}',
            'opened_at': session.opened_at.strftime("%d/%m/%Y à %H:%M"),
            'numero_session': session.numero_session,
            'type_caisse': type_attendu,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_module_access('caisse')
@require_POST
def cloturer_caisse(request):
    session = CaisseSession.objects.filter(user=request.user, is_open=True).first()
    if not session:
        return JsonResponse({'success': False, 'error': 'Aucune caisse ouverte'})
    try:
        data      = json.loads(request.body)
        today     = timezone.localdate()
        fond_reel = _dec(data.get('fond_reel', 0))
        prelev    = _dec(data.get('prelevement_banque', 0))
        banque    = data.get('banque', '')
        notes     = data.get('notes', '')

        stats = get_stats_session(session)

        # Solde théorique = fond initial + espèces encaissées pendant le shift − prélèvements banque
        solde_th = session.fond_caisse + _dec(stats['especes']) - prelev
        ecart    = solde_th - fond_reel

        session.closed_at          = timezone.now()
        session.is_open            = False
        session.fond_caisse_reel   = fond_reel
        session.total_especes      = stats['especes']
        session.total_mobile       = stats['mobile']
        session.total_carte        = stats['carte']
        session.total_virement     = stats['virement']
        session.total_general      = stats['total']
        session.prelevement_banque = prelev
        session.solde_theorique    = solde_th
        session.ecart              = ecart
        session.notes              = notes
        session.save()

        # Enregistrer le prélèvement banque si > 0
        if prelev > 0:
            PrelevementBanque.objects.create(
                session=session,
                montant=prelev,
                banque=banque,
                notes=notes,
                cree_par=request.user,
            )
            MouvementCaisse.objects.create(
                session=session,
                type='prelevement',
                module='banque',
                montant=prelev,
                mode_paiement='virement',
                description=f'Prélèvement banque à la clôture — {banque}',
                cree_par=request.user,
            )

        ecart_label = f"+{int(ecart):,} F (excédent)" if ecart > 0 else (f"{int(ecart):,} F (manquant)" if ecart < 0 else "0 F (équilibré)")

        return JsonResponse({
            'success':         True,
            'logout':          True,
            'session_id':      session.pk,
            'message':         f'Caisse clôturée — {session.numero_session}. Total: {int(stats["total"]):,} F | Écart: {ecart_label}',
            'total':           stats['total'],
            'prelev':          int(prelev),
            'solde_theorique': int(solde_th),
            'ecart':           int(ecart),
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_manager
@require_POST
def force_cloturer_caisse(request, session_id):
    """Clôture forcée d'une session bloquante par un manager."""
    session = get_object_or_404(CaisseSession, pk=session_id, is_open=True)
    try:
        data      = json.loads(request.body) if request.body else {}
        fond_reel = _dec(data.get('fond_reel', session.fond_caisse))
        notes     = data.get('notes', 'Clôture forcée par manager')

        date = session.date_session
        stats = get_stats_jour(date, type_caisse=session.type_caisse)

        solde_th = session.fond_caisse + _dec(stats['especes']) - session.prelevement_banque
        ecart    = solde_th - fond_reel

        session.closed_at        = timezone.now()
        session.is_open          = False
        session.fond_caisse_reel = fond_reel
        session.total_especes    = stats['especes']
        session.total_mobile     = stats['mobile']
        session.total_carte      = stats['carte']
        session.total_virement   = stats['virement']
        session.total_general    = stats['total']
        session.solde_theorique  = solde_th
        session.ecart            = ecart
        session.notes            = notes + f' — forcée par {request.user.get_full_name() or request.user.username}'
        session.save()

        return JsonResponse({
            'success': True,
            'message': f'Session {session.numero_session} du {session.date_session.strftime("%d/%m/%Y")} clôturée de force.',
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_module_access('caisse')
@require_POST
def cloture_oubliee(request):
    """La caissière clôture elle-même sa session oubliée d'un jour précédent."""
    today = timezone.localdate()
    session = CaisseSession.objects.filter(
        user=request.user, is_open=True, date_session__lt=today
    ).order_by('-date_session').first()

    if not session:
        return JsonResponse({'success': False, 'error': 'Aucune session oubliée à clôturer.'})

    try:
        data      = json.loads(request.body) if request.body else {}
        fond_reel = _dec(data.get('fond_reel', session.fond_caisse))
        notes     = data.get('notes', '').strip() or f'Clôture différée le {today.strftime("%d/%m/%Y")}'

        stats    = get_stats_session(session)
        solde_th = session.fond_caisse + _dec(stats['especes']) - session.prelevement_banque
        ecart    = solde_th - fond_reel

        session.closed_at        = timezone.now()
        session.is_open          = False
        session.fond_caisse_reel = fond_reel
        session.total_especes    = stats['especes']
        session.total_mobile     = stats['mobile']
        session.total_carte      = stats['carte']
        session.total_virement   = stats['virement']
        session.total_general    = stats['total']
        session.solde_theorique  = solde_th
        session.ecart            = ecart
        session.notes            = notes
        session.save()

        return JsonResponse({
            'success': True,
            'message': f'Session du {session.date_session.strftime("%d/%m/%Y")} clôturée. Vous pouvez maintenant ouvrir la session du jour.',
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_module_access('caisse')
@require_POST
def enregistrer_mouvement(request):
    """Dépense, remboursement, ajustement manuel."""
    try:
        data    = json.loads(request.body)
        session = CaisseSession.objects.filter(user=request.user, is_open=True).first()

        type_mv = data.get('type', 'depense')
        montant = _dec(data.get('montant', 0))
        if montant <= 0:
            return JsonResponse({'success': False, 'error': 'Montant invalide'})

        MouvementCaisse.objects.create(
            session=session,
            type=type_mv,
            module=data.get('module', 'caisse'),
            montant=montant,
            mode_paiement=data.get('mode_paiement', 'especes'),
            description=data.get('description', ''),
            reference=data.get('reference', ''),
            cree_par=request.user,
        )
        return JsonResponse({'success': True, 'message': f'Mouvement enregistré : {int(montant):,} F'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_module_access('caisse')
@require_POST
def prelevement_banque(request):
    """Prélèvement vers la banque en cours de journée."""
    try:
        data    = json.loads(request.body)
        montant = _dec(data.get('montant', 0))
        if montant <= 0:
            return JsonResponse({'success': False, 'error': 'Montant invalide'})
        session = CaisseSession.objects.filter(user=request.user, is_open=True).first()

        PrelevementBanque.objects.create(
            session=session,
            montant=montant,
            banque=data.get('banque', ''),
            reference=data.get('reference', ''),
            notes=data.get('notes', ''),
            cree_par=request.user,
        )
        MouvementCaisse.objects.create(
            session=session,
            type='prelevement',
            module='banque',
            montant=montant,
            mode_paiement='virement',
            description=f'Prélèvement banque — {data.get("banque", "")}',
            reference=data.get('reference', ''),
            cree_par=request.user,
        )
        return JsonResponse({
            'success': True,
            'message': f'Prélèvement de {int(montant):,} F enregistré',
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_module_access('caisse')
@require_POST
def api_modifier_ticket(request):
    """Correction d'un ticket pendant le shift actif : mode de paiement et/ou montant reçu."""
    try:
        data    = json.loads(request.body)
        session = CaisseSession.objects.filter(user=request.user, is_open=True).first()
        if not session:
            return JsonResponse({'success': False, 'error': 'Aucune session ouverte.'})

        ticket_id = data.get('ticket_id')
        ticket = get_object_or_404(Ticket, pk=ticket_id)

        # Vérifier que le ticket appartient à ce shift (fenêtre horaire)
        date_fin = session.closed_at or timezone.now()
        if not (session.opened_at <= ticket.date_creation < date_fin):
            return JsonResponse({'success': False, 'error': 'Ce ticket n\'appartient pas à votre shift.'})

        MODES_VALIDES = [
            'especes', 'mobile_money', 'orange_money', 'wave', 'mtn_money',
            'moov_money', 'carte_bancaire', 'carte', 'virement', 'cheque',
            'chambre', 'mixte', 'autre',
        ]
        nouveau_mode = data.get('mode_paiement')
        nouveau_montant_paye = data.get('montant_paye')

        if nouveau_mode and nouveau_mode not in MODES_VALIDES:
            return JsonResponse({'success': False, 'error': 'Mode de paiement invalide.'})

        modifs = []
        if nouveau_mode and nouveau_mode != ticket.mode_paiement:
            ticket.mode_paiement = nouveau_mode
            modifs.append(f'mode → {nouveau_mode}')
        if nouveau_montant_paye is not None:
            mp = _dec(nouveau_montant_paye)
            if mp >= 0:
                ticket.montant_paye = mp
                modifs.append(f'montant reçu → {int(mp):,} F')

        if not modifs:
            return JsonResponse({'success': False, 'error': 'Aucune modification détectée.'})

        ticket.save()
        return JsonResponse({'success': True, 'message': f'Ticket {ticket.numero} modifié : {", ".join(modifs)}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_module_access('caisse')
@require_POST
def api_modifier_mouvement(request):
    """Modifier un mouvement du journal (session active uniquement)."""
    try:
        data       = json.loads(request.body)
        session    = CaisseSession.objects.filter(user=request.user, is_open=True).first()
        if not session:
            return JsonResponse({'success': False, 'error': 'Aucune session ouverte.'})

        mv = get_object_or_404(MouvementCaisse, pk=data.get('mouvement_id'), session=session)

        nouveau_montant = data.get('montant')
        if nouveau_montant is not None:
            m = _dec(nouveau_montant)
            if m <= 0:
                return JsonResponse({'success': False, 'error': 'Montant invalide.'})
            mv.montant = m

        if data.get('description') is not None:
            mv.description = data['description'][:300]
        if data.get('mode_paiement'):
            mv.mode_paiement = data['mode_paiement']

        mv.save()
        return JsonResponse({'success': True, 'message': 'Mouvement mis à jour.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_module_access('caisse')
@require_POST
def api_supprimer_mouvement(request):
    """Supprimer un mouvement du journal (session active uniquement)."""
    try:
        data    = json.loads(request.body)
        session = CaisseSession.objects.filter(user=request.user, is_open=True).first()
        if not session:
            return JsonResponse({'success': False, 'error': 'Aucune session ouverte.'})

        mv = get_object_or_404(MouvementCaisse, pk=data.get('mouvement_id'), session=session)
        desc = str(mv)
        # Si prélèvement banque, supprimer aussi le PrelevementBanque associé
        if mv.type in ('prelevement', 'prelevement_banque') and mv.reference:
            PrelevementBanque.objects.filter(session=session, reference=mv.reference).delete()
        mv.delete()
        return JsonResponse({'success': True, 'message': f'Mouvement supprimé : {desc}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_module_access('caisse')
def rapport_caisse(request, session_id=None):
    """Rapport imprimable d'une session de caisse."""
    from utils.permissions import _is_manager
    from datetime import datetime
    is_manager = _is_manager(request.user)

    if session_id:
        session = get_object_or_404(CaisseSession, pk=session_id)
    else:
        date_str = request.GET.get('date')
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else timezone.localdate()
        except ValueError:
            target_date = timezone.localdate()

        qs = CaisseSession.objects.filter(opened_at__date=target_date)
        if not is_manager:
            qs = qs.filter(user=request.user)
        session = qs.order_by('-opened_at').first()

        # Fallback : si pas de session à cette date, prendre la session ouverte en cours
        if not session and not date_str:
            open_qs = CaisseSession.objects.filter(is_open=True)
            if not is_manager:
                open_qs = open_qs.filter(user=request.user)
            session = open_qs.order_by('-opened_at').first()

    if not session:
        if is_manager:
            return redirect('caisse:historique')
        return redirect('caisse:index')

    # Stats et réconciliation limités au shift (fenêtre horaire de la session)
    stats          = get_stats_session(session)
    reconciliation = get_reconciliation_session(session)
    mouvements     = MouvementCaisse.objects.filter(session=session, valide=True).order_by('date')
    prelevements   = PrelevementBanque.objects.filter(session=session, valide=True).order_by('date')
    solde_veille, _ = get_solde_veille()

    # Fond théorique = fond ouverture + versements reçus des modules – dépenses – prélèvements banque
    nouveau_fond = int(session.fond_caisse) + reconciliation['grand_total_verse'] - stats['prelevements'] - stats['depenses']

    auto_print = request.GET.get('auto_print', '0')

    return render(request, 'caisse/rapport.html', {
        'session':        session,
        'stats':          stats,
        'mouvements':     mouvements,
        'prelevements':   prelevements,
        'reconciliation': reconciliation,
        'solde_veille':   solde_veille,
        'nouveau_fond':   nouveau_fond,
        'auto_print':     auto_print,
    })


@require_module_access('caisse')
def rapport_caisse_excel(request, session_id=None):
    """Export Excel : rapport de session de caisse."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    if session_id:
        session = get_object_or_404(CaisseSession, pk=session_id)
    else:
        today = timezone.localdate()
        session = CaisseSession.objects.filter(
            opened_at__date=today, user=request.user
        ).order_by('-opened_at').first()

    if not session:
        from django.http import HttpResponseNotFound
        return HttpResponseNotFound("Aucune session trouvée.")

    mouvements   = MouvementCaisse.objects.filter(session=session, valide=True).order_by('date')
    prelevements = PrelevementBanque.objects.filter(session=session, valide=True).order_by('date')

    wb = openpyxl.Workbook()

    hf  = PatternFill("solid", fgColor="1a2535")
    hft = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    tf  = Font(name='Calibri', bold=True, size=14, color='1a2535')
    bf  = Font(name='Calibri', bold=True, size=10)
    nf  = Font(name='Calibri', size=10)
    th  = Side(border_style="thin", color="d4dce8")
    bd  = Border(left=th, right=th, top=th, bottom=th)
    ct  = Alignment(horizontal='center', vertical='center')
    rt  = Alignment(horizontal='right', vertical='center')

    # --- Feuille 1 : Mouvements ---
    ws = wb.active
    ws.title = "Mouvements Caisse"

    NC = 7
    ws.merge_cells(f'A1:{get_column_letter(NC)}1')
    ws['A1'] = f"RAPPORT DE CAISSE — {session.opened_at.strftime('%d/%m/%Y').upper()} — COMPLEXE BEHANIAN"
    ws['A1'].font = tf; ws['A1'].alignment = ct

    ws.merge_cells(f'A2:{get_column_letter(NC)}2')
    caissier = session.user.get_full_name() or session.user.username
    ws['A2'] = f"Caissier : {caissier} — Ouverture : {session.opened_at.strftime('%H:%M')} — Édité le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(name='Calibri', size=10, color='7a8b9c', italic=True)
    ws['A2'].alignment = ct

    ws.append([])
    headers = ['#', 'Date / Heure', 'Type', 'Module', 'Référence', 'Mode paiement', 'Montant (FCFA)']
    ws.append(headers)
    rh = ws.max_row
    for col in range(1, NC + 1):
        c = ws.cell(row=rh, column=col)
        c.fill = hf; c.font = hft; c.alignment = ct; c.border = bd

    enc_fill = PatternFill("solid", fgColor="f0fdf4")
    dep_fill = PatternFill("solid", fgColor="fef2f2")
    for i, m in enumerate(mouvements, 1):
        row_fill = enc_fill if m.type in ('encaissement', 'versement', 'fond_caisse') else dep_fill
        ws.append([
            i,
            m.date.strftime('%d/%m/%Y %H:%M'),
            m.get_type_display(),
            m.get_module_display(),
            m.reference or '—',
            m.get_mode_paiement_display(),
            float(m.montant),
        ])
        rw = ws.max_row
        for col in range(1, NC + 1):
            c = ws.cell(row=rw, column=col)
            c.font = bf if col == 7 else nf
            c.border = bd
            c.fill = row_fill
            c.alignment = rt if col == 7 else (ct if col in (1, 2) else Alignment(vertical='center'))
        ws.cell(row=rw, column=7).number_format = '#,##0'

    ws.append([])
    tr = ws.max_row + 1
    total = sum(float(m.montant) for m in mouvements)
    ws.cell(row=tr, column=6, value='TOTAL').font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=tr, column=7, value=total).font = Font(name='Calibri', bold=True, size=12, color='16a34a')
    ws.cell(row=tr, column=7).number_format = '#,##0'

    for col, w in enumerate([5, 18, 20, 16, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # --- Feuille 2 : Prélèvements ---
    if prelevements.exists():
        ws2 = wb.create_sheet("Prélèvements")
        ws2.merge_cells('A1:E1')
        ws2['A1'] = f"PRÉLÈVEMENTS BANQUE — {session.opened_at.strftime('%d/%m/%Y')}"
        ws2['A1'].font = tf; ws2['A1'].alignment = ct

        ws2.append([])
        h2 = ['#', 'Date / Heure', 'Montant (FCFA)', 'Banque', 'Description']
        ws2.append(h2)
        rh2 = ws2.max_row
        for col in range(1, 6):
            c = ws2.cell(row=rh2, column=col)
            c.fill = hf; c.font = hft; c.alignment = ct; c.border = bd

        for i, p in enumerate(prelevements, 1):
            ws2.append([
                i,
                p.date.strftime('%d/%m/%Y %H:%M'),
                float(p.montant),
                getattr(p, 'banque', '—') or '—',
                getattr(p, 'description', '') or '',
            ])
            rw2 = ws2.max_row
            for col in range(1, 6):
                c = ws2.cell(row=rw2, column=col)
                c.font = bf if col == 3 else nf
                c.border = bd
                c.alignment = rt if col == 3 else (ct if col in (1, 2) else Alignment(vertical='center'))
            ws2.cell(row=rw2, column=3).number_format = '#,##0'

        for col, w in enumerate([5, 18, 16, 16, 30], 1):
            ws2.column_dimensions[get_column_letter(col)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fname = f"Caisse_{session.opened_at.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


@require_module_access('caisse')
def api_stats_jour(request):
    """API stats pour une date donnée."""
    from datetime import date as dt
    date_str = request.GET.get('date')
    try:
        date = dt.fromisoformat(date_str) if date_str else timezone.now().date()
    except ValueError:
        date = timezone.now().date()
    stats = get_stats_jour(date)
    stats.pop('tickets', None)
    return JsonResponse(stats)


@require_module_access('caisse')
@require_POST
def sync_centrale(request):
    """Re-synchronise en temps réel les caisses modules/hotel dans la session centrale."""
    session = CaisseSession.objects.filter(
        user=request.user, is_open=True, type_caisse='centrale'
    ).first()
    if not session:
        return JsonResponse({'success': False, 'error': 'Aucune session centrale ouverte pour cet utilisateur'})

    try:
        today          = timezone.localdate()
        sessions_autres = CaisseSession.objects.filter(
            opened_at__date=today,
            type_caisse__in=['hotel', 'module']
        ).exclude(id=session.id)

        total_consolide = _dec(0)
        nb_updates      = 0

        for s in sessions_autres:
            stats_s   = get_stats_jour(today, type_caisse=s.type_caisse)
            montant_s = _dec(stats_s['total'])
            ref       = f'CONSOLIDATION-{s.pk}'

            existing = MouvementCaisse.objects.filter(session=session, reference=ref).first()
            if existing:
                if montant_s > 0 and existing.montant != montant_s:
                    existing.montant     = montant_s
                    existing.description = (
                        f'Consolidation sync — {s.get_type_caisse_display()} '
                        f'({s.user.get_full_name() or s.user.username})'
                    )
                    existing.save()
                    nb_updates += 1
            elif montant_s > 0:
                MouvementCaisse.objects.create(
                    session=session,
                    type='versement',
                    module='caisse',
                    reference=ref,
                    montant=montant_s,
                    mode_paiement='especes',
                    description=(
                        f'Consolidation — {s.get_type_caisse_display()} '
                        f'({s.user.get_full_name() or s.user.username})'
                    ),
                    cree_par=request.user,
                )
                nb_updates += 1

            total_consolide += montant_s

        return JsonResponse({
            'success': True,
            'message': f'Synchronisation OK — {int(total_consolide):,} F | {nb_updates} mise(s) à jour',
            'total': int(total_consolide),
            'nb_sessions': sessions_autres.count(),
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_module_access('caisse')
@require_POST
def api_autoriser_reouverture(request):
    """Manager : autoriser la réouverture d'une session clôturée pour un utilisateur."""
    from utils.permissions import _is_manager as _chk_manager
    if not (_chk_manager(request.user) or request.user.is_superuser):
        return JsonResponse({'success': False, 'error': 'Accès réservé aux responsables.'}, status=403)
    try:
        data = json.loads(request.body)
        session_id = data.get('session_id')
        session = CaisseSession.objects.get(pk=session_id, is_open=False)
        session.reouverture_autorisee = True
        session.reouverture_par = request.user
        session.save(update_fields=['reouverture_autorisee', 'reouverture_par'])
        nom = session.user.get_full_name() or session.user.username
        return JsonResponse({
            'success': True,
            'message': f'Réouverture autorisée pour {nom}. Elle peut maintenant ouvrir une nouvelle session.',
        })
    except CaisseSession.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Session introuvable.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_module_access('caisse')
def api_reconciliation(request):
    """API JSON — état des transactions vs versements par module pour le jour.
    Accepte ?session_id=X pour les managers (cohérence avec le sélecteur de session).
    """
    from utils.permissions import _is_manager as _chk_manager
    today = timezone.localdate()
    session_active = CaisseSession.objects.filter(user=request.user, is_open=True).first()
    if session_active:
        data = get_reconciliation_session(session_active)
    else:
        is_manager = _chk_manager(request.user) or request.user.is_superuser
        session_id = request.GET.get('session_id')
        if is_manager and session_id:
            try:
                session_filtre = CaisseSession.objects.get(pk=session_id, opened_at__date=today)
                data = get_reconciliation_session(session_filtre)
            except CaisseSession.DoesNotExist:
                data = get_reconciliation_jour(today)
        else:
            data = get_reconciliation_jour(today)

    # Inclure le nombre de sessions du jour pour détecter de nouvelles sessions côté JS
    nb_sessions = CaisseSession.objects.filter(opened_at__date=today).count()
    return JsonResponse({'success': True, 'reconciliation': data, 'nb_sessions': nb_sessions})


@require_module_access('caisse')
def rapport_transactions(request):
    """Rapport des transactions par module, mode de paiement et opérateur sur une période."""
    from datetime import datetime
    from collections import defaultdict

    today = timezone.localdate()

    def _parse_date(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return today

    date_debut     = _parse_date(request.GET.get('date_debut'))
    date_fin       = _parse_date(request.GET.get('date_fin')) if request.GET.get('date_fin') else date_debut
    filtre_module  = request.GET.get('module', '').strip()
    filtre_mode    = request.GET.get('mode', '').strip()

    if date_fin < date_debut:
        date_fin = date_debut

    MODE_LABELS = {
        'especes':        ('Espèces',            '#16a34a', '💵'),
        'wave':           ('Wave',               '#1d4ed8', '🔵'),
        'orange_money':   ('Orange Money',       '#c2410c', '🟠'),
        'mtn_money':      ('MTN Mobile Money',   '#854d0e', '🟡'),
        'moov_money':     ('Moov Money',         '#0f766e', '🟢'),
        'mobile_money':   ('Mobile Money',       '#7c3aed', '📱'),
        'mobile':         ('Mobile Money',       '#7c3aed', '📱'),
        'carte_bancaire': ('Carte Bancaire',     '#0891b2', '💳'),
        'carte':          ('Carte Bancaire',     '#0891b2', '💳'),
        'cheque':         ('Chèque',             '#475569', '🏷️'),
        'virement':       ('Virement',           '#d97706', '🏦'),
        'chambre':        ('Sur chambre',        '#6d28d9', '🛏️'),
        'autre':          ('Autre',              '#64748b', '—'),
    }

    MODULE_ORDER = ['hotel', 'restaurant', 'cave', 'piscine', 'espace', 'caisse']
    MODULE_LABELS = {
        'hotel':      ('Hôtel',                  '🏨'),
        'restaurant': ('Restaurant',             '🍽️'),
        'cave':       ('Cave / Bar',             '🍷'),
        'piscine':    ('Piscine',                '🏊'),
        'espace':     ('Espaces Événementiels',  '🎪'),
        'caisse':     ('Caisse',                 '🏦'),
    }

    # ── Queryset ────────────────────────────────────────────────────────────
    tickets_qs = Ticket.objects.filter(
        date_creation__date__gte=date_debut,
        date_creation__date__lte=date_fin,
    ).select_related('cree_par')

    if filtre_module:
        tickets_qs = tickets_qs.filter(module__startswith=filtre_module)

    MOBILE_SLUGS = ['mobile', 'mobile_money', 'wave', 'orange_money', 'mtn_money', 'moov_money']
    if filtre_mode:
        if filtre_mode == 'mobile':
            tickets_qs = tickets_qs.filter(mode_paiement__in=MOBILE_SLUGS)
        else:
            tickets_qs = tickets_qs.filter(mode_paiement=filtre_mode)

    # ── Agrégats module × mode ───────────────────────────────────────────────
    agg = (
        tickets_qs
        .values('module', 'mode_paiement')
        .annotate(total=Sum('montant_total'), nb=Count('id'))
        .order_by('module', 'mode_paiement')
    )

    par_module_raw = defaultdict(lambda: {'modes': {}, 'total': 0, 'nb': 0})
    par_mode_raw   = {}
    grand_total    = 0
    grand_nb       = 0

    for row in agg:
        mod_key  = (row['module'] or 'autre')
        mode_raw = (row['mode_paiement'] or 'especes')
        mode_info = MODE_LABELS.get(mode_raw, (mode_raw.replace('_', ' ').capitalize(), '#64748b', '—'))
        mode_lbl  = mode_info[0]
        amount    = int(row['total'] or 0)
        nb        = int(row['nb'] or 0)

        par_module_raw[mod_key]['modes'][mode_lbl] = par_module_raw[mod_key]['modes'].get(mode_lbl, 0) + amount
        par_module_raw[mod_key]['total'] += amount
        par_module_raw[mod_key]['nb']    += nb
        par_mode_raw[mode_lbl]            = par_mode_raw.get(mode_lbl, 0) + amount
        grand_total += amount
        grand_nb    += nb

    # Trier les modules dans l'ordre défini
    def _mod_sort(key):
        for i, prefix in enumerate(MODULE_ORDER):
            if key.startswith(prefix):
                return i
        return 99

    modules_list = []
    for mod_key in sorted(par_module_raw.keys(), key=_mod_sort):
        data     = par_module_raw[mod_key]
        mod_info = MODULE_LABELS.get(mod_key, (mod_key.capitalize(), '—'))
        modules_list.append({
            'key':   mod_key,
            'label': mod_info[0],
            'emoji': mod_info[1],
            'modes': sorted(data['modes'].items(), key=lambda x: -x[1]),
            'total': data['total'],
            'nb':    data['nb'],
        })

    global_modes = sorted(par_mode_raw.items(), key=lambda x: -x[1])

    # ── Détail tickets (max 1000) ────────────────────────────────────────────
    tickets_list = []
    for tk in tickets_qs.order_by('-date_creation')[:1000]:
        mod_key  = tk.module or 'autre'
        mode_raw = tk.mode_paiement or 'especes'
        mode_info = MODE_LABELS.get(mode_raw, (mode_raw.replace('_', ' ').capitalize(), '#64748b', '—'))
        mod_info  = MODULE_LABELS.get(mod_key, (mod_key.capitalize(), '—'))
        tickets_list.append({
            'numero':   tk.numero,
            'date':     tk.date_creation.strftime('%d/%m/%Y'),
            'heure':    tk.date_creation.strftime('%H:%M'),
            'module':   mod_info[0],
            'mod_emoji':mod_info[1],
            'mode':     mode_info[0],
            'mode_clr': mode_info[1],
            'mode_ico': mode_info[2],
            'montant':  int(tk.montant_total or 0),
            'caissier': (tk.cree_par.get_full_name() or tk.cree_par.username) if tk.cree_par else '—',
        })

    # Options de filtres pour le formulaire
    all_modules = [(k, v[0]) for k, v in MODULE_LABELS.items()]
    all_modes   = [(k, v[0]) for k, v in MODE_LABELS.items() if k not in ('mobile', 'carte')]

    periode = (
        f"{date_debut.strftime('%d/%m/%Y')} → {date_fin.strftime('%d/%m/%Y')}"
        if date_debut != date_fin else date_debut.strftime('%d/%m/%Y')
    )

    return render(request, 'caisse/rapport_transactions.html', {
        'date_debut':    date_debut,
        'date_fin':      date_fin,
        'filtre_module': filtre_module,
        'filtre_mode':   filtre_mode,
        'modules_list':  modules_list,
        'global_modes':  global_modes,
        'grand_total':   grand_total,
        'grand_nb':      grand_nb,
        'tickets_list':  tickets_list,
        'all_modules':   all_modules,
        'all_modes':     all_modes,
        'periode':       periode,
    })


@require_module_access('caisse')
def etat_journee(request):
    """État de fin de journée — point général des transactions par module et par mode de paiement."""
    from datetime import date as date_type
    date_str = request.GET.get('date')
    try:
        from datetime import datetime
        date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else timezone.localdate()
    except ValueError:
        date = timezone.localdate()

    # Labels pour les modes de paiement
    MODE_LABELS = {
        'especes':        'Espèces',
        'wave':           'Wave',
        'orange_money':   'Orange Money',
        'mtn_money':      'MTN Mobile Money',
        'moov_money':     'Moov Money',
        'mobile_money':   'Mobile Money',
        'mobile':         'Mobile Money',
        'carte_bancaire': 'Carte Bancaire',
        'carte':          'Carte Bancaire',
        'cheque':         'Chèque',
        'virement':       'Virement',
        'chambre':        'Sur chambre',
        'autre':          'Autre',
    }

    MODULES = [
        ('hotel',      'Hôtel'),
        ('restaurant', 'Restaurant'),
        ('cave',       'Cave / Bar'),
        ('piscine',    'Piscine'),
        ('espace',     'Espaces Événementiels'),
    ]

    tickets_jour = (
        Ticket.objects
        .filter(date_creation__date=date)
        .values('module', 'mode_paiement')
        .annotate(total=Sum('montant_total'))
        .order_by('module', 'mode_paiement')
    )

    # Construire {module: {mode_label: montant}}
    raw = {}
    for row in tickets_jour:
        mod  = row['module'] or 'autre'
        mode = row['mode_paiement'] or 'especes'
        lbl  = MODE_LABELS.get(mode, mode.replace('_', ' ').capitalize())
        raw.setdefault(mod, {})
        raw[mod][lbl] = raw[mod].get(lbl, 0) + int(row['total'] or 0)

    # Fusionner 'carte' + 'carte_bancaire' → 'Carte Bancaire'
    for mod_data in raw.values():
        # déjà fait via MODE_LABELS (les deux mappent au même label)
        pass

    # Construire la liste ordonnée pour le template
    modules_data = []
    grand_total = 0
    global_modes = {}

    for mod_key, mod_label in MODULES:
        modes = raw.get(mod_key, {})
        total_mod = sum(modes.values())
        if total_mod == 0:
            continue
        grand_total += total_mod
        for lbl, mt in modes.items():
            global_modes[lbl] = global_modes.get(lbl, 0) + mt
        modules_data.append({
            'label':      mod_label,
            'modes':      sorted(modes.items(), key=lambda x: -x[1]),
            'modes_dict': modes,
            'total':      total_mod,
        })

    # Trier les modes globaux par montant décroissant
    sorted_modes = sorted(global_modes.items(), key=lambda x: -x[1])
    mode_labels = [lbl for lbl, _ in sorted_modes]

    # Construire la matrice récapitulatif : une ligne par module avec valeur par mode
    recap_rows = []
    for mod in modules_data:
        vals = [mod['modes_dict'].get(lbl, 0) for lbl in mode_labels]
        recap_rows.append({
            'label': mod['label'],
            'vals':  vals,
            'total': mod['total'],
        })
    recap_mode_totals = [global_modes.get(lbl, 0) for lbl in mode_labels]

    # Shifts du jour : chaque session avec ses stats individuelles
    sessions_jour = CaisseSession.objects.filter(
        opened_at__date=date
    ).select_related('user').order_by('opened_at')
    # Inclure aussi les sessions ouvertes un jour antérieur mais encore actives ce jour
    sessions_veille_actives = CaisseSession.objects.filter(
        is_open=True, opened_at__date__lt=date
    ).select_related('user')
    toutes_sessions = list(sessions_veille_actives) + list(sessions_jour)

    sessions_avec_stats = []
    for s in toutes_sessions:
        s_stats = get_stats_session(s)
        if s_stats['total'] > 0 or s.is_open:
            sessions_avec_stats.append({
                'session': s,
                'stats':   s_stats,
            })

    return render(request, 'caisse/etat_journee.html', {
        'date':                date,
        'modules_data':        modules_data,
        'global_modes':        sorted_modes,
        'grand_total':         grand_total,
        'recap_rows':          recap_rows,
        'recap_mode_labels':   mode_labels,
        'recap_mode_totals':   recap_mode_totals,
        'sessions_avec_stats': sessions_avec_stats,
    })


@require_manager
def historique(request):
    """Historique complet des sessions — Manager."""
    from datetime import datetime, timedelta
    today = timezone.localdate()

    def _parse(p, fb):
        try:
            return datetime.strptime(request.GET[p], '%Y-%m-%d').date()
        except Exception:
            return fb

    date_debut = _parse('date_debut', today - timedelta(days=90))
    date_fin   = _parse('date_fin', today)
    if date_fin < date_debut:
        date_fin = date_debut

    sessions = CaisseSession.objects.filter(
        opened_at__date__gte=date_debut,
        opened_at__date__lte=date_fin,
    ).select_related('user').order_by('-opened_at')

    return render(request, 'caisse/historique.html', {
        'sessions':    sessions,
        'today':       today,
        'date_debut':  date_debut,
        'date_fin':    date_fin,
    })
